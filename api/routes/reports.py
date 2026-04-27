"""
Reports API endpoints for EDS Universal Requisition.

Provides spending analytics, vendor breakdowns, category analysis,
and budget reports. Requires admin/approver/reports-viewer role.

Access levels:
- ApprovalLevel >= 1 (admin/approver/manager): full read/write
- reports_viewer role: read-only access to reports data

Schema notes (actual EDS production schema):
- Requisitions: DateEntered (not CreatedAt), TotalRequisitionCost (not TotalAmount),
  StatusId (int FK → StatusTable), SchoolId (→ School → District, no direct DistrictId)
- Detail: line items with BidPrice, Quantity, VendorId (direct), no ExtendedPrice
- StatusTable: StatusId=1 On Hold, 2 Pending Approval, 3 Approved, 4 Rejected,
  5 At EDS, 6 PO Printed
- Vendors: VendorId, Code, Name
- Category: CategoryId, Name, Description
"""

from fastapi import APIRouter, HTTPException, Query, Header
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime, timedelta
from decimal import Decimal
import csv
import io
import logging

from api.database import execute_query, execute_single
from api.cache import get_cache, CACHE_TTL_MEDIUM
from api.middleware import SESSION_TIMEOUT_HOURS

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/reports", tags=["Reports"])

# Status IDs from StatusTable
STATUS_PENDING = 2     # Pending Approval
STATUS_APPROVED = 3    # Approved
STATUS_REJECTED = 4    # Rejected
STATUS_AT_EDS = 5      # At EDS
STATUS_PO_PRINTED = 6  # PO Printed
# Exclude from reports: On Hold (1) acts like draft
STATUS_EXCLUDE = (1,)  # On Hold


# ===========================================
# CACHE HELPERS
# ===========================================

REPORTS_CACHE_TTL = CACHE_TTL_MEDIUM  # 30 minutes — reports data doesn't change fast

_cache = get_cache()


def _cache_key(district_id: int, date_start: str, date_end: str, section: str) -> str:
    """Build a cache key for a reports query section."""
    return f"reports:{district_id}:{date_start}:{date_end}:{section}"


def _get_cached_or_query(district_id: int, date_start: str, date_end: str,
                         section: str, query_fn):
    """Check cache first, run query on miss, store result."""
    key = _cache_key(district_id, date_start, date_end, section)
    cached = _cache.get_sync(key)
    if cached is not None:
        logger.debug("Reports cache hit: %s for district %s", section, district_id)
        return cached
    result = query_fn(district_id, date_start, date_end)
    _cache.set_sync(key, result, REPORTS_CACHE_TTL)
    return result


def invalidate_reports_cache(district_id: int = None):
    """Clear reports cache when data changes (requisition submitted/approved/etc).

    If district_id is provided, only clears that district's cache.
    Otherwise clears all reports cache entries.
    """
    cache = get_cache()
    if district_id:
        # Clear all sections for this district
        keys_to_delete = [k for k in cache._cache if k.startswith(f"reports:{district_id}:")]
        for key in keys_to_delete:
            cache._cache.pop(key, None)
        logger.debug("Invalidated %d reports cache entries for district %s", len(keys_to_delete), district_id)
    else:
        keys_to_delete = [k for k in cache._cache if k.startswith("reports:")]
        for key in keys_to_delete:
            cache._cache.pop(key, None)
        logger.debug("Invalidated all %d reports cache entries", len(keys_to_delete))


# ===========================================
# AUTH HELPERS
# ===========================================

# Roles that can view reports without ApprovalLevel
REPORTS_VIEWER_ROLES = {"reports_viewer", "department_head", "principal"}


def _get_session_user(session_id: str, demo_approval_level: int = 1) -> dict:
    """Validate session and return user info with role details."""
    if not session_id or session_id == "demo":
        # Demo mode: approval level can be overridden via X-Demo-Approval-Level header
        user_type = "admin" if demo_approval_level >= 1 else "teacher"
        return {
            "UserId": 0, "DistrictId": 0, "SchoolId": 0,
            "ApprovalLevel": demo_approval_level, "FirstName": "Demo", "LastName": "User",
            "UserType": user_type, "DepartmentId": None,
        }
    try:
        sid = int(session_id)
    except (ValueError, TypeError):
        raise HTTPException(status_code=401, detail="Invalid session")

    user = execute_single("""
        SELECT
            s.UserId, s.DistrictId, s.SchoolId, s.ApprovalLevel,
            u.FirstName, u.LastName, u.Email, u.UserType,
            u.DepartmentId
        FROM SessionTable s
        LEFT JOIN Users u ON s.UserId = u.UserId
        WHERE s.SessionId = ?
        AND s.SessionEnd IS NULL
        AND DATEDIFF(HOUR, s.SessionStart, GETDATE()) < ?
    """, (sid, SESSION_TIMEOUT_HOURS))
    if not user:
        raise HTTPException(status_code=401, detail="Invalid or expired session")
    return user


def require_reports_access(user: dict) -> dict:
    """
    Verify user has reports access.

    Access granted if ANY of:
    - ApprovalLevel >= 1 (admin/approver/manager)
    - UserType is in REPORTS_VIEWER_ROLES (read-only)

    Raises 403 if neither condition is met.
    """
    approval_level = user.get("ApprovalLevel", 0) or 0
    user_type = (user.get("UserType") or "").lower().strip()

    if approval_level >= 1:
        return user  # Full admin access

    if user_type in REPORTS_VIEWER_ROLES:
        return user  # Read-only reports access

    raise HTTPException(
        status_code=403,
        detail="Reports access requires admin, approver, or reports viewer role"
    )


# ===========================================
# REPORTS SUMMARY ENDPOINT
# ===========================================

@router.get("/summary")
async def get_reports_summary(
    x_session_id: Optional[str] = Header(None, alias="X-Session-ID"),
    x_demo_approval: Optional[int] = Header(None, alias="X-Demo-Approval-Level"),
    period: str = Query("current", description="Report period: current, previous, ytd, custom"),
    date_start: Optional[str] = Query(None, description="Custom start date (YYYY-MM-DD), required if period=custom"),
    date_end: Optional[str] = Query(None, description="Custom end date (YYYY-MM-DD), required if period=custom"),
):
    """
    Get comprehensive reports summary data.

    Returns spending totals, vendor breakdown, category analysis,
    monthly trend, recent orders, and school-level budget estimates.

    Requires admin/approver/reports_viewer access.
    """
    demo_level = x_demo_approval if x_demo_approval is not None else 1
    user = _get_session_user(x_session_id, demo_approval_level=demo_level)
    require_reports_access(user)

    district_id = user.get("DistrictId", 0)

    if period == "custom" and date_start and date_end:
        # Validate custom dates
        try:
            datetime.strptime(date_start, "%Y-%m-%d")
            datetime.strptime(date_end, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")
        if date_start > date_end:
            raise HTTPException(status_code=400, detail="date_start must be before date_end")
    else:
        date_start, date_end = _compute_date_range(period)

    try:
        summary = _get_cached_or_query(district_id, date_start, date_end, "summary", _get_spending_summary)
        vendor_spend = _get_cached_or_query(district_id, date_start, date_end, "vendors", _get_vendor_spend)
        category_spend = _get_cached_or_query(district_id, date_start, date_end, "categories", _get_category_spend)
        monthly_trend = _get_cached_or_query(district_id, date_start, date_end, "monthly", _get_monthly_trend)
        recent_orders = _get_recent_orders(district_id, date_start, date_end)  # Not cached — always fresh
        budget_departments = _get_cached_or_query(district_id, date_start, date_end, "schools", _get_school_budgets)

        # Compute vs_last_period comparison
        summary = _add_comparison_metrics(summary, district_id, period)

        return {
            "summary": summary,
            "vendor_spend": vendor_spend,
            "category_spend": category_spend,
            "monthly_trend": monthly_trend,
            "recent_orders": recent_orders,
            "budget_departments": budget_departments,
            "period": period,
            "date_range": {"start": date_start, "end": date_end},
        }
    except Exception as e:
        logger.error("Reports query error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to load reports data")


def _add_comparison_metrics(summary: dict, district_id: int, period: str) -> dict:
    """Add vs_last_period and orders_vs_last by comparing to the prior period.

    For 'current' → compare against 'previous' year
    For 'ytd' → compare against same YTD window last year
    For 'previous' → compare against the year before that
    """
    try:
        now = datetime.now()
        if period == "previous":
            if now.month >= 12:
                cmp_start_year = now.year - 2
            else:
                cmp_start_year = now.year - 3
        elif period == "ytd":
            if now.month >= 12:
                cmp_start_year = now.year - 1
            else:
                cmp_start_year = now.year - 2
        else:  # current
            if now.month >= 12:
                cmp_start_year = now.year - 1
            else:
                cmp_start_year = now.year - 2

        cmp_start = f"{cmp_start_year}-12-01"
        if period == "ytd":
            # Same day range but one year earlier
            cmp_end = f"{cmp_start_year + 1}-{now.strftime('%m-%d')}"
        else:
            cmp_end = f"{cmp_start_year + 1}-11-30"

        prev = _get_cached_or_query(district_id, cmp_start, cmp_end, "summary", _get_spending_summary)
        prev_spend = prev.get("total_spend", 0) or 0
        prev_orders = prev.get("total_orders", 0) or 0
        cur_spend = summary.get("total_spend", 0) or 0
        cur_orders = summary.get("total_orders", 0) or 0

        if prev_spend > 0:
            summary["vs_last_period"] = round((cur_spend - prev_spend) / prev_spend * 100, 1)
        else:
            summary["vs_last_period"] = None

        if prev_orders > 0:
            summary["orders_vs_last"] = round((cur_orders - prev_orders) / prev_orders * 100, 1)
        else:
            summary["orders_vs_last"] = None
    except Exception as e:
        logger.debug("Comparison metrics failed: %s", e)
        summary.setdefault("vs_last_period", None)
        summary.setdefault("orders_vs_last", None)

    return summary


# ===========================================
# CSV EXPORT ENDPOINT
# ===========================================

def _compute_date_range(period: str):
    """Compute date_start, date_end for a given period string."""
    now = datetime.now()
    if period == "previous":
        start_year = now.year - 1 if now.month >= 12 else now.year - 2
        return f"{start_year}-12-01", f"{start_year + 1}-11-30"
    elif period == "ytd":
        start_year = now.year if now.month >= 12 else now.year - 1
        return f"{start_year}-12-01", now.strftime("%Y-%m-%d")
    else:
        start_year = now.year if now.month >= 12 else now.year - 1
        return f"{start_year}-12-01", f"{start_year + 1}-11-30"


@router.get("/export")
async def export_reports_csv(
    x_session_id: Optional[str] = Header(None, alias="X-Session-ID"),
    x_demo_approval: Optional[int] = Header(None, alias="X-Demo-Approval-Level"),
    period: str = Query("current", description="Report period: current, previous, ytd"),
    section: str = Query("all", description="Section to export: all, vendors, categories, orders, monthly, schools"),
):
    """
    Export reports data as CSV for Excel analysis.

    Sections:
    - all: Multiple sheets combined (summary + vendors + categories + orders)
    - vendors: Top vendor spending breakdown
    - categories: Category spending breakdown
    - orders: Recent requisitions list
    - monthly: Monthly spending trend
    - schools: School-level spending

    Requires admin/approver/reports_viewer access.
    """
    demo_level = x_demo_approval if x_demo_approval is not None else 1
    user = _get_session_user(x_session_id, demo_approval_level=demo_level)
    require_reports_access(user)

    district_id = user.get("DistrictId", 0)

    date_start, date_end = _compute_date_range(period)

    # Get data (from cache if available)
    data = {
        "summary": _get_cached_or_query(district_id, date_start, date_end, "summary", _get_spending_summary),
        "vendor_spend": _get_cached_or_query(district_id, date_start, date_end, "vendors", _get_vendor_spend),
        "category_spend": _get_cached_or_query(district_id, date_start, date_end, "categories", _get_category_spend),
        "monthly_trend": _get_cached_or_query(district_id, date_start, date_end, "monthly", _get_monthly_trend),
        "recent_orders": _get_recent_orders(district_id, date_start, date_end),
        "budget_departments": _get_cached_or_query(district_id, date_start, date_end, "schools", _get_school_budgets),
    }

    output = io.StringIO()
    writer = csv.writer(output)

    period_label = f"{date_start} to {date_end}"

    if section in ("all", "vendors"):
        if section == "all":
            writer.writerow(["=== VENDOR SPENDING ===", f"Period: {period_label}"])
        writer.writerow(["Rank", "Vendor", "Code", "Orders", "Spend", "% of Total"])
        for i, v in enumerate(data.get("vendor_spend", []), 1):
            writer.writerow([i, v["name"], v["code"], v["orders"], f'{v["spend"]:.2f}', f'{v["pct"]:.1f}%'])
        writer.writerow([])

    if section in ("all", "categories"):
        if section == "all":
            writer.writerow(["=== CATEGORY SPENDING ==="])
        writer.writerow(["Category", "Spend", "% of Total"])
        for c in data.get("category_spend", []):
            writer.writerow([c["name"], f'{c["spend"]:.2f}', f'{c["pct"]:.1f}%'])
        writer.writerow([])

    if section in ("all", "monthly"):
        if section == "all":
            writer.writerow(["=== MONTHLY TREND ==="])
        writer.writerow(["Month", "Amount"])
        for m in data.get("monthly_trend", []):
            writer.writerow([m["month"], f'{m["amount"]:.2f}'])
        writer.writerow([])

    if section in ("all", "orders"):
        if section == "all":
            writer.writerow(["=== RECENT ORDERS ==="])
        writer.writerow(["Order #", "Date", "Requester", "School/Dept", "Items", "Amount", "Status"])
        for o in data.get("recent_orders", []):
            writer.writerow([o["id"], o["date"], o["requester"], o["dept"],
                             o["items"], f'{o["amount"]:.2f}', o["status"]])
        writer.writerow([])

    if section in ("all", "schools"):
        if section == "all":
            writer.writerow(["=== SCHOOL SPENDING ==="])
        writer.writerow(["School", "Budget (Est.)", "Spent", "% Used"])
        for s in data.get("budget_departments", []):
            writer.writerow([s["name"], f'{s["budget"]:.2f}', f'{s["spent"]:.2f}', f'{s["pct"]:.1f}%'])

    output.seek(0)
    filename = f"eds-reports-{section}-{period}-{datetime.now().strftime('%Y%m%d')}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ===========================================
# PDF EXPORT ENDPOINT
# ===========================================

@router.get("/export/pdf")
async def export_reports_pdf(
    x_session_id: Optional[str] = Header(None, alias="X-Session-ID"),
    x_demo_approval: Optional[int] = Header(None, alias="X-Demo-Approval-Level"),
    period: str = Query("current", description="Report period: current, previous, ytd"),
):
    """
    Export a formatted PDF report with spending summary, vendor breakdown,
    category analysis, monthly trend, and recent orders.

    Requires admin/approver/reports_viewer access.
    """
    demo_level = x_demo_approval if x_demo_approval is not None else 1
    user = _get_session_user(x_session_id, demo_approval_level=demo_level)
    require_reports_access(user)

    district_id = user.get("DistrictId", 0)

    date_start, date_end = _compute_date_range(period)

    data = {
        "summary": _get_cached_or_query(district_id, date_start, date_end, "summary", _get_spending_summary),
        "vendor_spend": _get_cached_or_query(district_id, date_start, date_end, "vendors", _get_vendor_spend),
        "category_spend": _get_cached_or_query(district_id, date_start, date_end, "categories", _get_category_spend),
        "monthly_trend": _get_cached_or_query(district_id, date_start, date_end, "monthly", _get_monthly_trend),
        "recent_orders": _get_recent_orders(district_id, date_start, date_end),
        "budget_departments": _get_cached_or_query(district_id, date_start, date_end, "schools", _get_school_budgets),
    }

    pdf_bytes = _build_pdf_report(data, date_start, date_end, period)
    filename = f"eds-reports-{period}-{datetime.now().strftime('%Y%m%d')}.pdf"

    return StreamingResponse(
        iter([pdf_bytes]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ===========================================
# DISTRICT ITEMS XLSX EXPORT
# ===========================================
#
# One row per line item for a single district and date range, in an Excel
# workbook with freeze panes + auto-filter so the district can sort/filter
# by User #, Item #, Vendor, Price, etc. in Excel.
#
# Requested by Patrice Abate (EDS) for Colton-Pierrepont (7G): "I want to
# identify an item (say a Pilot pen) and then sort to see who is ordering
# that at what price."
#
# Default date range is the most recently completed EDS budget year
# (Dec 1 - Nov 30), computed from the current date.


def _default_budget_year_range() -> tuple[str, str, str]:
    """Return (date_start, date_end_exclusive, label) for the most recently
    completed EDS budget year, which runs Dec 1 through Nov 30.

    If today is on or after Dec 1, the most recently completed year started
    the *prior* Dec 1. Otherwise it started two Decembers ago.
    """
    now = datetime.now()
    if now.month >= 12:
        start_year = now.year - 1
    else:
        start_year = now.year - 2
    # Inclusive start, exclusive end (one day past Nov 30)
    return (
        f"{start_year}-12-01",
        f"{start_year + 1}-12-01",
        f"Dec {start_year} - Nov {start_year + 1}",
    )


DISTRICT_ITEMS_SQL = """
    SELECT
        r.RequisitionNumber                                    AS Requisition,
        CAST(r.DateEntered AS date)                            AS [Order Date],
        sch.Name                                               AS School,
        u.CometId                                              AS [User #],
        LTRIM(RTRIM(
            ISNULL(u.FirstName,'') + ' ' + ISNULL(u.LastName,'')
        ))                                                     AS [User Name],
        v.Code                                                 AS [Vendor Code],
        COALESCE(v.Name, 'N/A')                                AS Vendor,
        COALESCE(det.VendorItemCode, i.ItemCode, '')           AS [Item #],
        COALESCE(det.ItemCode, i.ItemCode, '')                 AS [EDS Item Code],
        COALESCE(det.Description, i.Description, '')           AS [Item Name],
        CAST(det.Quantity AS INT)                              AS Qty,
        COALESCE(det.BidPrice, det.CatalogPrice)               AS [Unit Price],
        CAST(det.Quantity AS money)
            * COALESCE(det.BidPrice, det.CatalogPrice)         AS [Line Total],
        ISNULL(st.Name, 'On Hold')                             AS Status
    FROM dbo.Requisitions r
    INNER JOIN dbo.Detail      det ON det.RequisitionId = r.RequisitionId
    INNER JOIN dbo.School      sch ON sch.SchoolId      = r.SchoolId
    INNER JOIN dbo.District    d   ON d.DistrictId      = sch.DistrictId
    LEFT  JOIN dbo.Users       u   ON u.UserId          = r.UserId
    LEFT  JOIN dbo.Vendors     v   ON v.VendorId        = det.VendorId
    LEFT  JOIN dbo.Items       i   ON i.ItemId          = det.ItemId
    LEFT  JOIN dbo.StatusTable st  ON st.StatusId       = r.StatusId
    WHERE d.DistrictCode = ?
      AND r.Active   = 1
      AND det.Active = 1
      AND r.DateEntered >= ?
      AND r.DateEntered <  ?
      AND r.StatusId IS NOT NULL
      AND r.StatusId NOT IN (1, 2, 4)  -- On Hold, Pending Approval, Rejected
    ORDER BY
        COALESCE(det.VendorItemCode, i.ItemCode, ''),
        v.Name,
        r.DateEntered,
        u.LastName,
        u.FirstName
"""

# 1-based column indexes used for Excel formatting
_XLSX_DATE_COLS     = {2}        # Order Date
_XLSX_ID_COLS       = {4}        # User # -- plain integer id, no thousands sep
_XLSX_INT_COLS      = {11}       # Qty -- thousands-separated integer
_XLSX_CURRENCY_COLS = {12, 13}   # Unit Price, Line Total

# EDS brand styling (matches scripts/orleans_top_items_report.py)
_EDS_BLUE = "1C1A83"
_EDS_LIGHT = "4A4890"
_EDS_RED = "B70C0D"
_XLSX_ALT_ROW = "EBEBF5"
_XLSX_MONEY_FMT = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'
_XLSX_QTY_FMT = "#,##0"          # quantities with thousands separator
_XLSX_ID_FMT = "0"               # identifiers as plain integers
_XLSX_DATE_FMT = "MM/DD/YYYY"

_XLSX_COL_WIDTHS = {
    "Requisition":     14,
    "Order Date":      12,
    "School":          28,
    "User #":          10,
    "User Name":       24,
    "Vendor Code":     12,
    "Vendor":          30,
    "Item #":          18,
    "EDS Item Code":   14,
    "Item Name":       48,
    "Qty":              8,
    "Unit Price":      13,
    "Line Total":      14,
    "Status":          18,
}


def _build_district_items_xlsx(
    rows: list[dict],
    col_names: list[str],
    district_code: str,
    district_name: str,
    date_start: str,
    date_end_exclusive: str,
    period_label: str,
) -> bytes:
    """Render the district items result set into an Excel workbook.

    openpyxl is imported lazily so the API still starts on deployments that
    don't have the [export] extra installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail=(
                "openpyxl is not installed on the API server. "
                "Install with: pip install openpyxl"
            ),
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Items Ordered"

    num_cols = len(col_names)
    last_col_letter = get_column_letter(num_cols)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    hdr_fill = PatternFill(start_color=_EDS_BLUE, end_color=_EDS_BLUE, fill_type="solid")
    data_font = Font(size=10, name="Calibri")
    alt_fill = PatternFill(start_color=_XLSX_ALT_ROW, end_color=_XLSX_ALT_ROW, fill_type="solid")

    # --- Title ---
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"{district_name} ({district_code}) -- Items Ordered -- {period_label}"
    ws["A1"].font = Font(bold=True, size=14, color=_EDS_BLUE, name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # --- Subtitle (total + timestamp) ---
    # Sum line totals as Decimal so the displayed total matches cents
    # exactly -- no float rounding drift. Wrap with Decimal(str(...)) so
    # the helper also accepts fixture values that are plain floats in tests.
    line_total_key = "Line Total"
    total_spend = Decimal("0")
    for r in rows:
        v = r.get(line_total_key)
        if v is not None:
            total_spend += v if isinstance(v, Decimal) else Decimal(str(v))

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = (
        f"{len(rows):,} line items   |   "
        f"Total: ${total_spend:,.2f}   |   "
        f"Generated {datetime.now().strftime('%B %d, %Y  %I:%M %p')}"
    )
    ws["A2"].font = Font(italic=True, size=10, color="666666", name="Calibri")

    ws.merge_cells(f"A3:{last_col_letter}3")
    ws["A3"] = "Click the filter arrows in row 4 to sort/filter by any column."
    ws["A3"].font = Font(italic=True, size=10, color=_EDS_LIGHT, name="Calibri")

    # --- Header row (row 4) ---
    for col_idx, header in enumerate(col_names, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = _XLSX_COL_WIDTHS.get(header, 14)
    ws.row_dimensions[4].height = 32

    # --- Data rows ---
    first_data_row = 5
    for r_offset, row in enumerate(rows):
        excel_row = first_data_row + r_offset
        zebra = (r_offset % 2 == 1)
        for c_idx, header in enumerate(col_names, start=1):
            value = row.get(header)
            cell = ws.cell(row=excel_row, column=c_idx)
            cell.border = border
            cell.font = data_font

            if value is None:
                cell.value = None
            elif c_idx in _XLSX_CURRENCY_COLS:
                # Assign pyodbc's Decimal directly -- openpyxl supports it
                # natively and avoids any float-precision drift on money.
                cell.value = value
                cell.number_format = _XLSX_MONEY_FMT
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx in _XLSX_ID_COLS:
                cell.value = int(value)
                cell.number_format = _XLSX_ID_FMT
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx in _XLSX_INT_COLS:
                cell.value = int(value)
                cell.number_format = _XLSX_QTY_FMT
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx in _XLSX_DATE_COLS:
                cell.value = value
                cell.number_format = _XLSX_DATE_FMT
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")

            if zebra:
                cell.fill = alt_fill

    last_data_row = first_data_row + len(rows) - 1 if rows else 4

    # --- Totals row (Qty + Line Total) ---
    if rows:
        totals_row = last_data_row + 1
        label_cell = ws.cell(row=totals_row, column=1, value="TOTAL")
        label_cell.font = Font(bold=True, size=11, color=_EDS_RED, name="Calibri")
        label_cell.alignment = Alignment(horizontal="right")

        qty_letter = get_column_letter(11)
        total_letter = get_column_letter(13)

        qty_sum = ws.cell(row=totals_row, column=11)
        qty_sum.value = f"=SUM({qty_letter}{first_data_row}:{qty_letter}{last_data_row})"
        qty_sum.number_format = _XLSX_QTY_FMT
        qty_sum.font = Font(bold=True, size=11, name="Calibri")
        qty_sum.alignment = Alignment(horizontal="right")
        qty_sum.border = Border(top=Side(style="double"))

        total_sum = ws.cell(row=totals_row, column=13)
        total_sum.value = f"=SUM({total_letter}{first_data_row}:{total_letter}{last_data_row})"
        total_sum.number_format = _XLSX_MONEY_FMT
        total_sum.font = Font(bold=True, size=11, name="Calibri")
        total_sum.alignment = Alignment(horizontal="right")
        total_sum.border = Border(top=Side(style="double"))

    # --- Freeze + auto-filter ---
    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A4:{last_col_letter}{last_data_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@router.get("/district-items")
async def export_district_items_xlsx(
    x_session_id: Optional[str] = Header(None, alias="X-Session-ID"),
    x_demo_approval: Optional[int] = Header(None, alias="X-Demo-Approval-Level"),
    district_code: str = Query(..., description="District code, e.g. '7G' for Colton-Pierrepont"),
    date_start: Optional[str] = Query(None, description="Inclusive start date (YYYY-MM-DD). Defaults to most recently completed budget year."),
    date_end: Optional[str] = Query(None, description="Inclusive end date (YYYY-MM-DD). Defaults to Nov 30 of the most recently completed budget year."),
):
    """
    Download an Excel workbook with every line item ordered by a single
    district over a date range.

    One row per Detail line, sortable/filterable in Excel. Columns:
    Requisition, Order Date, School, User #, User Name, Vendor Code,
    Vendor, Item #, EDS Item Code, Item Name, Qty, Unit Price,
    Line Total, Status.

    Default date range is the most recently completed EDS budget year
    (Dec 1 - Nov 30).

    Access rules:
    - Requires admin/approver/reports_viewer (same as the rest of /reports).
    - A user's DistrictId must match the requested district, unless the
      user has ApprovalLevel >= 8 (EDS Administration / System Admin),
      which can pull any district.
    """
    demo_level = x_demo_approval if x_demo_approval is not None else 1
    user = _get_session_user(x_session_id, demo_approval_level=demo_level)
    require_reports_access(user)

    district_code = (district_code or "").strip()
    if not district_code:
        raise HTTPException(status_code=400, detail="district_code is required")
    if len(district_code) > 4:
        raise HTTPException(status_code=400, detail="district_code must be 1-4 characters")

    # --- Resolve district ---
    district = execute_single(
        "SELECT DistrictId, DistrictCode, Name FROM dbo.District WHERE DistrictCode = ?",
        (district_code,),
    )
    if not district:
        raise HTTPException(status_code=404, detail=f"District '{district_code}' not found")

    requested_district_id = int(district["DistrictId"])
    user_district_id = int(user.get("DistrictId") or 0)
    user_level = int(user.get("ApprovalLevel") or 0)

    # Cross-district access is gated to EDS-level admins.
    if user_district_id and user_district_id != requested_district_id and user_level < 8:
        raise HTTPException(
            status_code=403,
            detail="You do not have access to another district's data",
        )

    # --- Resolve date range ---
    default_start, default_end_excl, default_label = _default_budget_year_range()

    if date_start or date_end:
        if not (date_start and date_end):
            raise HTTPException(
                status_code=400,
                detail="Provide both date_start and date_end, or neither (to use the default budget year).",
            )
        try:
            start_dt = datetime.strptime(date_start, "%Y-%m-%d")
            end_dt = datetime.strptime(date_end, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="Dates must be in YYYY-MM-DD format")
        if start_dt > end_dt:
            raise HTTPException(status_code=400, detail="date_start must be on or before date_end")

        sql_start = date_start
        # Upper bound is exclusive, so bump end date forward one day to make
        # the user-supplied "inclusive" date_end actually inclusive.
        sql_end_exclusive = (end_dt + timedelta(days=1)).strftime("%Y-%m-%d")
        inclusive_end_label = date_end
        period_label = f"{date_start} to {date_end}"
    else:
        sql_start = default_start
        sql_end_exclusive = default_end_excl
        inclusive_end_label = (
            datetime.strptime(default_end_excl, "%Y-%m-%d") - timedelta(days=1)
        ).strftime("%Y-%m-%d")
        period_label = default_label

    # --- Query ---
    try:
        rows = execute_query(
            DISTRICT_ITEMS_SQL,
            (district_code, sql_start, sql_end_exclusive),
        )
    except Exception as e:
        logger.error("District items export query failed: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to run district items query")

    # Column ordering (matches the SELECT list above)
    col_names = [
        "Requisition", "Order Date", "School", "User #", "User Name",
        "Vendor Code", "Vendor", "Item #", "EDS Item Code", "Item Name",
        "Qty", "Unit Price", "Line Total", "Status",
    ]

    xlsx_bytes = _build_district_items_xlsx(
        rows=rows,
        col_names=col_names,
        district_code=district["DistrictCode"],
        district_name=district["Name"] or district["DistrictCode"],
        date_start=sql_start,
        date_end_exclusive=sql_end_exclusive,
        period_label=period_label,
    )

    safe_name = (district["Name"] or district_code).replace(" ", "_").replace("/", "_")
    filename = (
        f"{safe_name}_{district_code}_Items_Ordered_"
        f"{sql_start}_to_{inclusive_end_label}.xlsx"
    )

    return StreamingResponse(
        iter([xlsx_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_pdf_report(data: dict, date_start: str, date_end: str, period: str) -> bytes:
    """Build a formatted PDF report using ReportLab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.6 * inch, bottomMargin=0.5 * inch,
                            leftMargin=0.6 * inch, rightMargin=0.6 * inch)
    styles = getSampleStyleSheet()

    # Custom styles
    eds_blue = colors.HexColor("#1c1a83")
    eds_red = colors.HexColor("#b70c0d")
    title_style = ParagraphStyle("EDSTitle", parent=styles["Title"], textColor=eds_blue,
                                  fontSize=20, spaceAfter=4)
    subtitle_style = ParagraphStyle("EDSSub", parent=styles["Normal"], textColor=colors.gray,
                                     fontSize=10, spaceAfter=12)
    heading_style = ParagraphStyle("EDSHeading", parent=styles["Heading2"], textColor=eds_blue,
                                    fontSize=14, spaceBefore=16, spaceAfter=8)
    normal = styles["Normal"]

    elements = []

    # --- Header ---
    period_labels = {"current": "Current Budget Year", "previous": "Previous Budget Year", "ytd": "Year to Date"}
    elements.append(Paragraph("EDS Spending Report", title_style))
    elements.append(Paragraph(f"{period_labels.get(period, period)} &bull; {date_start} to {date_end}", subtitle_style))
    elements.append(HRFlowable(width="100%", thickness=1, color=eds_blue, spaceAfter=12))

    # --- Summary Stats ---
    s = data.get("summary", {})
    fmt = lambda v: f"${v:,.2f}" if v else "$0.00"
    summary_data = [
        ["Total Spend", "Total Orders", "Avg Order Value", "Items Ordered", "Pending", "Approved (Month)"],
        [fmt(s.get("total_spend", 0)), str(s.get("total_orders", 0)), fmt(s.get("avg_order_value", 0)),
         str(s.get("total_items", 0)), str(s.get("pending_orders", 0)), str(s.get("approved_this_month", 0))],
    ]
    t = Table(summary_data, colWidths=[1.2 * inch] * 6)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), eds_blue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, 1), 10),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(t)

    # --- Top Vendors ---
    vendors = data.get("vendor_spend", [])
    if vendors:
        elements.append(Paragraph("Top Vendors by Spend", heading_style))
        vdata = [["#", "Vendor", "Code", "Orders", "Spend", "Share"]]
        for i, v in enumerate(vendors, 1):
            vdata.append([str(i), v["name"], v["code"], str(v["orders"]),
                          fmt(v["spend"]), f'{v["pct"]:.1f}%'])
        vt = Table(vdata, colWidths=[0.3 * inch, 2.5 * inch, 0.6 * inch, 0.6 * inch, 1.1 * inch, 0.6 * inch])
        vt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), eds_blue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f8ff")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(vt)

    # --- Category Spending ---
    cats = data.get("category_spend", [])
    if cats:
        elements.append(Paragraph("Spending by Category", heading_style))
        cdata = [["Category", "Spend", "Share"]]
        for c in cats:
            cdata.append([c["name"], fmt(c["spend"]), f'{c["pct"]:.1f}%'])
        ct = Table(cdata, colWidths=[3.0 * inch, 1.2 * inch, 0.8 * inch])
        ct.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), eds_blue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f8ff")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(ct)

    # --- Monthly Trend ---
    trend = data.get("monthly_trend", [])
    if trend:
        elements.append(Paragraph("Monthly Spending Trend", heading_style))
        mdata = [["Month", "Amount"]]
        for m in trend:
            mdata.append([m["month"], fmt(m["amount"])])
        mt = Table(mdata, colWidths=[1.0 * inch, 1.2 * inch])
        mt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), eds_blue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f8ff")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(mt)

    # --- Recent Orders ---
    orders = data.get("recent_orders", [])
    if orders:
        elements.append(Paragraph("Recent Orders", heading_style))
        odata = [["Order #", "Date", "Requester", "Dept", "Items", "Amount", "Status"]]
        for o in orders[:15]:  # Limit to 15 for PDF
            odata.append([o["id"], o["date"], o["requester"][:20], o["dept"][:15],
                          str(o["items"]), fmt(o["amount"]), o["status"]])
        ot = Table(odata, colWidths=[1.0 * inch, 0.75 * inch, 1.3 * inch, 0.9 * inch, 0.45 * inch, 0.9 * inch, 0.9 * inch])
        ot.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), eds_blue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (4, 0), (5, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f8ff")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(ot)

    # --- Footer ---
    elements.append(Spacer(1, 16))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.gray, spaceAfter=4))
    elements.append(Paragraph(
        f"Generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')} &bull; EDS Universal Requisition System",
        ParagraphStyle("Footer", parent=normal, fontSize=7, textColor=colors.gray, alignment=1)
    ))

    doc.build(elements)
    return buf.getvalue()


# ===========================================
# DRILL-DOWN ENDPOINTS
# ===========================================

@router.get("/drilldown/vendor")
async def vendor_drilldown(
    vendor_code: str = Query(..., description="Vendor code to drill into"),
    x_session_id: Optional[str] = Header(None, alias="X-Session-ID"),
    x_demo_approval: Optional[int] = Header(None, alias="X-Demo-Approval-Level"),
    period: str = Query("current"),
    date_start: Optional[str] = Query(None),
    date_end: Optional[str] = Query(None),
):
    """
    Get line-item detail for a specific vendor.
    Returns top requisition line items for the vendor in the given period.
    """
    demo_level = x_demo_approval if x_demo_approval is not None else 1
    user = _get_session_user(x_session_id, demo_approval_level=demo_level)
    require_reports_access(user)

    district_id = user.get("DistrictId", 0)


    if period == "custom" and date_start and date_end:
        ds, de = date_start, date_end
    else:
        ds, de = _compute_date_range(period)

    try:
        district_filter = "AND sch.DistrictId = ?" if district_id else ""
        district_params = [district_id] if district_id else []
        rows = execute_query(f"""
            SELECT TOP 50
                d.Description AS item_name,
                d.ItemCode AS sku,
                d.Quantity AS qty,
                d.BidPrice AS unit_price,
                (d.BidPrice * d.Quantity) AS line_total,
                r.RequisitionNumber AS req_number,
                CONVERT(VARCHAR(10), r.DateEntered, 23) AS date,
                ISNULL(sch.Name, '') AS school
            FROM Detail d
            JOIN Requisitions r ON d.RequisitionId = r.RequisitionId
            JOIN School sch ON r.SchoolId = sch.SchoolId
            JOIN Vendors v ON d.VendorId = v.VendorId
            WHERE v.Code = ?
            AND r.DateEntered >= ? AND r.DateEntered <= ?
            AND r.StatusId NOT IN (1)
            AND (d.Active IS NULL OR d.Active = 1)
            {district_filter}
            ORDER BY (d.BidPrice * d.Quantity) DESC
        """, tuple([vendor_code, ds, de] + district_params))

        vendor_row = execute_single(
            "SELECT Name, Code FROM Vendors WHERE Code = ?", (vendor_code,))
        vendor_name = vendor_row.get("Name", vendor_code) if vendor_row else vendor_code

        return {
            "vendor": {"name": vendor_name, "code": vendor_code},
            "period": {"start": ds, "end": de},
            "items": [
                {
                    "item_name": r.get("item_name") or "Unknown",
                    "sku": r.get("sku") or "",
                    "qty": r.get("qty") or 0,
                    "unit_price": float(r.get("unit_price") or 0),
                    "line_total": float(r.get("line_total") or 0),
                    "req_number": r.get("req_number") or "",
                    "date": r.get("date") or "",
                    "school": r.get("school") or "",
                }
                for r in rows
            ],
            "total_items": len(rows),
        }
    except Exception as e:
        logger.error("Vendor drilldown error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to load vendor drilldown data")


@router.get("/drilldown/category")
async def category_drilldown(
    category_name: str = Query(..., description="Category name to drill into"),
    x_session_id: Optional[str] = Header(None, alias="X-Session-ID"),
    x_demo_approval: Optional[int] = Header(None, alias="X-Demo-Approval-Level"),
    period: str = Query("current"),
    date_start: Optional[str] = Query(None),
    date_end: Optional[str] = Query(None),
):
    """
    Get requisition-level detail for a specific category.
    Returns top requisitions in the given category and period.
    """
    demo_level = x_demo_approval if x_demo_approval is not None else 1
    user = _get_session_user(x_session_id, demo_approval_level=demo_level)
    require_reports_access(user)

    district_id = user.get("DistrictId", 0)


    if period == "custom" and date_start and date_end:
        ds, de = date_start, date_end
    else:
        ds, de = _compute_date_range(period)

    try:
        district_filter = "AND sch.DistrictId = ?" if district_id else ""
        district_params = [district_id] if district_id else []
        rows = execute_query(f"""
            SELECT TOP 50
                r.RequisitionNumber AS req_number,
                CONVERT(VARCHAR(10), r.DateEntered, 23) AS date,
                ISNULL(u.FirstName + ' ' + u.LastName, 'Unknown') AS requester,
                ISNULL(sch.Name, '') AS school,
                r.TotalRequisitionCost AS amount,
                ISNULL(st.Name, '') AS status,
                (SELECT COUNT(*) FROM Detail d2
                 WHERE d2.RequisitionId = r.RequisitionId
                 AND (d2.Active IS NULL OR d2.Active = 1)) AS item_count
            FROM Requisitions r
            JOIN School sch ON r.SchoolId = sch.SchoolId
            LEFT JOIN Category c ON r.CategoryId = c.CategoryId
            LEFT JOIN Users u ON r.UserId = u.UserId
            LEFT JOIN StatusTable st ON r.StatusId = st.StatusId
            WHERE c.Name = ?
            AND r.DateEntered >= ? AND r.DateEntered <= ?
            AND r.StatusId NOT IN (1)
            {district_filter}
            ORDER BY r.TotalRequisitionCost DESC
        """, tuple([category_name, ds, de] + district_params))

        return {
            "category": category_name,
            "period": {"start": ds, "end": de},
            "items": [
                {
                    "req_number": r.get("req_number") or "",
                    "date": r.get("date") or "",
                    "requester": r.get("requester") or "Unknown",
                    "school": r.get("school") or "",
                    "amount": float(r.get("amount") or 0),
                    "status": r.get("status") or "",
                    "item_count": r.get("item_count") or 0,
                }
                for r in rows
            ],
            "total_items": len(rows),
        }
    except Exception as e:
        logger.error("Category drilldown error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to load category drilldown data")



# ===========================================
# DATA QUERY FUNCTIONS (actual EDS schema)
# ===========================================

def _get_spending_summary(district_id: int, date_start: str, date_end: str) -> dict:
    """Get overall spending summary for the period.

    Uses Requisitions table with:
    - DateEntered (date column)
    - TotalRequisitionCost (money column)
    - StatusId (FK to StatusTable)
    - SchoolId (FK to School → District)
    """
    district_filter = "AND s.DistrictId = ?" if district_id else ""
    district_params = [district_id] if district_id else []
    result = execute_single(f"""
        SELECT
            ISNULL(SUM(r.TotalRequisitionCost), 0) AS total_spend,
            COUNT(*) AS total_orders,
            ISNULL(AVG(r.TotalRequisitionCost), 0) AS avg_order_value,
            SUM(CASE WHEN r.StatusId = ? THEN 1 ELSE 0 END) AS pending_orders,
            SUM(CASE WHEN r.StatusId = ?
                AND MONTH(r.DateEntered) = MONTH(GETDATE())
                AND YEAR(r.DateEntered) = YEAR(GETDATE()) THEN 1 ELSE 0 END) AS approved_this_month
        FROM Requisitions r
        JOIN School s ON r.SchoolId = s.SchoolId
        WHERE r.DateEntered >= ?
        AND r.DateEntered <= ?
        AND r.StatusId NOT IN (?)
        {district_filter}
    """, tuple([STATUS_PENDING, STATUS_APPROVED, date_start, date_end, 1] + district_params))

    if not result:
        return {"total_spend": 0, "total_orders": 0, "avg_order_value": 0,
                "total_items": 0, "pending_orders": 0, "approved_this_month": 0,
                "budget_total": 0, "budget_spent": 0, "budget_remaining": 0, "budget_percent": 0}

    total_spend = float(result.get("total_spend", 0) or 0)
    total_orders = result.get("total_orders", 0) or 0

    # Get total line items count
    items_result = execute_single(f"""
        SELECT COUNT(*) AS total_items
        FROM Detail d
        JOIN Requisitions r ON d.RequisitionId = r.RequisitionId
        JOIN School s ON r.SchoolId = s.SchoolId
        WHERE r.DateEntered >= ? AND r.DateEntered <= ?
        AND r.StatusId NOT IN (1)
        {district_filter}
    """, tuple([date_start, date_end] + district_params))
    total_items = (items_result.get("total_items", 0) or 0) if items_result else 0

    # No Departments/Budget table in production — estimate from total
    # Use district's historical average as rough budget proxy
    budget_total = total_spend * 1.75 if total_spend > 0 else 0

    return {
        "total_spend": total_spend,
        "total_orders": total_orders,
        "avg_order_value": round(float(result.get("avg_order_value", 0) or 0), 2),
        "total_items": total_items,
        "pending_orders": result.get("pending_orders", 0) or 0,
        "approved_this_month": result.get("approved_this_month", 0) or 0,
        "budget_total": round(budget_total, 2),
        "budget_spent": total_spend,
        "budget_remaining": round(max(0, budget_total - total_spend), 2),
        "budget_percent": round(total_spend / budget_total * 100, 1) if budget_total > 0 else 0,
    }


def _get_vendor_spend(district_id: int, date_start: str, date_end: str) -> list:
    """Get top vendors by spending.

    Detail has a direct VendorId column, and BidPrice is the unit price.
    Spend = SUM(BidPrice * Quantity) for each vendor.
    """
    district_filter = "AND s.DistrictId = ?" if district_id else ""
    district_params = [district_id] if district_id else []
    rows = execute_query(f"""
        SELECT TOP 10
            v.Name AS name,
            v.Code AS code,
            ISNULL(SUM(d.BidPrice * d.Quantity), 0) AS spend,
            COUNT(DISTINCT r.RequisitionId) AS orders
        FROM Detail d
        JOIN Requisitions r ON d.RequisitionId = r.RequisitionId
        JOIN School s ON r.SchoolId = s.SchoolId
        LEFT JOIN Vendors v ON d.VendorId = v.VendorId
        WHERE r.DateEntered >= ? AND r.DateEntered <= ?
        AND r.StatusId NOT IN (1)
        AND (d.Active IS NULL OR d.Active = 1)
        {district_filter}
        GROUP BY v.Name, v.Code
        ORDER BY spend DESC
    """, tuple([date_start, date_end] + district_params))

    if not rows:
        return []

    total = sum(float(r.get("spend", 0) or 0) for r in rows)
    return [
        {
            "name": r.get("name") or "Unknown Vendor",
            "code": r.get("code") or "",
            "spend": float(r.get("spend", 0) or 0),
            "orders": r.get("orders", 0) or 0,
            "pct": round(float(r.get("spend", 0) or 0) / total * 100, 1) if total > 0 else 0,
        }
        for r in rows
    ]


def _get_category_spend(district_id: int, date_start: str, date_end: str) -> list:
    """Get spending by bid category.

    Requisitions link to Category via CategoryId on the Requisitions table.
    """
    cat_styles = {
        "classroom": {"icon": "fa-pen", "color": "#607d8b"},
        "general": {"icon": "fa-box", "color": "#607d8b"},
        "athletic": {"icon": "fa-running", "color": "#e74c3c"},
        "sport": {"icon": "fa-running", "color": "#e74c3c"},
        "art": {"icon": "fa-palette", "color": "#e91e63"},
        "science": {"icon": "fa-flask", "color": "#1abc9c"},
        "technology": {"icon": "fa-laptop", "color": "#3498db"},
        "custodial": {"icon": "fa-broom", "color": "#27ae60"},
        "janitorial": {"icon": "fa-broom", "color": "#27ae60"},
        "maintenance": {"icon": "fa-wrench", "color": "#795548"},
        "furniture": {"icon": "fa-chair", "color": "#8e44ad"},
        "health": {"icon": "fa-medkit", "color": "#f44336"},
        "medical": {"icon": "fa-medkit", "color": "#f44336"},
        "music": {"icon": "fa-music", "color": "#9c27b0"},
        "office": {"icon": "fa-pen", "color": "#607d8b"},
        "textbook": {"icon": "fa-book", "color": "#ff9800"},
        "food": {"icon": "fa-utensils", "color": "#4caf50"},
        "paper": {"icon": "fa-newspaper", "color": "#78909c"},
    }

    district_filter = "AND s.DistrictId = ?" if district_id else ""
    district_params = [district_id] if district_id else []
    rows = execute_query(f"""
        SELECT TOP 10
            ISNULL(c.Name, 'Other') AS name,
            ISNULL(SUM(r.TotalRequisitionCost), 0) AS spend
        FROM Requisitions r
        JOIN School s ON r.SchoolId = s.SchoolId
        LEFT JOIN Category c ON r.CategoryId = c.CategoryId
        WHERE r.DateEntered >= ? AND r.DateEntered <= ?
        AND r.StatusId NOT IN (1)
        {district_filter}
        GROUP BY c.Name
        ORDER BY spend DESC
    """, tuple([date_start, date_end] + district_params))

    if not rows:
        return []

    total = sum(float(r.get("spend", 0) or 0) for r in rows)
    result = []
    for r in rows:
        name = r.get("name") or "Other"
        name_lower = name.lower()
        style = {"icon": "fa-box", "color": "#9e9e9e"}
        for key, val in cat_styles.items():
            if key in name_lower:
                style = val
                break
        result.append({
            "name": name,
            "spend": float(r.get("spend", 0) or 0),
            "pct": round(float(r.get("spend", 0) or 0) / total * 100, 1) if total > 0 else 0,
            "icon": style["icon"],
            "color": style["color"],
        })
    return result


def _get_monthly_trend(district_id: int, date_start: str, date_end: str) -> list:
    """Get monthly spending trend using DateEntered."""
    months = ["Dec", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov"]

    district_filter = "AND s.DistrictId = ?" if district_id else ""
    district_params = [district_id] if district_id else []
    rows = execute_query(f"""
        SELECT
            MONTH(r.DateEntered) AS month_num,
            ISNULL(SUM(r.TotalRequisitionCost), 0) AS amount
        FROM Requisitions r
        JOIN School s ON r.SchoolId = s.SchoolId
        WHERE r.DateEntered >= ? AND r.DateEntered <= ?
        AND r.StatusId NOT IN (1)
        {district_filter}
        GROUP BY MONTH(r.DateEntered)
    """, tuple([date_start, date_end] + district_params))

    month_data = {r.get("month_num"): float(r.get("amount", 0) or 0) for r in rows}
    # Budget year order: Dec, Jan, Feb, ..., Nov
    month_order = [12, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
    return [
        {"month": months[i], "amount": month_data.get(month_order[i], 0)}
        for i in range(12)
    ]


def _get_recent_orders(district_id: int, date_start: str, date_end: str) -> list:
    """Get recent requisitions with status names from StatusTable."""
    district_filter = "AND sch.DistrictId = ?" if district_id else ""
    district_params = [district_id] if district_id else []
    rows = execute_query(f"""
        SELECT TOP 20
            r.RequisitionId,
            r.RequisitionNumber AS id,
            CONVERT(VARCHAR(10), r.DateEntered, 23) AS date,
            ISNULL(u.FirstName + ' ' + u.LastName, 'Unknown') AS requester,
            ISNULL(sch.Name, '') AS dept,
            (SELECT COUNT(*) FROM Detail d WHERE d.RequisitionId = r.RequisitionId AND (d.Active IS NULL OR d.Active = 1)) AS items,
            r.TotalRequisitionCost AS amount,
            ISNULL(st.Name, 'Unknown') AS status
        FROM Requisitions r
        JOIN School sch ON r.SchoolId = sch.SchoolId
        LEFT JOIN Users u ON r.UserId = u.UserId
        LEFT JOIN StatusTable st ON r.StatusId = st.StatusId
        WHERE r.DateEntered >= ? AND r.DateEntered <= ?
        {district_filter}
        ORDER BY r.DateEntered DESC
    """, tuple([date_start, date_end] + district_params))

    return [
        {
            "id": r.get("id") or f"REQ-{r.get('RequisitionId', 0)}",
            "date": r.get("date", ""),
            "requester": r.get("requester", "Unknown"),
            "dept": r.get("dept", ""),
            "items": r.get("items", 0) or 0,
            "amount": float(r.get("amount", 0) or 0),
            "status": r.get("status", "Unknown"),
        }
        for r in rows
    ]


def _get_school_budgets(district_id: int, date_start: str, date_end: str) -> list:
    """Get per-school spending breakdown (acts as department budget view).

    Since there's no Departments table with budgets, we show spending
    by school within the district as a budget utilization proxy.
    """
    try:
        district_filter = "AND sch.DistrictId = ?" if district_id else ""
        district_params = [district_id] if district_id else []
        rows = execute_query(f"""
            SELECT TOP 10
                sch.Name AS name,
                COUNT(*) AS req_count,
                ISNULL(SUM(r.TotalRequisitionCost), 0) AS spent
            FROM Requisitions r
            JOIN School sch ON r.SchoolId = sch.SchoolId
            WHERE r.DateEntered >= ? AND r.DateEntered <= ?
            AND r.StatusId NOT IN (1)
            {district_filter}
            GROUP BY sch.Name
            ORDER BY spent DESC
        """, tuple([date_start, date_end] + district_params))
    except Exception:
        logger.warning("School budget query failed — returning empty data")
        return []

    if not rows:
        return []

    # Use the highest spender as a rough budget benchmark (120% of their spend)
    max_spent = max(float(r.get("spent", 0) or 0) for r in rows)

    return [
        {
            "name": r.get("name") or "Unknown School",
            "budget": round(float(r.get("spent", 0) or 0) * 1.2, 2),  # Estimated budget
            "spent": float(r.get("spent", 0) or 0),
            "pct": round(float(r.get("spent", 0) or 0) / (float(r.get("spent", 0) or 0) * 1.2) * 100, 1)
                   if float(r.get("spent", 0) or 0) > 0 else 0,
        }
        for r in rows
    ]


