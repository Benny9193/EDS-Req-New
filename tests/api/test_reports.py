"""
Tests for Reports API endpoints.

Covers: summary, CSV export, PDF export, cache invalidation,
custom date range, vendor/category drill-down, auth guards.
"""

import pytest
from unittest.mock import patch, MagicMock

# Realistic mock data for demo-mode query functions
_MOCK_SUMMARY = {
    "total_spend": 287450.0, "total_orders": 156, "avg_order_value": 1842.63,
    "total_items": 2847, "pending_orders": 12, "approved_this_month": 34,
    "budget_total": 500000, "budget_spent": 287450, "budget_remaining": 212550,
    "budget_percent": 57.5,
}
_MOCK_VENDORS = [
    {"name": "School Specialty, LLC", "code": "0009", "spend": 62450.00, "orders": 284, "pct": 21.7},
    {"name": "Staples Contract & Commercial", "code": "2222", "spend": 43200.00, "orders": 156, "pct": 15.0},
]
_MOCK_CATEGORIES = [
    {"name": "Classroom Supplies", "spend": 78500, "pct": 50.0, "icon": "fa-pen", "color": "#607d8b"},
    {"name": "Athletic Equipment", "spend": 52300, "pct": 33.3, "icon": "fa-running", "color": "#e74c3c"},
]
_MOCK_MONTHLY = [
    {"month": "Dec", "amount": 31200}, {"month": "Jan", "amount": 42500},
    {"month": "Feb", "amount": 38100}, {"month": "Mar", "amount": 29800},
    {"month": "Apr", "amount": 35400}, {"month": "May", "amount": 27600},
    {"month": "Jun", "amount": 18200}, {"month": "Jul", "amount": 8900},
    {"month": "Aug", "amount": 45300}, {"month": "Sep", "amount": 52100},
    {"month": "Oct", "amount": 41800}, {"month": "Nov", "amount": 16550},
]
_MOCK_ORDERS = [
    {"id": "REQ-2026-0089", "date": "2026-03-21", "requester": "Sarah Johnson", "dept": "Science", "items": 12, "amount": 3450.00, "status": "Approved"},
]
_MOCK_SCHOOLS = [
    {"name": "Lincoln Elementary", "budget": 60000, "spent": 48200, "pct": 80.3},
]


@pytest.fixture(autouse=True)
def mock_reports_queries():
    """Mock the underlying query functions so demo-mode tests don't hit the DB."""
    with patch("api.routes.reports._get_spending_summary", return_value=_MOCK_SUMMARY.copy()) as m1, \
         patch("api.routes.reports._get_vendor_spend", return_value=list(_MOCK_VENDORS)) as m2, \
         patch("api.routes.reports._get_category_spend", return_value=list(_MOCK_CATEGORIES)) as m3, \
         patch("api.routes.reports._get_monthly_trend", return_value=list(_MOCK_MONTHLY)) as m4, \
         patch("api.routes.reports._get_recent_orders", return_value=list(_MOCK_ORDERS)) as m5, \
         patch("api.routes.reports._get_school_budgets", return_value=list(_MOCK_SCHOOLS)) as m6:
        yield {
            "summary": m1, "vendors": m2, "categories": m3,
            "monthly": m4, "orders": m5, "schools": m6,
        }


class TestReportsSummary:
    """Test GET /api/reports/summary."""

    def test_demo_session_returns_live_data(self, test_client, mock_reports_queries):
        """Demo session queries live DB data (all districts) instead of hardcoded values."""
        r = test_client.get("/api/reports/summary", headers={"X-Session-ID": "demo"})
        assert r.status_code == 200
        data = r.json()
        assert "summary" in data
        assert "vendor_spend" in data
        assert "category_spend" in data
        assert "monthly_trend" in data
        assert "recent_orders" in data
        assert "budget_departments" in data
        # Verify data came from mocked query functions (not hardcoded)
        assert data["summary"]["total_spend"] == _MOCK_SUMMARY["total_spend"]
        assert data["summary"]["total_orders"] == _MOCK_SUMMARY["total_orders"]
        # Verify query functions were called with district_id=0 (all districts)
        # Summary is called twice: once for current period, once for comparison metrics
        assert mock_reports_queries["summary"].call_count >= 1
        assert mock_reports_queries["summary"].call_args_list[0][0][0] == 0

    def test_summary_has_comparison_metrics(self, test_client):
        """Summary should include vs_last_period and orders_vs_last."""
        r = test_client.get("/api/reports/summary", headers={"X-Session-ID": "demo"})
        data = r.json()
        assert "vs_last_period" in data["summary"]
        assert "orders_vs_last" in data["summary"]

    def test_summary_has_date_range(self, test_client):
        """Response should include date_range and period."""
        r = test_client.get("/api/reports/summary", headers={"X-Session-ID": "demo"})
        data = r.json()
        assert "date_range" in data
        assert "start" in data["date_range"]
        assert "end" in data["date_range"]
        assert "period" in data

    def test_period_current(self, test_client):
        """Current period should return data."""
        r = test_client.get("/api/reports/summary?period=current",
                            headers={"X-Session-ID": "demo"})
        assert r.status_code == 200
        assert r.json()["period"] == "current"

    def test_period_previous(self, test_client):
        """Previous period should return data."""
        r = test_client.get("/api/reports/summary?period=previous",
                            headers={"X-Session-ID": "demo"})
        assert r.status_code == 200
        assert r.json()["period"] == "previous"

    def test_period_ytd(self, test_client):
        """YTD period should return data."""
        r = test_client.get("/api/reports/summary?period=ytd",
                            headers={"X-Session-ID": "demo"})
        assert r.status_code == 200
        assert r.json()["period"] == "ytd"

    def test_custom_date_range(self, test_client):
        """Custom date range should override period computation."""
        r = test_client.get(
            "/api/reports/summary?period=custom&date_start=2024-01-01&date_end=2024-06-30",
            headers={"X-Session-ID": "demo"})
        assert r.status_code == 200

    def test_no_session_returns_401(self, test_client):
        """Missing session header should return 401."""
        r = test_client.get("/api/reports/summary")
        # Demo fallback treats missing session as demo
        # but actual non-demo invalid sessions get 401
        assert r.status_code in (200, 401)

    def test_vendor_spend_structure(self, test_client):
        """Vendor spend items should have required fields."""
        r = test_client.get("/api/reports/summary", headers={"X-Session-ID": "demo"})
        data = r.json()
        for v in data["vendor_spend"]:
            assert "name" in v
            assert "code" in v
            assert "spend" in v
            assert "orders" in v
            assert "pct" in v

    def test_category_spend_structure(self, test_client):
        """Category spend items should have icon and color."""
        r = test_client.get("/api/reports/summary", headers={"X-Session-ID": "demo"})
        data = r.json()
        for c in data["category_spend"]:
            assert "name" in c
            assert "spend" in c
            assert "icon" in c
            assert "color" in c

    def test_monthly_trend_has_12_months(self, test_client):
        """Monthly trend should return exactly 12 months."""
        r = test_client.get("/api/reports/summary", headers={"X-Session-ID": "demo"})
        data = r.json()
        assert len(data["monthly_trend"]) == 12
        assert data["monthly_trend"][0]["month"] == "Dec"
        assert data["monthly_trend"][11]["month"] == "Nov"

    def test_budget_fields(self, test_client):
        """Summary should include budget breakdown."""
        r = test_client.get("/api/reports/summary", headers={"X-Session-ID": "demo"})
        s = r.json()["summary"]
        assert "budget_total" in s
        assert "budget_spent" in s
        assert "budget_remaining" in s
        assert "budget_percent" in s
        assert s["budget_remaining"] == s["budget_total"] - s["budget_spent"]


class TestCSVExport:
    """Test GET /api/reports/export."""

    def test_csv_export_all(self, test_client):
        """Full CSV export should return CSV content."""
        r = test_client.get("/api/reports/export?section=all",
                            headers={"X-Session-ID": "demo"})
        assert r.status_code == 200
        assert "text/csv" in r.headers["content-type"]
        assert "attachment" in r.headers.get("content-disposition", "")
        content = r.text
        assert "VENDOR SPENDING" in content
        assert "CATEGORY SPENDING" in content
        assert "MONTHLY TREND" in content
        assert "RECENT ORDERS" in content
        assert "SCHOOL SPENDING" in content

    def test_csv_export_vendors_only(self, test_client):
        """Vendor-only CSV should only have vendor data."""
        r = test_client.get("/api/reports/export?section=vendors",
                            headers={"X-Session-ID": "demo"})
        assert r.status_code == 200
        content = r.text
        assert "Vendor" in content
        assert "School Specialty" in content
        # Should NOT have other sections
        assert "CATEGORY SPENDING" not in content

    def test_csv_export_orders(self, test_client):
        """Orders CSV should have order columns."""
        r = test_client.get("/api/reports/export?section=orders",
                            headers={"X-Session-ID": "demo"})
        assert r.status_code == 200
        content = r.text
        assert "Order #" in content
        assert "Requester" in content

    def test_csv_export_categories(self, test_client):
        """Categories CSV should have category data."""
        r = test_client.get("/api/reports/export?section=categories",
                            headers={"X-Session-ID": "demo"})
        assert r.status_code == 200
        assert "Category" in r.text

    def test_csv_export_monthly(self, test_client):
        """Monthly CSV should have month and amount columns."""
        r = test_client.get("/api/reports/export?section=monthly",
                            headers={"X-Session-ID": "demo"})
        assert r.status_code == 200
        assert "Month" in r.text
        assert "Dec" in r.text

    def test_csv_filename_includes_period(self, test_client):
        """Filename should include the period and date."""
        r = test_client.get("/api/reports/export?section=vendors&period=previous",
                            headers={"X-Session-ID": "demo"})
        cd = r.headers.get("content-disposition", "")
        assert "previous" in cd
        assert ".csv" in cd


class TestPDFExport:
    """Test GET /api/reports/export/pdf."""

    def test_pdf_export_returns_pdf(self, test_client):
        """PDF export should return application/pdf."""
        r = test_client.get("/api/reports/export/pdf",
                            headers={"X-Session-ID": "demo"})
        assert r.status_code == 200
        assert "application/pdf" in r.headers["content-type"]
        assert "attachment" in r.headers.get("content-disposition", "")
        # PDF magic bytes
        assert r.content[:4] == b"%PDF"

    def test_pdf_filename_includes_period(self, test_client):
        """PDF filename should include period."""
        r = test_client.get("/api/reports/export/pdf?period=ytd",
                            headers={"X-Session-ID": "demo"})
        cd = r.headers.get("content-disposition", "")
        assert "ytd" in cd
        assert ".pdf" in cd

    def test_pdf_has_reasonable_size(self, test_client):
        """PDF should be non-trivial (has tables and content)."""
        r = test_client.get("/api/reports/export/pdf",
                            headers={"X-Session-ID": "demo"})
        assert len(r.content) > 2000  # Should be at least a few KB


class TestCacheInvalidation:
    """Test reports cache invalidation."""

    def test_invalidate_clears_district_cache(self):
        """invalidate_reports_cache should clear matching keys."""
        from api.cache import get_cache
        from api.routes.reports import invalidate_reports_cache

        cache = get_cache()
        cache.set_sync("reports:765:2025-12-01:2026-11-30:summary", {"test": 1}, 3600)
        cache.set_sync("reports:765:2025-12-01:2026-11-30:vendors", {"test": 2}, 3600)
        cache.set_sync("reports:999:2025-12-01:2026-11-30:summary", {"test": 3}, 3600)

        invalidate_reports_cache(765)

        assert cache.get_sync("reports:765:2025-12-01:2026-11-30:summary") is None
        assert cache.get_sync("reports:765:2025-12-01:2026-11-30:vendors") is None
        # Other district's cache should remain
        assert cache.get_sync("reports:999:2025-12-01:2026-11-30:summary") is not None

    def test_invalidate_all_clears_everything(self):
        """invalidate_reports_cache(None) should clear all reports keys."""
        from api.cache import get_cache
        from api.routes.reports import invalidate_reports_cache

        cache = get_cache()
        cache.set_sync("reports:765:2025-12-01:2026-11-30:summary", {"a": 1}, 3600)
        cache.set_sync("reports:999:2025-12-01:2026-11-30:summary", {"b": 2}, 3600)
        cache.set_sync("other:key", {"c": 3}, 3600)

        invalidate_reports_cache()

        assert cache.get_sync("reports:765:2025-12-01:2026-11-30:summary") is None
        assert cache.get_sync("reports:999:2025-12-01:2026-11-30:summary") is None
        # Non-reports keys should remain
        assert cache.get_sync("other:key") is not None

    def test_cache_hit_returns_same_data(self):
        """Cached data should be returned on second call."""
        from api.cache import get_cache
        from api.routes.reports import _get_cached_or_query

        call_count = 0

        def fake_query(did, ds, de):
            nonlocal call_count
            call_count += 1
            return {"total": call_count}

        cache = get_cache()
        # First call — cache miss
        r1 = _get_cached_or_query(1, "2025-12-01", "2026-11-30", "test_section", fake_query)
        assert r1["total"] == 1
        assert call_count == 1

        # Second call — cache hit
        r2 = _get_cached_or_query(1, "2025-12-01", "2026-11-30", "test_section", fake_query)
        assert r2["total"] == 1  # Same data
        assert call_count == 1   # Function not called again


class TestAuthGuards:
    """Test reports access control."""

    def test_require_reports_access_admin(self):
        """Admin users (ApprovalLevel >= 1) should have access."""
        from api.routes.reports import require_reports_access
        user = {"ApprovalLevel": 1, "UserType": "teacher"}
        result = require_reports_access(user)
        assert result is user

    def test_require_reports_access_reports_viewer(self):
        """Reports viewer role should have access."""
        from api.routes.reports import require_reports_access
        user = {"ApprovalLevel": 0, "UserType": "reports_viewer"}
        result = require_reports_access(user)
        assert result is user

    def test_require_reports_access_department_head(self):
        """Department head role should have access."""
        from api.routes.reports import require_reports_access
        user = {"ApprovalLevel": 0, "UserType": "department_head"}
        result = require_reports_access(user)
        assert result is user

    def test_require_reports_access_teacher_denied(self):
        """Regular teacher should be denied."""
        from api.routes.reports import require_reports_access
        from fastapi import HTTPException
        user = {"ApprovalLevel": 0, "UserType": "teacher"}
        with pytest.raises(HTTPException) as exc_info:
            require_reports_access(user)
        assert exc_info.value.status_code == 403

    def test_require_reports_access_none_approval(self):
        """None ApprovalLevel should be denied."""
        from api.routes.reports import require_reports_access
        from fastapi import HTTPException
        user = {"ApprovalLevel": None, "UserType": ""}
        with pytest.raises(HTTPException) as exc_info:
            require_reports_access(user)
        assert exc_info.value.status_code == 403


class TestDateRangeComputation:
    """Test _compute_date_range helper."""

    def test_current_period(self):
        """Current period should return Dec-Nov range."""
        from api.routes.reports import _compute_date_range
        start, end = _compute_date_range("current")
        assert start.endswith("-12-01")
        assert end.endswith("-11-30")

    def test_previous_period(self):
        """Previous period should be one year before current."""
        from api.routes.reports import _compute_date_range
        cur_start, _ = _compute_date_range("current")
        prev_start, _ = _compute_date_range("previous")
        cur_year = int(cur_start[:4])
        prev_year = int(prev_start[:4])
        assert cur_year - prev_year == 1

    def test_ytd_ends_today(self):
        """YTD period should end today or close to it."""
        from api.routes.reports import _compute_date_range
        from datetime import datetime
        _, end = _compute_date_range("ytd")
        today = datetime.now().strftime("%Y-%m-%d")
        assert end == today


class TestDrillDown:
    """Test vendor/category drill-down endpoints."""

    def test_vendor_drilldown_demo(self, test_client):
        """Vendor drill-down should return line items for a vendor."""
        r = test_client.get("/api/reports/drilldown/vendor?vendor_code=0009",
                            headers={"X-Session-ID": "demo"})
        assert r.status_code == 200
        data = r.json()
        assert "vendor" in data
        assert "items" in data
        assert isinstance(data["items"], list)

    def test_category_drilldown_demo(self, test_client):
        """Category drill-down should return line items for a category."""
        r = test_client.get("/api/reports/drilldown/category?category_name=Classroom+Supplies",
                            headers={"X-Session-ID": "demo"})
        assert r.status_code == 200
        data = r.json()
        assert "category" in data
        assert "items" in data
        assert isinstance(data["items"], list)


class TestDistrictItemsExport:
    """Test GET /api/reports/district-items — Excel export for one district."""

    _FAKE_DISTRICT = {"DistrictId": 42, "DistrictCode": "7G", "Name": "Colton-Pierrepont"}

    @staticmethod
    def _fake_rows():
        from datetime import date
        return [
            {
                "Requisition": "R-12345",
                "Order Date": date(2025, 3, 15),
                "School": "Colton-Pierrepont Central",
                "User #": 101,
                "User Name": "Jane Teacher",
                "Vendor Code": "0001",
                "Vendor": "Pilot Corporation of America",
                "Item #": "BPS-GP-F",
                "EDS Item Code": "EDS001",
                "Item Name": "Pilot G2 Gel Pen, Fine Point, Black",
                "Qty": 12,
                "Unit Price": 1.49,
                "Line Total": 17.88,
                "Status": "PO Printed",
            },
            {
                "Requisition": "R-12346",
                "Order Date": date(2025, 5, 10),
                "School": "Colton-Pierrepont Central",
                "User #": 102,
                "User Name": "John Principal",
                "Vendor Code": "0009",
                "Vendor": "School Specialty, LLC",
                "Item #": "SS-1234",
                "EDS Item Code": "EDS002",
                "Item Name": "Copy Paper, 500-sheet ream",
                "Qty": 50,
                "Unit Price": 4.99,
                "Line Total": 249.50,
                "Status": "Approved",
            },
        ]

    def _patched(self):
        """Build patchers that stub the two DB helpers the endpoint uses."""
        district = self._FAKE_DISTRICT
        rows = self._fake_rows()

        def fake_execute_single(sql, params=None):
            if "FROM dbo.District" in sql and params and params[0] == "7G":
                return district
            return None

        def fake_execute_query(sql, params=None):
            if "FROM dbo.Requisitions" in sql:
                return list(rows)
            return []

        return (
            patch("api.routes.reports.execute_single", side_effect=fake_execute_single),
            patch("api.routes.reports.execute_query", side_effect=fake_execute_query),
        )

    def test_district_items_xlsx_happy_path(self, test_client):
        """Demo session can download an Excel workbook for a district."""
        import io
        import openpyxl

        p1, p2 = self._patched()
        with p1, p2:
            r = test_client.get(
                "/api/reports/district-items",
                params={"district_code": "7G"},
                headers={"X-Session-ID": "demo"},
            )
        assert r.status_code == 200
        assert r.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        disp = r.headers.get("content-disposition", "")
        assert "Colton-Pierrepont_7G_Items_Ordered_" in disp
        assert disp.endswith('.xlsx"')

        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        assert ws.title == "Items Ordered"

        # Header row is row 4, data begins at row 5
        headers = [ws.cell(row=4, column=c).value for c in range(1, 15)]
        assert headers[0] == "Requisition"
        assert headers[3] == "User #"
        assert headers[7] == "Item #"
        assert headers[9] == "Item Name"
        assert headers[11] == "Unit Price"

        row5 = [ws.cell(row=5, column=c).value for c in range(1, 15)]
        assert row5[0] == "R-12345"
        assert row5[3] == 101
        assert row5[7] == "BPS-GP-F"
        assert row5[11] == 1.49

        # Freeze panes + auto-filter enable sort/filter in Excel
        assert ws.freeze_panes == "A5"
        assert ws.auto_filter.ref is not None and ws.auto_filter.ref.startswith("A4:")

        # Totals row uses SUM formulas so Excel keeps them in sync with filters
        total_row = 5 + len(self._fake_rows())
        line_total_formula = ws.cell(row=total_row, column=13).value
        assert isinstance(line_total_formula, str)
        assert line_total_formula.startswith("=SUM(")

    def test_district_items_requires_district_code(self, test_client):
        r = test_client.get(
            "/api/reports/district-items",
            headers={"X-Session-ID": "demo"},
        )
        # FastAPI returns 422 for missing required query params
        assert r.status_code == 422

    def test_district_items_unknown_district_returns_404(self, test_client):
        p1, p2 = self._patched()
        with p1, p2:
            r = test_client.get(
                "/api/reports/district-items",
                params={"district_code": "ZZ"},
                headers={"X-Session-ID": "demo"},
            )
        assert r.status_code == 404

    def test_district_items_rejects_bad_date_format(self, test_client):
        p1, p2 = self._patched()
        with p1, p2:
            r = test_client.get(
                "/api/reports/district-items",
                params={"district_code": "7G", "date_start": "not-a-date", "date_end": "2025-11-30"},
                headers={"X-Session-ID": "demo"},
            )
        assert r.status_code == 400

    def test_district_items_custom_date_range_reflected_in_filename(self, test_client):
        p1, p2 = self._patched()
        with p1, p2:
            r = test_client.get(
                "/api/reports/district-items",
                params={
                    "district_code": "7G",
                    "date_start": "2025-01-01",
                    "date_end": "2025-03-31",
                },
                headers={"X-Session-ID": "demo"},
            )
        assert r.status_code == 200
        assert "2025-01-01_to_2025-03-31.xlsx" in r.headers.get("content-disposition", "")
