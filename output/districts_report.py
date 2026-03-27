import pyodbc
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

conn = pyodbc.connect(
    'DRIVER={ODBC Driver 17 for SQL Server};'
    'SERVER=eds-sqlserver.eastus2.cloudapp.azure.com;'
    'DATABASE=EDS;'
    'UID=EDSAdmin;'
    'PWD=Consultant~!;'
    'TrustServerCertificate=yes;'
)
cursor = conn.cursor()

cursor.execute("""
    SELECT
        d.DistrictCode,
        d.Name AS DistrictName,
        ISNULL(af.Description, 'Not Assigned') AS AccountingSoftware,
        CASE WHEN d.AllowElectronicPOs = 1 THEN 'Yes' ELSE 'No' END AS UsesEPO
    FROM District d
    LEFT JOIN AccountingFormats af ON d.AccountingFormatId = af.AccountingFormatId
    WHERE d.Active = 1
      AND ISNULL(d.HidefromDistrictLists, 0) = 0
    ORDER BY d.Name
""")

rows = cursor.fetchall()
print(f"Total active districts: {len(rows)}")

# Build workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Districts"

# Styles
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1C1A83")  # EDS primary blue
alt_fill    = PatternFill("solid", fgColor="EEEEF8")
yes_fill    = PatternFill("solid", fgColor="C6EFCE")
no_fill     = PatternFill("solid", fgColor="FFCCCC")
center      = Alignment(horizontal="center", vertical="center")
left        = Alignment(horizontal="left",   vertical="center")
thin        = Side(style="thin", color="CCCCCC")
border      = Border(left=thin, right=thin, top=thin, bottom=thin)

# Header row
headers = ["District Code", "District Name", "Accounting Software", "Uses ePO?"]
col_widths = [15, 45, 35, 12]

for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font   = header_font
    cell.fill   = header_fill
    cell.alignment = center
    cell.border = border
    ws.column_dimensions[cell.column_letter].width = w

ws.row_dimensions[1].height = 22

# Data rows
for row_idx, row in enumerate(rows, start=2):
    district_code, district_name, accounting_software, uses_epo = row
    values = [district_code, district_name, accounting_software, uses_epo]
    is_alt = (row_idx % 2 == 0)

    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.border = border
        cell.alignment = center if col in (1, 4) else left

        if col == 4:  # ePO column - color coded
            cell.fill = yes_fill if val == "Yes" else no_fill
        elif is_alt:
            cell.fill = alt_fill

ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:D{len(rows) + 1}"

out_path = r"C:\EDS\output\Districts_Accounting_ePO.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
