#!/usr/bin/env python3
"""
Orleans (Olean City School District) Top-Ordered Items Report
School Year 2025-2026  (BudgetId 1774722)

Queries EDS for all approved/processed requisition line items
for Olean City School District and ranks them by total quantity ordered.
"""

import os
import sys
from datetime import datetime

import pyodbc

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required.  pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Connection
# ---------------------------------------------------------------------------
CONNECTION_STRING = (
    'DRIVER={ODBC Driver 17 for SQL Server};'
    'SERVER=eds-sqlserver.eastus2.cloudapp.azure.com;'
    'DATABASE=EDS;UID=EDSAdmin;PWD=Consultant~!;'
    'Encrypt=yes;TrustServerCertificate=yes;'
)

DISTRICT_NAME = "Olean City School District"
BUDGET_ID     = 1774722          # 2025-2026 school year budget
YEAR_LABEL    = "2025-2026"

# StatusIds to EXCLUDE (cancelled / pending / rejected / on-hold only)
EXCLUDED_STATUSES = (1, 2, 4)    # On Hold, Pending Approval, Rejected

# ---------------------------------------------------------------------------
# EDS brand colours
# ---------------------------------------------------------------------------
EDS_BLUE  = "1C1A83"
EDS_LIGHT = "4A4890"
WHITE     = "FFFFFF"
ALT_ROW   = "EBEBF5"   # light lavender


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def style_header(ws, num_cols):
    hdr_font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    hdr_fill = PatternFill(start_color=EDS_BLUE, end_color=EDS_BLUE, fill_type="solid")
    bdr      = thin_border()
    for col in range(1, num_cols + 1):
        cell           = ws.cell(row=1, column=col)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.border    = bdr
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value),
            default=10,
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 60)


def shade_alt_rows(ws, num_data_rows, num_cols):
    fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
    for row in range(2, num_data_rows + 2):
        if row % 2 == 0:
            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).fill = fill


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def run_report():
    print("Connecting to EDS...")
    conn = pyodbc.connect(CONNECTION_STRING, timeout=30)
    cur  = conn.cursor()

    excl = ",".join(str(s) for s in EXCLUDED_STATUSES)

    # -----------------------------------------------------------------------
    # 1. Top-ordered items
    # -----------------------------------------------------------------------
    items_sql = f"""
        SELECT
            ROW_NUMBER() OVER (ORDER BY SUM(d.Quantity) DESC)      AS [Rank],
            COALESCE(d.ItemCode, i.ItemCode, '')                    AS [Item Code],
            COALESCE(d.Description, i.Description, 'Unknown')      AS [Description],
            COALESCE(cat.Name, 'Uncategorized')                     AS [Category],
            COALESCE(v.Name, 'N/A')                                 AS [Vendor],
            COALESCE(d.UnitCode, '')                                AS [UOM],
            CAST(SUM(d.Quantity) AS INT)                            AS [Total Qty Ordered],
            COUNT(DISTINCT r.RequisitionId)                         AS [# Requisitions],
            MIN(COALESCE(d.BidPrice, d.CatalogPrice))               AS [Min Unit Price],
            MAX(COALESCE(d.BidPrice, d.CatalogPrice))               AS [Max Unit Price],
            AVG(COALESCE(d.BidPrice, d.CatalogPrice))               AS [Avg Unit Price],
            SUM(d.Quantity * COALESCE(d.BidPrice, d.CatalogPrice))  AS [Total Spend]
        FROM   Requisitions r
        JOIN   Detail     d   ON r.RequisitionId = d.RequisitionId
        LEFT JOIN Items   i   ON d.ItemId        = i.ItemId
        LEFT JOIN Category cat ON i.CategoryId   = cat.CategoryId
        LEFT JOIN Vendors  v   ON d.VendorId     = v.VendorId
        WHERE  r.BudgetId = {BUDGET_ID}
          AND  r.StatusId NOT IN ({excl})
        GROUP BY
            COALESCE(d.ItemCode, i.ItemCode, ''),
            COALESCE(d.Description, i.Description, 'Unknown'),
            COALESCE(cat.Name, 'Uncategorized'),
            COALESCE(v.Name, 'N/A'),
            COALESCE(d.UnitCode, '')
        ORDER BY SUM(d.Quantity) DESC
    """

    print("Querying top-ordered items for 2025-26...")
    cur.execute(items_sql)
    rows     = cur.fetchall()
    col_names = [desc[0] for desc in cur.description]
    print(f"  {len(rows)} unique items found.")

    # -----------------------------------------------------------------------
    # 2. Summary totals
    # -----------------------------------------------------------------------
    summary_sql = f"""
        SELECT
            COUNT(DISTINCT r.RequisitionId)                            AS TotalReqs,
            COUNT(DISTINCT r.SchoolId)                                 AS TotalSchools,
            COUNT(d.DetailId)                                          AS TotalLines,
            SUM(CAST(d.Quantity AS BIGINT))                            AS TotalQty,
            SUM(d.Quantity * COALESCE(d.BidPrice, d.CatalogPrice))     AS TotalSpend
        FROM   Requisitions r
        JOIN   Detail d ON r.RequisitionId = d.RequisitionId
        WHERE  r.BudgetId = {BUDGET_ID}
          AND  r.StatusId NOT IN ({excl})
    """
    cur.execute(summary_sql)
    sm = cur.fetchone()

    conn.close()

    # -----------------------------------------------------------------------
    # 3. Build workbook
    # -----------------------------------------------------------------------
    wb = Workbook()

    # -- Sheet 1: Ranked Items -----------------------------------------------
    ws = wb.active
    ws.title = "Top Ordered Items"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 36

    ws.append(col_names)
    for row in rows:
        ws.append(list(row))

    num_cols = len(col_names)
    num_rows = len(rows)

    style_header(ws, num_cols)
    shade_alt_rows(ws, num_rows, num_cols)

    # Per-cell formatting
    bdr          = thin_border()
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(vertical="center")
    currency_cols = {9, 10, 11, 12}   # Min/Max/Avg price, Total Spend
    integer_cols  = {1, 7, 8}         # Rank, Total Qty, # Reqs

    for r_idx in range(2, num_rows + 2):
        for c_idx in range(1, num_cols + 1):
            cell        = ws.cell(row=r_idx, column=c_idx)
            cell.border = bdr
            if c_idx in currency_cols:
                cell.number_format = '"$"#,##0.00'
                cell.alignment     = Alignment(horizontal="right", vertical="center")
            elif c_idx in integer_cols:
                cell.number_format = "#,##0"
                cell.alignment     = center_align
            else:
                cell.alignment = left_align

    auto_width(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_rows + 1}"

    # -- Sheet 2: Summary ----------------------------------------------------
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 28

    blue_font = Font(bold=True, size=16, color=EDS_BLUE,  name="Calibri")
    sub_font  = Font(italic=True, size=12, color=EDS_LIGHT, name="Calibri")
    meta_font = Font(size=10, color="888888", name="Calibri")
    hdr_fill2 = PatternFill(start_color=EDS_BLUE, end_color=EDS_BLUE, fill_type="solid")
    hdr_font2 = Font(bold=True, color=WHITE, size=11, name="Calibri")
    lbl_font  = Font(bold=True, size=11, name="Calibri")
    val_font  = Font(size=11, name="Calibri")

    ws2["A1"] = DISTRICT_NAME
    ws2["A1"].font = blue_font
    ws2["A2"] = f"Top Ordered Items -- School Year {YEAR_LABEL}"
    ws2["A2"].font = sub_font
    ws2["A3"] = f"Generated {datetime.now().strftime('%B %d, %Y  %I:%M %p')}"
    ws2["A3"].font = meta_font

    ws2.row_dimensions[5].height = 22
    for col, label in [(1, "Metric"), (2, "Value")]:
        c           = ws2.cell(row=5, column=col, value=label)
        c.font      = hdr_font2
        c.fill      = hdr_fill2
        c.alignment = Alignment(horizontal="center", vertical="center")

    data = [
        ("School Year",            f"{YEAR_LABEL}  (Jul 1, 2025 - Jun 30, 2026)"),
        ("Total Requisitions",     sm.TotalReqs    if sm else "N/A"),
        ("Schools / Locations",    sm.TotalSchools if sm else "N/A"),
        ("Total Line Items",       sm.TotalLines   if sm else "N/A"),
        ("Total Qty Ordered",      sm.TotalQty     if sm else "N/A"),
        ("Total Spend",            sm.TotalSpend   if sm else "N/A"),
        ("Unique Items in Report", num_rows),
    ]
    alt_fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
    for i, (label, val) in enumerate(data, start=6):
        lc      = ws2.cell(row=i, column=1, value=label)
        lc.font = lbl_font
        vc      = ws2.cell(row=i, column=2, value=val)
        vc.font = val_font

        if label == "Total Spend" and isinstance(val, (int, float)):
            vc.number_format = '"$"#,##0.00'
        elif label in ("Total Requisitions", "Total Qty Ordered",
                       "Total Line Items", "Unique Items in Report",
                       "Schools / Locations") and isinstance(val, (int, float)):
            vc.number_format = "#,##0"

        if i % 2 == 0:
            lc.fill = alt_fill
            vc.fill = alt_fill

    # -----------------------------------------------------------------------
    # 4. Save
    # -----------------------------------------------------------------------
    os.makedirs("output", exist_ok=True)
    out_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "output",
        f"Orleans_TopItems_{YEAR_LABEL.replace('-', '_')}.xlsx",
    )
    wb.save(out_path)
    print(f"\n[OK] Report saved to: {out_path}")
    return out_path


if __name__ == "__main__":
    run_report()
