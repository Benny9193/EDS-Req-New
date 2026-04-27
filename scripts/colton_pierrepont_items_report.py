#!/usr/bin/env python3
"""
Colton-Pierrepont (district code 7G) Items Ordered Report

Request from Patrice Abate / Educational Data Services:
    One row per line item ordered by Colton-Pierrepont so the district can
    sort/filter in Excel to see who is ordering a given item (e.g. a Pilot
    pen) and at what price, or to see every item a given user has ordered.

Default date range: most recently completed budget year,
    Dec 1, 2024 - Nov 30, 2025 (EDS budget year runs Dec 1 - Nov 30).

To run the report for a different district or budget year, edit the
constants in the "Report parameters" section below.

Usage:
    # Set DB creds in .env (DB_SERVER / DB_USERNAME / DB_PASSWORD), then:
    python scripts/colton_pierrepont_items_report.py

Requirements:
    pip install pyodbc openpyxl python-dotenv
    Microsoft ODBC Driver for SQL Server

The output workbook is saved under `output/` relative to the project root.
The header row has auto-filter and freeze-panes enabled so the district can
click any column's filter arrow to sort by User #, Item #, Vendor, Price,
etc. without touching the data.
"""

import os
import sys
from datetime import datetime
from decimal import Decimal

# Reuse the shared secure DB helper (reads DB_SERVER/DB_USERNAME/DB_PASSWORD
# from .env rather than hardcoding credentials in source control).
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from db_utils import DatabaseConnection, DatabaseConnectionError  # noqa: E402

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required.  pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Report parameters -- edit to reuse the script for another district / year
# ---------------------------------------------------------------------------
DISTRICT_CODE     = "7G"                         # Colton-Pierrepont
DISTRICT_LABEL    = "Colton-Pierrepont"
START_DATE        = "2024-12-01"                 # Inclusive (Dec 1)
END_DATE          = "2025-12-01"                 # Exclusive (= Nov 30 23:59)
BUDGET_YEAR_LABEL = "Dec 2024 - Nov 2025"

# Status filter: exclude requisitions that never became actual orders.
# StatusTable values: 1=On Hold, 2=Pending Approval, 4=Rejected.
# NULL statuses are also excluded -- they indicate uninitialized/draft
# requisitions that never reached an ordered state.
EXCLUDED_STATUSES = (1, 2, 4)


# ---------------------------------------------------------------------------
# SQL
# ---------------------------------------------------------------------------
# One row per active Detail line on a requisition entered during the budget
# year for the target district. Columns requested by Patrice:
#     User #, User Name, Item #, Vendor, Item Name, Price
# Supporting context included so the district can cross-reference in Excel:
#     Req #, Order Date, School, EDS Item Code, Qty, Line Total, Status
ITEMS_SQL = """
    SELECT
        r.RequisitionNumber                                    AS Requisition,
        CAST(r.DateEntered AS date)                            AS [Order Date],
        sch.Name                                               AS School,
        u.CometId                                              AS [User #],
        LTRIM(RTRIM(
            ISNULL(u.FirstName,'') + ' ' + ISNULL(u.LastName,'')
        ))                                                     AS [User Name],
        v.Code                                                 AS [Vendor Code],
        COALESCE(v.Name, 'N/A')                                AS Vendor,
        COALESCE(det.VendorItemCode, i.ItemCode, '')           AS [Item #],
        COALESCE(det.ItemCode, i.ItemCode, '')                 AS [EDS Item Code],
        COALESCE(det.Description, i.Description, '')           AS [Item Name],
        CAST(det.Quantity AS INT)                              AS Qty,
        COALESCE(det.BidPrice, det.CatalogPrice)               AS [Unit Price],
        CAST(det.Quantity AS money)
            * COALESCE(det.BidPrice, det.CatalogPrice)         AS [Line Total],
        ISNULL(st.Name, 'On Hold')                             AS Status
    FROM dbo.Requisitions r
    INNER JOIN dbo.Detail      det ON det.RequisitionId = r.RequisitionId
    INNER JOIN dbo.School      sch ON sch.SchoolId      = r.SchoolId
    INNER JOIN dbo.District    d   ON d.DistrictId      = sch.DistrictId
    LEFT  JOIN dbo.Users       u   ON u.UserId          = r.UserId
    LEFT  JOIN dbo.Vendors     v   ON v.VendorId        = det.VendorId
    LEFT  JOIN dbo.Items       i   ON i.ItemId          = det.ItemId
    LEFT  JOIN dbo.StatusTable st  ON st.StatusId       = r.StatusId
    WHERE d.DistrictCode = ?
      AND r.Active   = 1
      AND det.Active = 1
      AND r.DateEntered >= ?
      AND r.DateEntered <  ?
      AND r.StatusId IS NOT NULL
      AND r.StatusId NOT IN ({excluded})
    ORDER BY
        COALESCE(det.VendorItemCode, i.ItemCode, ''),
        v.Name,
        r.DateEntered,
        u.LastName,
        u.FirstName
""".format(excluded=",".join(str(s) for s in EXCLUDED_STATUSES))


# ---------------------------------------------------------------------------
# EDS brand styling
# ---------------------------------------------------------------------------
EDS_BLUE  = "1C1A83"
EDS_LIGHT = "4A4890"
EDS_RED   = "B70C0D"
WHITE     = "FFFFFF"
ALT_ROW   = "EBEBF5"    # light lavender zebra striping

MONEY_FMT = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'
QTY_FMT   = "#,##0"     # thousands-separated, for item quantities
ID_FMT    = "0"         # plain integer, for identifiers like User # / CometId
DATE_FMT  = "MM/DD/YYYY"

# 1-based column indexes for formatting
DATE_COLS     = {2}          # Order Date
ID_COLS       = {4}          # User # (no thousands separator -- it's an id)
INT_COLS      = {11}         # Qty (thousands-separated)
CURRENCY_COLS = {12, 13}     # Unit Price, Line Total


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def style_header_row(ws, row_idx, num_cols):
    hdr_font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    hdr_fill = PatternFill(start_color=EDS_BLUE, end_color=EDS_BLUE, fill_type="solid")
    bdr      = thin_border()
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.border    = bdr
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def shade_alt_rows(ws, first_data_row, last_data_row, num_cols):
    fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
    for row in range(first_data_row, last_data_row + 1):
        if (row - first_data_row) % 2 == 1:
            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).fill = fill


def run_report():
    print("Connecting to EDS...")

    # DatabaseConnection is a context manager -- it guarantees the
    # connection is closed even if we raise mid-query or mid-workbook build.
    try:
        with DatabaseConnection(database="EDS") as db:
            # Confirm the district exists (catches typos in DISTRICT_CODE).
            district_rows = db.execute_query(
                "SELECT DistrictId, DistrictCode, Name FROM dbo.District WHERE DistrictCode = ?",
                params=(DISTRICT_CODE,),
            )
            if not district_rows:
                raise SystemExit(f"District with code '{DISTRICT_CODE}' not found.")
            drow = district_rows[0]
            print(f"  District: [{drow.DistrictCode}] {drow.Name}  (DistrictId={drow.DistrictId})")

            print(f"Querying line items {START_DATE} through {END_DATE} (exclusive)...")
            # fetch_all=False returns a live cursor so we can read column names
            # from its description, then pull rows.
            cursor = db.execute_query(
                ITEMS_SQL,
                params=(DISTRICT_CODE, START_DATE, END_DATE),
                fetch_all=False,
            )
            try:
                col_names = [desc[0] for desc in cursor.description]
                rows = cursor.fetchall()
            finally:
                cursor.close()
    except DatabaseConnectionError as exc:
        raise SystemExit(f"Database connection failed: {exc}") from exc

    print(f"  Fetched {len(rows):,} line items.")

    # -----------------------------------------------------------------------
    # Build workbook
    # -----------------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Items Ordered"

    num_cols = len(col_names)
    last_col_letter = get_column_letter(num_cols)

    # ---- Title + subtitle ------------------------------------------------
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"{DISTRICT_LABEL} ({DISTRICT_CODE}) -- Items Ordered -- {BUDGET_YEAR_LABEL}"
    ws["A1"].font = Font(bold=True, size=14, color=EDS_BLUE, name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # Pre-compute aggregate totals for the subtitle using Decimal so the
    # displayed total matches cents exactly (no float rounding drift).
    line_total_idx = col_names.index("Line Total")
    total_spend: Decimal = sum(
        (r[line_total_idx] for r in rows if r[line_total_idx] is not None),
        start=Decimal("0"),
    )

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = (
        f"{len(rows):,} line items   |   "
        f"Total: ${total_spend:,.2f}   |   "
        f"Generated {datetime.now().strftime('%B %d, %Y  %I:%M %p')}"
    )
    ws["A2"].font = Font(italic=True, size=10, color="666666", name="Calibri")

    ws.merge_cells(f"A3:{last_col_letter}3")
    ws["A3"] = "Click the filter arrows in row 4 to sort/filter by any column."
    ws["A3"].font = Font(italic=True, size=10, color=EDS_LIGHT, name="Calibri")

    # ---- Header row (row 4) ---------------------------------------------
    for col_idx, header in enumerate(col_names, start=1):
        ws.cell(row=4, column=col_idx, value=header)
    style_header_row(ws, row_idx=4, num_cols=num_cols)
    ws.row_dimensions[4].height = 32

    # ---- Data rows (start at row 5) -------------------------------------
    first_data_row = 5
    bdr = thin_border()
    for r_offset, row in enumerate(rows):
        excel_row = first_data_row + r_offset
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=c_idx)
            cell.border = bdr
            cell.font   = Font(size=10, name="Calibri")

            if value is None:
                cell.value = None
            elif c_idx in CURRENCY_COLS:
                # Assign Decimal directly -- openpyxl handles it and no
                # Decimal->float rounding error is introduced.
                cell.value = value
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx in ID_COLS:
                cell.value = int(value)
                cell.number_format = ID_FMT
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx in INT_COLS:
                cell.value = int(value)
                cell.number_format = QTY_FMT
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx in DATE_COLS:
                cell.value = value
                cell.number_format = DATE_FMT
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    last_data_row = first_data_row + len(rows) - 1 if rows else 4
    if rows:
        shade_alt_rows(ws, first_data_row, last_data_row, num_cols)

    # ---- Totals row ------------------------------------------------------
    if rows:
        totals_row = last_data_row + 1
        label_cell = ws.cell(row=totals_row, column=1, value="TOTAL")
        label_cell.font = Font(bold=True, size=11, color=EDS_RED, name="Calibri")
        label_cell.alignment = Alignment(horizontal="right")

        qty_letter   = get_column_letter(11)  # Qty
        total_letter = get_column_letter(13)  # Line Total

        qty_sum = ws.cell(row=totals_row, column=11)
        qty_sum.value = f"=SUM({qty_letter}{first_data_row}:{qty_letter}{last_data_row})"
        qty_sum.number_format = QTY_FMT
        qty_sum.font = Font(bold=True, size=11, name="Calibri")
        qty_sum.alignment = Alignment(horizontal="right")
        qty_sum.border = Border(top=Side(style="double"))

        total_sum = ws.cell(row=totals_row, column=13)
        total_sum.value = f"=SUM({total_letter}{first_data_row}:{total_letter}{last_data_row})"
        total_sum.number_format = MONEY_FMT
        total_sum.font = Font(bold=True, size=11, name="Calibri")
        total_sum.alignment = Alignment(horizontal="right")
        total_sum.border = Border(top=Side(style="double"))

    # ---- Column widths ---------------------------------------------------
    widths = {
        "Requisition":     14,
        "Order Date":      12,
        "School":          28,
        "User #":          10,
        "User Name":       24,
        "Vendor Code":     12,
        "Vendor":          30,
        "Item #":          18,
        "EDS Item Code":   14,
        "Item Name":       48,
        "Qty":              8,
        "Unit Price":      13,
        "Line Total":      14,
        "Status":          18,
    }
    for col_idx, header in enumerate(col_names, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 14)

    # ---- Freeze + auto-filter -------------------------------------------
    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A4:{last_col_letter}{last_data_row}"

    # -----------------------------------------------------------------------
    # Save
    # -----------------------------------------------------------------------
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_dir = os.path.join(project_root, "output")
    os.makedirs(out_dir, exist_ok=True)

    safe_year = BUDGET_YEAR_LABEL.replace(" ", "_").replace("-", "").replace(",", "")
    out_path = os.path.join(
        out_dir,
        f"{DISTRICT_LABEL.replace(' ', '_').replace('-', '_')}"
        f"_{DISTRICT_CODE}_Items_Ordered_{safe_year}.xlsx",
    )
    wb.save(out_path)
    print(f"\n[OK] Report saved to: {out_path}")
    print(f"     {len(rows):,} rows   |   Total ${total_spend:,.2f}")
    return out_path


if __name__ == "__main__":
    run_report()
