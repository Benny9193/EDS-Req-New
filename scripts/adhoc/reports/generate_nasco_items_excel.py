"""
Generate Nasco Items-by-Bid Excel report from pre-queried data.
Aggregates PODetailItems by Bid + Item, with bid-level summary sheet.
"""
import pickle, os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# --- Load data ---
with open(r"C:\EDS\nasco_items_raw3.pkl", "rb") as f:
    data = pickle.load(f)

pdi_df = pd.DataFrame(data["pdi"]["rows"], columns=data["pdi"]["cols"])
items_df = pd.DataFrame(data["items"]["rows"], columns=data["items"]["cols"])
awards_df = pd.DataFrame(data["awards"]["rows"], columns=data["awards"]["cols"])

# --- Merge ---
df = pdi_df.merge(items_df, on="ItemId", how="left")
df = df.merge(awards_df, on="AwardId", how="left")

# Compute line amount
df["LineAmount"] = df["Quantity"] * df["BidPrice"]

# --- Aggregate: items grouped by bid + item ---
detail_agg = df.groupby([
    "CategoryName", "CategoryCode",
    "BidId", "BidName", "VendorBidNumber",
    "BidHeaderId", "BidHeaderDescription",
    "EffectiveFrom", "EffectiveUntil",
    "ItemCode", "VendorItemCode", "Description"
]).agg(
    TotalQtyOrdered=("Quantity", "sum"),
    MinBidPrice=("BidPrice", "min"),
    MaxBidPrice=("BidPrice", "max"),
    TotalAmount=("LineAmount", "sum"),
    POCount=("POId", "nunique"),
).reset_index()

detail_agg = detail_agg.sort_values(
    ["CategoryName", "BidName", "TotalAmount"],
    ascending=[True, True, False]
)

# --- Aggregate: bid-level summary ---
bid_agg = df.groupby([
    "CategoryName", "CategoryCode",
    "BidId", "BidName", "VendorBidNumber",
    "BidHeaderId", "BidHeaderDescription",
    "EffectiveFrom", "EffectiveUntil",
]).agg(
    UniqueItems=("ItemCode", "nunique"),
    TotalQtyOrdered=("Quantity", "sum"),
    TotalAmount=("LineAmount", "sum"),
    POCount=("POId", "nunique"),
    FirstPODate=("PODate", "min"),
    LastPODate=("PODate", "max"),
).reset_index()

bid_agg = bid_agg.sort_values(
    ["CategoryName", "TotalAmount"],
    ascending=[True, False]
)

print(f"Detail rows: {len(detail_agg)}")
print(f"Bid summary rows: {len(bid_agg)}")

# ===================== EXCEL GENERATION =====================
EDS_PRIMARY   = "1C1A83"
EDS_SECONDARY = "4A4890"
EDS_ACCENT    = "B70C0D"

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color=EDS_PRIMARY, end_color=EDS_PRIMARY, fill_type="solid")
subheader_fill = PatternFill(start_color=EDS_SECONDARY, end_color=EDS_SECONDARY, fill_type="solid")
accent_font = Font(name="Calibri", bold=True, color=EDS_ACCENT, size=11)
data_font = Font(name="Calibri", size=10)
money_fmt = '#,##0.00'
qty_fmt = '#,##0'
date_fmt = 'MM/DD/YYYY'
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

wb = Workbook()

# ---- Sheet 1: Bid Summary ----
ws1 = wb.active
ws1.title = "Bid Summary"

# Title row
ws1.merge_cells("A1:N1")
ws1["A1"].value = "Nasco Education (0518) — Items Ordered by Bid — Dec 2024 – Nov 2025"
ws1["A1"].font = Font(name="Calibri", bold=True, color=EDS_PRIMARY, size=14)
ws1["A1"].alignment = Alignment(horizontal="left")

ws1.merge_cells("A2:N2")
ws1["A2"].value = f"Bid-Level Summary  |  {len(bid_agg)} bids  |  {df['LineAmount'].sum():,.2f} total spend"
ws1["A2"].font = Font(name="Calibri", italic=True, color="666666", size=10)

bid_headers = [
    ("Category", 22), ("Cat Code", 10), ("Bid ID", 8), ("Bid Name", 35),
    ("Vendor Bid #", 14), ("Bid Header ID", 14), ("Bid Header Desc", 30),
    ("Effective From", 14), ("Effective Until", 14),
    ("Unique Items", 12), ("Total Qty", 12), ("Total Amount", 16),
    ("PO Count", 10), ("First PO Date", 14), ("Last PO Date", 14),
]

for col_idx, (hdr, width) in enumerate(bid_headers, 1):
    cell = ws1.cell(row=4, column=col_idx, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border
    ws1.column_dimensions[get_column_letter(col_idx)].width = width

for row_idx, row in enumerate(bid_agg.itertuples(index=False), 5):
    vals = [
        row.CategoryName, row.CategoryCode, row.BidId, row.BidName,
        row.VendorBidNumber, row.BidHeaderId, row.BidHeaderDescription,
        row.EffectiveFrom, row.EffectiveUntil,
        row.UniqueItems, row.TotalQtyOrdered, row.TotalAmount,
        row.POCount, row.FirstPODate, row.LastPODate,
    ]
    for col_idx, val in enumerate(vals, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        if col_idx == 12:  # TotalAmount
            cell.number_format = money_fmt
        elif col_idx in (10, 11, 13):  # counts/qty
            cell.number_format = qty_fmt
        elif col_idx in (14, 15):  # dates
            cell.number_format = date_fmt
        if row_idx % 2 == 0:
            cell.fill = PatternFill(start_color="F2F2F7", end_color="F2F2F7", fill_type="solid")

# Totals row
total_row = 5 + len(bid_agg)
ws1.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=11, color=EDS_ACCENT)
for col_idx in (10, 11, 12, 13):
    cell = ws1.cell(row=total_row, column=col_idx)
    col_letter = get_column_letter(col_idx)
    cell.value = f"=SUM({col_letter}5:{col_letter}{total_row-1})"
    cell.font = Font(name="Calibri", bold=True, size=11)
    cell.border = Border(top=Side(style="double"))
    if col_idx == 12:
        cell.number_format = money_fmt
    else:
        cell.number_format = qty_fmt

ws1.freeze_panes = "A5"
ws1.auto_filter.ref = f"A4:O{total_row - 1}"

# ---- Sheet 2: Item Detail ----
ws2 = wb.create_sheet("Item Detail by Bid")

ws2.merge_cells("A1:Q1")
ws2["A1"].value = "Nasco Education (0518) — Item Detail by Bid — Dec 2024 – Nov 2025"
ws2["A1"].font = Font(name="Calibri", bold=True, color=EDS_PRIMARY, size=14)

ws2.merge_cells("A2:Q2")
ws2["A2"].value = f"Item-Level Detail  |  {len(detail_agg)} item-bid combinations"
ws2["A2"].font = Font(name="Calibri", italic=True, color="666666", size=10)

detail_headers = [
    ("Category", 22), ("Cat Code", 10), ("Bid ID", 8), ("Bid Name", 35),
    ("Vendor Bid #", 14), ("Bid Header ID", 14), ("Bid Header Desc", 30),
    ("Effective From", 14), ("Effective Until", 14),
    ("Item Code", 14), ("Vendor Item Code", 16), ("Item Description", 40),
    ("Total Qty", 12), ("Min Price", 12), ("Max Price", 12),
    ("Total Amount", 16), ("PO Count", 10),
]

for col_idx, (hdr, width) in enumerate(detail_headers, 1):
    cell = ws2.cell(row=4, column=col_idx, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

for row_idx, row in enumerate(detail_agg.itertuples(index=False), 5):
    vals = [
        row.CategoryName, row.CategoryCode, row.BidId, row.BidName,
        row.VendorBidNumber, row.BidHeaderId, row.BidHeaderDescription,
        row.EffectiveFrom, row.EffectiveUntil,
        row.ItemCode, row.VendorItemCode, row.Description,
        row.TotalQtyOrdered, row.MinBidPrice, row.MaxBidPrice,
        row.TotalAmount, row.POCount,
    ]
    for col_idx, val in enumerate(vals, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        if col_idx in (14, 15, 16):  # money
            cell.number_format = money_fmt
        elif col_idx in (13, 17):  # qty/count
            cell.number_format = qty_fmt
        if row_idx % 2 == 0:
            cell.fill = PatternFill(start_color="F2F2F7", end_color="F2F2F7", fill_type="solid")

# Totals
total_row2 = 5 + len(detail_agg)
ws2.cell(row=total_row2, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=11, color=EDS_ACCENT)
for col_idx in (13, 16, 17):
    cell = ws2.cell(row=total_row2, column=col_idx)
    col_letter = get_column_letter(col_idx)
    cell.value = f"=SUM({col_letter}5:{col_letter}{total_row2-1})"
    cell.font = Font(name="Calibri", bold=True, size=11)
    cell.border = Border(top=Side(style="double"))
    if col_idx == 16:
        cell.number_format = money_fmt
    else:
        cell.number_format = qty_fmt

ws2.freeze_panes = "A5"
ws2.auto_filter.ref = f"A4:Q{total_row2 - 1}"

# ---- Save ----
out_dir = r"C:\EDS\output"
os.makedirs(out_dir, exist_ok=True)
out_path = os.path.join(out_dir, "Nasco_0518_Items_by_Bid_Dec2024_Nov2025.xlsx")
wb.save(out_path)
print(f"\nSaved: {out_path}", flush=True)
print(f"  Bid Summary: {len(bid_agg)} rows", flush=True)
print(f"  Item Detail: {len(detail_agg)} rows", flush=True)
print(f"  Total Spend: ${df['LineAmount'].sum():,.2f}", flush=True)
