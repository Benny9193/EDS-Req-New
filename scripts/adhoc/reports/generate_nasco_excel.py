import pickle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

with open("nasco_data.pkl", "rb") as f:
    data = pickle.load(f)

detail_rows, detail_cols = data["detail"]
cat_rows, cat_cols = data["cats"]

wb = Workbook()

# ── Styles ─────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", start_color="1C1A83")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(name="Arial", bold=True, size=14, color="1C1A83")
BODY_FONT   = Font(name="Arial", size=10)
SUBTTL_FILL = PatternFill("solid", start_color="4A4890")
SUBTTL_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TOTAL_FILL  = PatternFill("solid", start_color="B70C0D")
TOTAL_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", start_color="F0F0FA")
TITLE_BG    = PatternFill("solid", start_color="F0F0FB")
MONEY_FMT   = '$#,##0.00'
DATE_FMT    = 'MM/DD/YYYY'
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")


def hdr(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER


def money_cell(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=float(val) if val else 0.0)
    c.number_format = MONEY_FMT
    c.alignment = RIGHT
    c.font = BODY_FONT
    c.border = BORDER


def int_cell(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=int(val) if val else 0)
    c.number_format = "#,##0"
    c.alignment = RIGHT
    c.font = BODY_FONT
    c.border = BORDER


def date_cell(ws, row, col, val):
    if val is None:
        c = ws.cell(row=row, column=col, value="")
    else:
        c = ws.cell(row=row, column=col,
                    value=val.date() if hasattr(val, "date") else val)
        c.number_format = DATE_FMT
    c.alignment = CENTER
    c.font = BODY_FONT
    c.border = BORDER


def text_cell(ws, row, col, val, bold=False, fill=None):
    c = ws.cell(row=row, column=col, value=str(val) if val else "")
    c.font = Font(name="Arial", size=10, bold=bold)
    c.alignment = LEFT
    c.border = BORDER
    if fill:
        c.fill = fill


def apply_fill_row(ws, row, ncols, fill):
    for ci in range(1, ncols + 1):
        ws.cell(row, ci).fill = fill


# ═══════════════════════════════════════════════════════════════════════
# SHEET 1 – Category Summary
# ═══════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Category Summary"
ws1.sheet_view.showGridLines = False

ws1.merge_cells("A1:G1")
ws1["A1"].value = "NASCO (Vendor 0518) – Purchase Orders by Category  |  Dec 1, 2024 – Nov 30, 2025"
ws1["A1"].font  = TITLE_FONT
ws1["A1"].alignment = CENTER
ws1["A1"].fill  = TITLE_BG

ws1.merge_cells("A2:G2")
ws1["A2"].value = (
    f"School Specialty, LLC dba Nasco Education  |  "
    f"Generated: {datetime.now().strftime('%B %d, %Y')}"
)
ws1["A2"].font = Font(name="Arial", size=10, italic=True, color="555555")
ws1["A2"].alignment = CENTER
ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 18
ws1.row_dimensions[3].height = 6

hdrs1 = ["Category", "Code", "Unique Bids", "Total POs",
         "Total PO Amount", "First PO Date", "Last PO Date"]
for ci, h in enumerate(hdrs1, 1):
    hdr(ws1, 4, ci, h)
ws1.row_dimensions[4].height = 20

for ri, row in enumerate(cat_rows, 5):
    d = dict(zip(cat_cols, row))
    use_fill = ALT_FILL if ri % 2 == 0 else None
    text_cell(ws1, ri, 1, d["CategoryName"], fill=use_fill)
    text_cell(ws1, ri, 2, d["CategoryCode"], fill=use_fill)
    int_cell(ws1, ri, 3, d["BidCount"])
    int_cell(ws1, ri, 4, d["POCount"])
    money_cell(ws1, ri, 5, d["TotalAmount"])
    date_cell(ws1, ri, 6, d["FirstPODate"])
    date_cell(ws1, ri, 7, d["LastPODate"])
    if use_fill:
        apply_fill_row(ws1, ri, 7, use_fill)
    ws1.row_dimensions[ri].height = 16

last_data = 4 + len(cat_rows)
tr = last_data + 1
ws1.row_dimensions[tr].height = 22
for ci in range(1, 8):
    c = ws1.cell(tr, ci)
    c.fill = TOTAL_FILL
    c.font = TOTAL_FONT
    c.border = BORDER

total_bids = sum(r[2] for r in cat_rows)
total_pos  = sum(r[3] for r in cat_rows)
total_amt  = sum(float(r[4]) for r in cat_rows if r[4])

ws1.cell(tr, 1).value = "GRAND TOTAL"
ws1.cell(tr, 1).alignment = LEFT
ws1.cell(tr, 3).value = total_bids
ws1.cell(tr, 3).number_format = "#,##0"
ws1.cell(tr, 3).alignment = RIGHT
ws1.cell(tr, 4).value = total_pos
ws1.cell(tr, 4).number_format = "#,##0"
ws1.cell(tr, 4).alignment = RIGHT
ws1.cell(tr, 5).value = total_amt
ws1.cell(tr, 5).number_format = MONEY_FMT
ws1.cell(tr, 5).alignment = RIGHT

ws1.column_dimensions["A"].width = 36
ws1.column_dimensions["B"].width = 8
ws1.column_dimensions["C"].width = 14
ws1.column_dimensions["D"].width = 12
ws1.column_dimensions["E"].width = 18
ws1.column_dimensions["F"].width = 14
ws1.column_dimensions["G"].width = 14
ws1.freeze_panes = "A5"

# ═══════════════════════════════════════════════════════════════════════
# SHEET 2 – Detail by Category & Bid
# ═══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Detail by Category & Bid")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:J1")
ws2["A1"].value = "NASCO (Vendor 0518) – Purchase Orders by Category and Bid  |  Dec 1, 2024 – Nov 30, 2025"
ws2["A1"].font  = TITLE_FONT
ws2["A1"].alignment = CENTER
ws2["A1"].fill  = TITLE_BG

ws2.merge_cells("A2:J2")
ws2["A2"].value = (
    f"School Specialty, LLC dba Nasco Education  |  "
    f"Generated: {datetime.now().strftime('%B %d, %Y')}"
)
ws2["A2"].font = Font(name="Arial", size=10, italic=True, color="555555")
ws2["A2"].alignment = CENTER
ws2.row_dimensions[1].height = 30
ws2.row_dimensions[2].height = 18
ws2.row_dimensions[3].height = 6

hdrs2 = [
    "Category", "Cat Code", "Bid ID", "Bid Name",
    "Vendor Bid #", "Bid Header Key", "Effective From",
    "Effective Until", "PO Count", "Total PO Amount"
]
for ci, h in enumerate(hdrs2, 1):
    hdr(ws2, 4, ci, h)
ws2.row_dimensions[4].height = 20

current_cat = None
row_num = 5
cat_po_count = 0
cat_amount = 0.0
grand_po_count = 0
grand_amount = 0.0

for dr in detail_rows:
    d = dict(zip(detail_cols, dr))
    cat = d["CategoryName"]

    if cat != current_cat:
        if current_cat is not None:
            st = row_num
            ws2.row_dimensions[st].height = 17
            for ci in range(1, 11):
                c = ws2.cell(st, ci)
                c.fill = SUBTTL_FILL
                c.font = SUBTTL_FONT
                c.border = BORDER
            ws2.cell(st, 1).value = f"  Subtotal – {current_cat}"
            ws2.cell(st, 1).alignment = LEFT
            ws2.cell(st, 9).value = cat_po_count
            ws2.cell(st, 9).number_format = "#,##0"
            ws2.cell(st, 9).alignment = RIGHT
            ws2.cell(st, 10).value = cat_amount
            ws2.cell(st, 10).number_format = MONEY_FMT
            ws2.cell(st, 10).alignment = RIGHT
            row_num += 1

        current_cat = cat
        cat_po_count = 0
        cat_amount = 0.0

    po_cnt = int(d["POCount"]) if d["POCount"] else 0
    amt = float(d["TotalAmount"]) if d["TotalAmount"] else 0.0
    cat_po_count += po_cnt
    cat_amount += amt
    grand_po_count += po_cnt
    grand_amount += amt

    use_fill = ALT_FILL if row_num % 2 == 0 else None
    text_cell(ws2, row_num, 1, d["CategoryName"], fill=use_fill)
    text_cell(ws2, row_num, 2, d["CategoryCode"], fill=use_fill)
    text_cell(ws2, row_num, 3, str(d["BidId"]), fill=use_fill)
    text_cell(ws2, row_num, 4, d["BidName"] or "", fill=use_fill)
    text_cell(ws2, row_num, 5, d["BidVendorNumber"] or "", fill=use_fill)
    text_cell(ws2, row_num, 6, d["BidHeaderKey"] or "", fill=use_fill)
    date_cell(ws2, row_num, 7, d["EffectiveFrom"])
    date_cell(ws2, row_num, 8, d["EffectiveUntil"])
    int_cell(ws2, row_num, 9, po_cnt)
    money_cell(ws2, row_num, 10, amt)
    if use_fill:
        apply_fill_row(ws2, row_num, 10, use_fill)
    ws2.row_dimensions[row_num].height = 15
    row_num += 1

# last category subtotal
st = row_num
ws2.row_dimensions[st].height = 17
for ci in range(1, 11):
    c = ws2.cell(st, ci)
    c.fill = SUBTTL_FILL
    c.font = SUBTTL_FONT
    c.border = BORDER
ws2.cell(st, 1).value = f"  Subtotal – {current_cat}"
ws2.cell(st, 1).alignment = LEFT
ws2.cell(st, 9).value = cat_po_count
ws2.cell(st, 9).number_format = "#,##0"
ws2.cell(st, 9).alignment = RIGHT
ws2.cell(st, 10).value = cat_amount
ws2.cell(st, 10).number_format = MONEY_FMT
ws2.cell(st, 10).alignment = RIGHT
row_num += 1

# Grand total
gt = row_num
ws2.row_dimensions[gt].height = 22
for ci in range(1, 11):
    c = ws2.cell(gt, ci)
    c.fill = TOTAL_FILL
    c.font = TOTAL_FONT
    c.border = BORDER
ws2.cell(gt, 1).value = "GRAND TOTAL"
ws2.cell(gt, 1).alignment = LEFT
ws2.cell(gt, 9).value = grand_po_count
ws2.cell(gt, 9).number_format = "#,##0"
ws2.cell(gt, 9).alignment = RIGHT
ws2.cell(gt, 10).value = grand_amount
ws2.cell(gt, 10).number_format = MONEY_FMT
ws2.cell(gt, 10).alignment = RIGHT

ws2.column_dimensions["A"].width = 34
ws2.column_dimensions["B"].width = 9
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 16
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 16
ws2.column_dimensions["G"].width = 14
ws2.column_dimensions["H"].width = 14
ws2.column_dimensions["I"].width = 12
ws2.column_dimensions["J"].width = 18
ws2.freeze_panes = "A5"

os.makedirs("output", exist_ok=True)
outfile = "output/Nasco_0518_PO_by_Category_and_Bid_Dec2024_Nov2025.xlsx"
wb.save(outfile)
print(f"Saved: {outfile}")
