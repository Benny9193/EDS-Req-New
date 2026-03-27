"""Generate Excel report: Bids/Trades by County with Vendor Bid Counts (Category Type = 3)"""
import pyodbc, os
from collections import defaultdict
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

load_dotenv()
server = os.getenv('DB_SERVER')
database = 'EDS'
username = os.getenv('DB_USERNAME')
password = os.getenv('DB_PASSWORD')
conn_str = f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'
conn = pyodbc.connect(conn_str)
cursor = conn.cursor()

# ---- QUERY 1: Full detail with state ----
detail_query = """
SELECT
    s.Code AS StateCode,
    s.Name AS StateName,
    c.Name AS CountyName,
    t.Description AS TradeName,
    bh.Description AS BidDescription,
    bh.BidHeaderId,
    CONVERT(varchar, bh.EffectiveFrom, 101) AS EffectiveFrom,
    CONVERT(varchar, bh.EffectiveUntil, 101) AS EffectiveUntil,
    COUNT(DISTINCT bi.VendorId) AS VendorBidCount
FROM BidHeaders bh WITH (NOLOCK)
JOIN Category cat WITH (NOLOCK) ON cat.CategoryId = bh.CategoryId AND cat.Type = 3
JOIN States s WITH (NOLOCK) ON s.StateId = bh.StateId
JOIN BidTrades bt WITH (NOLOCK) ON bt.BidHeaderId = bh.BidHeaderId
JOIN BidTradeCounties btc WITH (NOLOCK) ON btc.BidTradeId = bt.BidTradeId
JOIN Counties c WITH (NOLOCK) ON c.CountyId = btc.CountyId AND c.StateId = s.StateId
JOIN Trades t WITH (NOLOCK) ON t.TradeId = bt.TradeId AND t.Active = 1
LEFT JOIN BidImportCounties bic WITH (NOLOCK) ON bic.BidTradeCountyId = btc.BidTradeCountyId AND bic.Active = 1
LEFT JOIN BidImports bi WITH (NOLOCK) ON bi.BidImportId = bic.BidImportId AND bi.Active = 1
WHERE bh.Active = 1
  AND GETDATE() BETWEEN bh.EffectiveFrom AND bh.EffectiveUntil
GROUP BY s.Code, s.Name, c.Name, t.Description, bh.Description, bh.BidHeaderId,
         bh.EffectiveFrom, bh.EffectiveUntil
ORDER BY s.Code, c.Name, t.Description
"""
cursor.execute(detail_query)
detail_rows = cursor.fetchall()
print(f"Detail rows: {len(detail_rows)}")

# ---- QUERY 2: Summary by State + County ----
summary_query = """
SELECT
    s.Code AS StateCode,
    s.Name AS StateName,
    c.Name AS CountyName,
    COUNT(DISTINCT bt.BidTradeId) AS TotalBidTrades,
    COUNT(DISTINCT bi.VendorId) AS UniqueVendorsBid,
    COUNT(DISTINCT bh.BidHeaderId) AS BidHeaders
FROM BidHeaders bh WITH (NOLOCK)
JOIN Category cat WITH (NOLOCK) ON cat.CategoryId = bh.CategoryId AND cat.Type = 3
JOIN States s WITH (NOLOCK) ON s.StateId = bh.StateId
JOIN BidTrades bt WITH (NOLOCK) ON bt.BidHeaderId = bh.BidHeaderId
JOIN BidTradeCounties btc WITH (NOLOCK) ON btc.BidTradeId = bt.BidTradeId
JOIN Counties c WITH (NOLOCK) ON c.CountyId = btc.CountyId AND c.StateId = s.StateId
JOIN Trades t WITH (NOLOCK) ON t.TradeId = bt.TradeId AND t.Active = 1
LEFT JOIN BidImportCounties bic WITH (NOLOCK) ON bic.BidTradeCountyId = btc.BidTradeCountyId AND bic.Active = 1
LEFT JOIN BidImports bi WITH (NOLOCK) ON bi.BidImportId = bic.BidImportId AND bi.Active = 1
WHERE bh.Active = 1
  AND GETDATE() BETWEEN bh.EffectiveFrom AND bh.EffectiveUntil
GROUP BY s.Code, s.Name, c.Name
ORDER BY s.Code, c.Name
"""
cursor.execute(summary_query)
summary_rows = cursor.fetchall()
print(f"Summary rows: {len(summary_rows)}")

# ---- QUERY 3: State totals ----
state_query = """
SELECT
    s.Code AS StateCode,
    s.Name AS StateName,
    COUNT(DISTINCT c.CountyId) AS Counties,
    COUNT(DISTINCT bt.BidTradeId) AS TotalBidTrades,
    COUNT(DISTINCT bh.BidHeaderId) AS BidHeaders,
    COUNT(DISTINCT bi.VendorId) AS UniqueVendorsBid
FROM BidHeaders bh WITH (NOLOCK)
JOIN Category cat WITH (NOLOCK) ON cat.CategoryId = bh.CategoryId AND cat.Type = 3
JOIN States s WITH (NOLOCK) ON s.StateId = bh.StateId
JOIN BidTrades bt WITH (NOLOCK) ON bt.BidHeaderId = bh.BidHeaderId
JOIN BidTradeCounties btc WITH (NOLOCK) ON btc.BidTradeId = bt.BidTradeId
JOIN Counties c WITH (NOLOCK) ON c.CountyId = btc.CountyId AND c.StateId = s.StateId
JOIN Trades t WITH (NOLOCK) ON t.TradeId = bt.TradeId AND t.Active = 1
LEFT JOIN BidImportCounties bic WITH (NOLOCK) ON bic.BidTradeCountyId = btc.BidTradeCountyId AND bic.Active = 1
LEFT JOIN BidImports bi WITH (NOLOCK) ON bi.BidImportId = bic.BidImportId AND bi.Active = 1
WHERE bh.Active = 1
  AND GETDATE() BETWEEN bh.EffectiveFrom AND bh.EffectiveUntil
GROUP BY s.Code, s.Name
ORDER BY s.Code
"""
cursor.execute(state_query)
state_rows = cursor.fetchall()
for r in state_rows:
    print(f"  {r.StateCode} - {r.StateName}: {r.Counties} counties, {r.TotalBidTrades} trades, {r.UniqueVendorsBid} vendors")

# ---- QUERY 4: Vendor-level drill-down ----
vendor_query = """
SELECT
    s.Code AS StateCode,
    c.Name AS CountyName,
    t.Description AS TradeName,
    v.Name AS VendorName,
    v.VendorId,
    bh.Description AS BidDescription,
    bh.BidHeaderId,
    CONVERT(varchar, bh.EffectiveFrom, 101) AS EffectiveFrom,
    CONVERT(varchar, bh.EffectiveUntil, 101) AS EffectiveUntil
FROM BidHeaders bh WITH (NOLOCK)
JOIN Category cat WITH (NOLOCK) ON cat.CategoryId = bh.CategoryId AND cat.Type = 3
JOIN States s WITH (NOLOCK) ON s.StateId = bh.StateId
JOIN BidTrades bt WITH (NOLOCK) ON bt.BidHeaderId = bh.BidHeaderId
JOIN BidTradeCounties btc WITH (NOLOCK) ON btc.BidTradeId = bt.BidTradeId
JOIN Counties c WITH (NOLOCK) ON c.CountyId = btc.CountyId AND c.StateId = s.StateId
JOIN Trades t WITH (NOLOCK) ON t.TradeId = bt.TradeId AND t.Active = 1
JOIN BidImportCounties bic WITH (NOLOCK) ON bic.BidTradeCountyId = btc.BidTradeCountyId AND bic.Active = 1
JOIN BidImports bi WITH (NOLOCK) ON bi.BidImportId = bic.BidImportId AND bi.Active = 1
JOIN Vendors v WITH (NOLOCK) ON v.VendorId = bi.VendorId
WHERE bh.Active = 1
  AND GETDATE() BETWEEN bh.EffectiveFrom AND bh.EffectiveUntil
ORDER BY s.Code, c.Name, t.Description, v.Name
"""
cursor.execute(vendor_query)
vendor_rows = cursor.fetchall()
print(f"Vendor drill-down rows: {len(vendor_rows)}")

conn.close()

# ---- BUILD EXCEL ----
wb = Workbook()

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1C1A83", end_color="1C1A83", fill_type="solid")
nj_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
ny_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
total_font = Font(bold=True, size=11)
total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
zero_font = Font(color="C0C0C0")


def style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border


def auto_width(ws, col_count, max_rows=100):
    for col in range(1, col_count + 1):
        max_len = 0
        for row in range(1, min(max_rows + 1, ws.max_row + 1)):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 55)


def get_fill(state_code):
    return nj_fill if state_code == "NJ" else ny_fill


# ============ SHEET 1: State Summary ============
ws1 = wb.active
ws1.title = "State Summary"
headers1 = ["State", "State Name", "Counties", "Bid/Trades", "Bid Headers", "Unique Vendors Who Bid"]
for i, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=i, value=h)
style_header(ws1, len(headers1))

for r_idx, row in enumerate(state_rows, 2):
    vals = [row.StateCode, row.StateName, row.Counties, row.TotalBidTrades, row.BidHeaders, row.UniqueVendorsBid]
    fill = get_fill(row.StateCode)
    for c, v in enumerate(vals, 1):
        cell = ws1.cell(row=r_idx, column=c, value=v)
        cell.fill = fill
        cell.border = thin_border

# Totals row
total_row = len(state_rows) + 2
ws1.cell(row=total_row, column=1, value="TOTAL").font = total_font
ws1.cell(row=total_row, column=3, value=sum(r.Counties for r in state_rows)).font = total_font
ws1.cell(row=total_row, column=4, value=sum(r.TotalBidTrades for r in state_rows)).font = total_font
ws1.cell(row=total_row, column=5, value=sum(r.BidHeaders for r in state_rows)).font = total_font
for c in range(1, len(headers1) + 1):
    ws1.cell(row=total_row, column=c).fill = total_fill
    ws1.cell(row=total_row, column=c).border = thin_border

auto_width(ws1, len(headers1))
ws1.freeze_panes = "A2"

# ============ SHEET 2: County Summary ============
ws2 = wb.create_sheet("County Summary")
headers2 = ["State", "County", "Bid/Trades", "Unique Vendors Who Bid", "Bid Headers"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, len(headers2))

for r_idx, row in enumerate(summary_rows, 2):
    vals = [row.StateCode, row.CountyName, row.TotalBidTrades, row.UniqueVendorsBid, row.BidHeaders]
    fill = get_fill(row.StateCode)
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=r_idx, column=c, value=v)
        cell.fill = fill
        cell.border = thin_border

auto_width(ws2, len(headers2))
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:E{len(summary_rows) + 1}"

# ============ DETAIL SHEET HELPER ============
detail_headers = ["State", "County", "Trade", "Bid Description", "BidHeaderId", "Effective From", "Effective Until", "Vendors Who Bid"]


def write_detail_sheet(ws, rows_to_write, state_filter=None):
    for i, h in enumerate(detail_headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(detail_headers))

    r_idx = 2
    for row in rows_to_write:
        if state_filter and row.StateCode != state_filter:
            continue
        vals = [row.StateCode, row.CountyName, row.TradeName, row.BidDescription,
                row.BidHeaderId, row.EffectiveFrom, row.EffectiveUntil, row.VendorBidCount]
        fill = get_fill(row.StateCode)
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            cell.fill = fill
            cell.border = thin_border
        if row.VendorBidCount == 0:
            ws.cell(row=r_idx, column=8).font = zero_font
        r_idx += 1

    auto_width(ws, len(detail_headers), max_rows=50)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{r_idx - 1}"
    return r_idx - 2


# ============ PIVOT SHEET HELPER ============
# Color scale: white (0) -> light green -> dark green (max)
def get_heatmap_color(value, max_val):
    """Return a fill color on a white-to-green gradient based on value."""
    if value == 0 or max_val == 0:
        return None  # no fill for zeros
    ratio = min(value / max_val, 1.0)
    # Interpolate from white (255,255,255) toward dark green (39,125,60)
    r = int(255 - ratio * (255 - 39))
    g = int(255 - ratio * (255 - 174))
    b = int(255 - ratio * (255 - 96))
    hex_color = f"{r:02X}{g:02X}{b:02X}"
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def get_font_for_heatmap(value, max_val):
    """Return white font for dark cells, black for light cells."""
    if max_val == 0 or value == 0:
        return Font(color="C0C0C0", size=9)  # gray for zeros
    ratio = value / max_val
    if ratio > 0.6:
        return Font(color="FFFFFF", size=9, bold=True)
    return Font(size=9)


def write_pivot_sheet(ws, detail_data, state_filter=None, sheet_label=""):
    """Build a pivot: trades (rows) x counties (columns) with vendor bid counts."""
    # Filter data
    filtered = [r for r in detail_data if (state_filter is None or r.StateCode == state_filter)]
    if not filtered:
        ws.cell(row=1, column=1, value="No data")
        return

    # Build pivot dict: {trade_name: {county_name: vendor_count}}
    pivot = defaultdict(lambda: defaultdict(int))
    counties_set = set()
    trades_set = set()
    for row in filtered:
        pivot[row.TradeName][row.CountyName] += row.VendorBidCount
        counties_set.add(row.CountyName)
        trades_set.add(row.TradeName)

    # Sort
    counties = sorted(counties_set)
    trades = sorted(trades_set)

    # Find max value for heatmap scaling
    max_val = 0
    for t in trades:
        for c in counties:
            max_val = max(max_val, pivot[t][c])

    # Header row: "Trade" + county names
    trade_header_fill = PatternFill(start_color="1C1A83", end_color="1C1A83", fill_type="solid")
    county_header_fill = PatternFill(start_color="4A4890", end_color="4A4890", fill_type="solid")
    total_col_fill = PatternFill(start_color="2E2E8F", end_color="2E2E8F", fill_type="solid")

    # Row 1: Header
    cell = ws.cell(row=1, column=1, value="Trade")
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = trade_header_fill
    cell.alignment = Alignment(horizontal="left", wrap_text=True)
    cell.border = thin_border

    for c_idx, county in enumerate(counties, 2):
        cell = ws.cell(row=1, column=c_idx, value=county)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = county_header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True, text_rotation=90)
        cell.border = thin_border

    # Total column header
    total_col = len(counties) + 2
    cell = ws.cell(row=1, column=total_col, value="TOTAL")
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = total_col_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border

    # Data rows
    for r_idx, trade in enumerate(trades, 2):
        # Trade name
        cell = ws.cell(row=r_idx, column=1, value=trade)
        cell.font = Font(size=9)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True)

        row_total = 0
        for c_idx, county in enumerate(counties, 2):
            val = pivot[trade][county]
            row_total += val
            cell = ws.cell(row=r_idx, column=c_idx, value=val if val > 0 else "")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if val > 0:
                hm_fill = get_heatmap_color(val, max_val)
                if hm_fill:
                    cell.fill = hm_fill
                cell.font = get_font_for_heatmap(val, max_val)
            else:
                cell.font = Font(color="E0E0E0", size=9)

        # Row total
        cell = ws.cell(row=r_idx, column=total_col, value=row_total)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Bottom totals row
    totals_row = len(trades) + 2
    cell = ws.cell(row=totals_row, column=1, value="TOTAL")
    cell.font = Font(bold=True, size=10)
    cell.fill = total_fill
    cell.border = thin_border

    grand_total = 0
    for c_idx, county in enumerate(counties, 2):
        col_total = sum(pivot[t][county] for t in trades)
        grand_total += col_total
        cell = ws.cell(row=totals_row, column=c_idx, value=col_total)
        cell.font = Font(bold=True, size=9)
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    cell = ws.cell(row=totals_row, column=total_col, value=grand_total)
    cell.font = Font(bold=True, size=10)
    cell.fill = total_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 45  # Trade name column
    for c_idx in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 5
    # Row height for header
    ws.row_dimensions[1].height = 120

    ws.freeze_panes = "B2"
    print(f"  {sheet_label}: {len(trades)} trades x {len(counties)} counties, max vendor count = {max_val}")


# ============ SHEET 3: Full Detail ============
ws3 = wb.create_sheet("All Detail")
total = write_detail_sheet(ws3, detail_rows)
print(f"All Detail sheet: {total} rows")

# ============ SHEET 4: NJ Detail ============
ws4 = wb.create_sheet("NJ Detail")
nj_count = write_detail_sheet(ws4, detail_rows, state_filter="NJ")
print(f"NJ Detail sheet: {nj_count} rows")

# ============ SHEET 5: NY Detail ============
ws5 = wb.create_sheet("NY Detail")
ny_count = write_detail_sheet(ws5, detail_rows, state_filter="NY")
print(f"NY Detail sheet: {ny_count} rows")

# ============ DRILL-DOWN SHEET HELPER ============
drilldown_headers = ["State", "County", "Trade", "Vendor Name", "VendorId",
                     "Bid Description", "BidHeaderId", "Effective From", "Effective Until"]

# Alternating row fills for grouped readability
group_fill_a = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
group_fill_b = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")


def write_drilldown_sheet(ws, vendor_data, state_filter=None, sheet_label=""):
    """Write vendor-level drill-down with grouped shading by trade/county."""
    for i, h in enumerate(drilldown_headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(drilldown_headers))

    r_idx = 2
    prev_group_key = None
    use_fill_a = True

    for row in vendor_data:
        if state_filter and row.StateCode != state_filter:
            continue

        # Alternate fill every time the county+trade group changes
        group_key = (row.CountyName, row.TradeName)
        if group_key != prev_group_key:
            use_fill_a = not use_fill_a
            prev_group_key = group_key
        fill = group_fill_a if use_fill_a else group_fill_b

        vals = [row.StateCode, row.CountyName, row.TradeName, row.VendorName,
                row.VendorId, row.BidDescription, row.BidHeaderId,
                row.EffectiveFrom, row.EffectiveUntil]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            cell.fill = fill
            cell.border = thin_border
            cell.font = Font(size=9)
        r_idx += 1

    row_count = r_idx - 2
    auto_width(ws, len(drilldown_headers), max_rows=80)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(drilldown_headers))}{r_idx - 1}"
    print(f"  {sheet_label}: {row_count} vendor entries")
    return row_count


# ============ SHEET 6: Vendor Drill-Down (All) ============
print("\nBuilding drill-down sheets...")
ws_dd_all = wb.create_sheet("All Vendors")
write_drilldown_sheet(ws_dd_all, vendor_rows, state_filter=None, sheet_label="All Vendors")

# ============ SHEET 7: Vendor Drill-Down (NJ) ============
ws_dd_nj = wb.create_sheet("NJ Vendors")
write_drilldown_sheet(ws_dd_nj, vendor_rows, state_filter="NJ", sheet_label="NJ Vendors")

# ============ SHEET 8: Vendor Drill-Down (NY) ============
ws_dd_ny = wb.create_sheet("NY Vendors")
write_drilldown_sheet(ws_dd_ny, vendor_rows, state_filter="NY", sheet_label="NY Vendors")

# ============ SHEET 9: NJ Pivot (Trade x County) ============
print("\nBuilding pivot sheets...")
ws_piv_nj = wb.create_sheet("NJ Pivot")
write_pivot_sheet(ws_piv_nj, detail_rows, state_filter="NJ", sheet_label="NJ Pivot")

# ============ SHEET 10: NY Pivot (Trade x County) ============
ws_piv_ny = wb.create_sheet("NY Pivot")
write_pivot_sheet(ws_piv_ny, detail_rows, state_filter="NY", sheet_label="NY Pivot")

# ============ SHEET 11: All States Pivot (Trade x County) ============
ws_piv_all = wb.create_sheet("All Pivot")
write_pivot_sheet(ws_piv_all, detail_rows, state_filter=None, sheet_label="All Pivot")

# Save
output_path = r"C:\EDS\output\BidTrades_VendorBidCount_by_County_v3.xlsx"
os.makedirs(r"C:\EDS\output", exist_ok=True)
wb.save(output_path)
print(f"\nSaved to: {output_path}")
