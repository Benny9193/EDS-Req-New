#!/usr/bin/env python3
"""
Export Data Dictionary to Excel

Creates a multi-sheet Excel workbook with all database documentation
for easy searching and filtering.
"""

import os
import sys
from datetime import datetime

# Add scripts directory to path
sys.path.insert(0, os.path.dirname(__file__))

from db_utils import DatabaseConnection, DatabaseConnectionError
from logging_config import setup_logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def style_header_row(ws, num_cols):
    """Apply header styling to first row."""
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border


def auto_adjust_columns(ws):
    """Auto-adjust column widths based on content."""
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column].width = adjusted_width


def add_autofilter(ws, num_cols):
    """Add autofilter to header row."""
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}1"


def export_to_excel(database='EDS', output_dir='docs'):
    """Export database documentation to Excel workbook."""

    logger = setup_logging('export_to_excel')
    logger.info(f"Exporting {database} documentation to Excel")

    db = DatabaseConnection(database=database)
    db.connect()

    # Create workbook
    wb = Workbook()

    # =========================================================================
    # Sheet 1: Tables
    # =========================================================================
    logger.info("Extracting tables...")
    ws_tables = wb.active
    ws_tables.title = "Tables"

    tables = db.execute_query('''
        SELECT
            s.name AS [Schema],
            t.name AS [Table],
            ISNULL(CAST(ep.value AS NVARCHAR(500)), '') AS [Description],
            ISNULL(p.row_count, 0) AS [Rows],
            t.create_date AS [Created],
            t.modify_date AS [Modified],
            CASE WHEN pk.object_id IS NOT NULL THEN 'Yes' ELSE 'No' END AS [Has PK]
        FROM sys.tables t
        INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
        LEFT JOIN sys.dm_db_partition_stats p ON t.object_id = p.object_id AND p.index_id < 2
        LEFT JOIN (
            SELECT DISTINCT object_id FROM sys.indexes WHERE is_primary_key = 1
        ) pk ON t.object_id = pk.object_id
        LEFT JOIN sys.extended_properties ep ON t.object_id = ep.major_id
            AND ep.minor_id = 0 AND ep.name = 'MS_Description'
        ORDER BY s.name, t.name
    ''')

    ws_tables.append(['Schema', 'Table', 'Description', 'Rows', 'Created', 'Modified', 'Has PK'])
    for row in tables:
        ws_tables.append(list(row))

    style_header_row(ws_tables, 7)
    add_autofilter(ws_tables, 7)
    auto_adjust_columns(ws_tables)
    logger.info(f"  Added {len(tables)} tables")

    # =========================================================================
    # Sheet 2: Columns
    # =========================================================================
    logger.info("Extracting columns...")
    ws_columns = wb.create_sheet("Columns")

    columns = db.execute_query('''
        SELECT
            s.name AS [Schema],
            t.name AS [Table],
            c.name AS [Column],
            ISNULL(CAST(ep.value AS NVARCHAR(500)), '') AS [Description],
            ty.name AS [DataType],
            c.max_length AS [MaxLength],
            c.precision AS [Precision],
            c.scale AS [Scale],
            CASE WHEN c.is_nullable = 1 THEN 'Yes' ELSE 'No' END AS [Nullable],
            CASE WHEN c.is_identity = 1 THEN 'Yes' ELSE 'No' END AS [Identity],
            ISNULL(dc.definition, '') AS [Default],
            CASE WHEN pk.column_id IS NOT NULL THEN 'PK' ELSE '' END AS [Key]
        FROM sys.columns c
        INNER JOIN sys.tables t ON c.object_id = t.object_id
        INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
        INNER JOIN sys.types ty ON c.user_type_id = ty.user_type_id
        LEFT JOIN sys.default_constraints dc ON c.default_object_id = dc.object_id
        LEFT JOIN (
            SELECT ic.object_id, ic.column_id
            FROM sys.index_columns ic
            INNER JOIN sys.indexes i ON ic.object_id = i.object_id AND ic.index_id = i.index_id
            WHERE i.is_primary_key = 1
        ) pk ON c.object_id = pk.object_id AND c.column_id = pk.column_id
        LEFT JOIN sys.extended_properties ep ON c.object_id = ep.major_id
            AND c.column_id = ep.minor_id AND ep.name = 'MS_Description'
        ORDER BY s.name, t.name, c.column_id
    ''')

    ws_columns.append(['Schema', 'Table', 'Column', 'Description', 'DataType', 'MaxLength', 'Precision', 'Scale', 'Nullable', 'Identity', 'Default', 'Key'])
    for row in columns:
        ws_columns.append(list(row))

    style_header_row(ws_columns, 12)
    add_autofilter(ws_columns, 12)
    auto_adjust_columns(ws_columns)
    logger.info(f"  Added {len(columns)} columns")

    # =========================================================================
    # Sheet 3: Indexes
    # =========================================================================
    logger.info("Extracting indexes...")
    ws_indexes = wb.create_sheet("Indexes")

    indexes = db.execute_query('''
        SELECT
            s.name AS [Schema],
            t.name AS [Table],
            i.name AS [IndexName],
            i.type_desc AS [Type],
            CASE WHEN i.is_unique = 1 THEN 'Yes' ELSE 'No' END AS [Unique],
            CASE WHEN i.is_primary_key = 1 THEN 'Yes' ELSE 'No' END AS [PrimaryKey],
            STRING_AGG(c.name, ', ') WITHIN GROUP (ORDER BY ic.key_ordinal) AS [Columns]
        FROM sys.indexes i
        INNER JOIN sys.index_columns ic ON i.object_id = ic.object_id AND i.index_id = ic.index_id
        INNER JOIN sys.columns c ON ic.object_id = c.object_id AND ic.column_id = c.column_id
        INNER JOIN sys.tables t ON i.object_id = t.object_id
        INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
        WHERE i.type > 0
        GROUP BY s.name, t.name, i.name, i.type_desc, i.is_unique, i.is_primary_key
        ORDER BY s.name, t.name, i.name
    ''')

    ws_indexes.append(['Schema', 'Table', 'IndexName', 'Type', 'Unique', 'PrimaryKey', 'Columns'])
    for row in indexes:
        ws_indexes.append(list(row))

    style_header_row(ws_indexes, 7)
    add_autofilter(ws_indexes, 7)
    auto_adjust_columns(ws_indexes)
    logger.info(f"  Added {len(indexes)} indexes")

    # =========================================================================
    # Sheet 4: Stored Procedures
    # =========================================================================
    logger.info("Extracting stored procedures...")
    ws_procs = wb.create_sheet("StoredProcedures")

    procs = db.execute_query('''
        SELECT
            s.name AS [Schema],
            p.name AS [Procedure],
            ISNULL(CAST(ep.value AS NVARCHAR(500)), '') AS [Description],
            p.create_date AS [Created],
            p.modify_date AS [Modified],
            (SELECT COUNT(*) FROM sys.parameters WHERE object_id = p.object_id) AS [ParamCount]
        FROM sys.procedures p
        INNER JOIN sys.schemas s ON p.schema_id = s.schema_id
        LEFT JOIN sys.extended_properties ep ON p.object_id = ep.major_id
            AND ep.minor_id = 0 AND ep.name = 'MS_Description'
        ORDER BY s.name, p.name
    ''')

    ws_procs.append(['Schema', 'Procedure', 'Description', 'Created', 'Modified', 'ParamCount'])
    for row in procs:
        ws_procs.append(list(row))

    style_header_row(ws_procs, 6)
    add_autofilter(ws_procs, 6)
    auto_adjust_columns(ws_procs)
    logger.info(f"  Added {len(procs)} procedures")

    # =========================================================================
    # Sheet 5: Views
    # =========================================================================
    logger.info("Extracting views...")
    ws_views = wb.create_sheet("Views")

    views = db.execute_query('''
        SELECT
            s.name AS [Schema],
            v.name AS [View],
            ISNULL(CAST(ep.value AS NVARCHAR(500)), '') AS [Description],
            v.create_date AS [Created],
            v.modify_date AS [Modified],
            (SELECT COUNT(*) FROM sys.columns WHERE object_id = v.object_id) AS [ColumnCount]
        FROM sys.views v
        INNER JOIN sys.schemas s ON v.schema_id = s.schema_id
        LEFT JOIN sys.extended_properties ep ON v.object_id = ep.major_id
            AND ep.minor_id = 0 AND ep.name = 'MS_Description'
        ORDER BY s.name, v.name
    ''')

    ws_views.append(['Schema', 'View', 'Description', 'Created', 'Modified', 'ColumnCount'])
    for row in views:
        ws_views.append(list(row))

    style_header_row(ws_views, 6)
    add_autofilter(ws_views, 6)
    auto_adjust_columns(ws_views)
    logger.info(f"  Added {len(views)} views")

    # =========================================================================
    # Sheet 6: View Columns
    # =========================================================================
    logger.info("Extracting view columns...")
    ws_view_cols = wb.create_sheet("ViewColumns")

    view_columns = db.execute_query('''
        SELECT
            s.name AS [Schema],
            v.name AS [View],
            c.name AS [Column],
            ISNULL(CAST(ep.value AS NVARCHAR(500)), '') AS [Description],
            ty.name AS [DataType],
            c.max_length AS [MaxLength],
            c.precision AS [Precision],
            c.scale AS [Scale],
            CASE WHEN c.is_nullable = 1 THEN 'Yes' ELSE 'No' END AS [Nullable]
        FROM sys.columns c
        INNER JOIN sys.views v ON c.object_id = v.object_id
        INNER JOIN sys.schemas s ON v.schema_id = s.schema_id
        INNER JOIN sys.types ty ON c.user_type_id = ty.user_type_id
        LEFT JOIN sys.extended_properties ep ON c.object_id = ep.major_id
            AND c.column_id = ep.minor_id AND ep.name = 'MS_Description'
        ORDER BY s.name, v.name, c.column_id
    ''')

    ws_view_cols.append(['Schema', 'View', 'Column', 'Description', 'DataType', 'MaxLength', 'Precision', 'Scale', 'Nullable'])
    for row in view_columns:
        ws_view_cols.append(list(row))

    style_header_row(ws_view_cols, 9)
    add_autofilter(ws_view_cols, 9)
    auto_adjust_columns(ws_view_cols)
    logger.info(f"  Added {len(view_columns)} view columns")

    # =========================================================================
    # Sheet 7: Foreign Keys
    # =========================================================================
    logger.info("Extracting foreign keys...")
    ws_fks = wb.create_sheet("ForeignKeys")

    fks = db.execute_query('''
        SELECT
            s.name AS [Schema],
            t.name AS [Table],
            fk.name AS [FKName],
            c.name AS [Column],
            rs.name AS [RefSchema],
            rt.name AS [RefTable],
            rc.name AS [RefColumn]
        FROM sys.foreign_key_columns fkc
        INNER JOIN sys.foreign_keys fk ON fkc.constraint_object_id = fk.object_id
        INNER JOIN sys.tables t ON fkc.parent_object_id = t.object_id
        INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
        INNER JOIN sys.columns c ON fkc.parent_object_id = c.object_id AND fkc.parent_column_id = c.column_id
        INNER JOIN sys.tables rt ON fkc.referenced_object_id = rt.object_id
        INNER JOIN sys.schemas rs ON rt.schema_id = rs.schema_id
        INNER JOIN sys.columns rc ON fkc.referenced_object_id = rc.object_id AND fkc.referenced_column_id = rc.column_id
        ORDER BY s.name, t.name, fk.name
    ''')

    ws_fks.append(['Schema', 'Table', 'FKName', 'Column', 'RefSchema', 'RefTable', 'RefColumn'])
    for row in fks:
        ws_fks.append(list(row))

    style_header_row(ws_fks, 7)
    add_autofilter(ws_fks, 7)
    auto_adjust_columns(ws_fks)
    logger.info(f"  Added {len(fks)} foreign keys")

    db.disconnect()

    # Save workbook
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f'{database}_DATA_DICTIONARY.xlsx')
    wb.save(output_file)

    logger.info(f"[OK] Excel workbook saved to {output_file}")
    print(f"\nExcel data dictionary saved to: {output_file}")
    return output_file


def main():
    """Main entry point."""
    import argparse

    parser = argparse.ArgumentParser(description='Export data dictionary to Excel')
    parser.add_argument('--database', '-d', default='EDS', help='Database name (default: EDS)')
    parser.add_argument('--output', '-o', default='docs', help='Output directory (default: docs)')

    args = parser.parse_args()

    try:
        export_to_excel(database=args.database, output_dir=args.output)
    except DatabaseConnectionError as e:
        print(f"Connection failed: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
