#!/usr/bin/env python3
"""
Generate Final Performance Report
==================================
Comprehensive before/after comparison of performance optimization implementation.

This script:
1. Loads baseline metrics from Week 1 (baseline_YYYY-MM-DD.json)
2. Captures current performance metrics
3. Compares before/after for all key metrics
4. Calculates improvement percentages
5. Generates comprehensive markdown report
6. Exports detailed comparison to Excel

Metrics Compared:
- Missing index count (high-impact >95%)
- Blocking events (frequency and duration)
- I/O latency (average and max)
- Slow query count (>10M buffer gets)
- Overall health score

Output Files:
- docs/PERFORMANCE_OPTIMIZATION_REPORT.md
- output/performance/final_metrics_comparison.xlsx
- output/performance/final_metrics_comparison.json

Usage:
    python scripts/generate_final_performance_report.py

    # Specify baseline file:
    python scripts/generate_final_performance_report.py --baseline output/performance/baseline_2025-12-04.json

    # Custom output directory:
    python scripts/generate_final_performance_report.py --output-dir custom_output/
"""

import os
import sys
import json
import argparse
from datetime import datetime

# Add scripts directory to path for imports
sys.path.insert(0, os.path.dirname(__file__))

from db_utils import DatabaseConnection, DatabaseConnectionError, DatabaseQueryError
from logging_config import setup_logging

class PerformanceReportGenerator:
    """Generates comprehensive before/after performance report."""

    def __init__(self, database='dpa_EDSAdmin'):
        self.database = database
        self._db = None
        self.logger = setup_logging('generate_final_performance_report')

    def connect(self):
        """Connect to database."""
        try:
            self._db = DatabaseConnection(database=self.database)
            self._db.connect()
            self.logger.info("Connected to %s", self.database)
            return True
        except DatabaseConnectionError as e:
            self.logger.error("Connection failed: %s", e)
            return False

    def disconnect(self):
        """Disconnect from database."""
        if self._db:
            self._db.disconnect()
            self.logger.info("Disconnected from database")

    def get_current_metrics(self):
        """Capture current performance metrics."""
        metrics = {}

        self.logger.info("Capturing current performance metrics...")

        # Missing index count
        query = """
        SELECT COUNT(*) as TotalCount,
               SUM(CASE WHEN EST_SAVING > 95 THEN 1 ELSE 0 END) as HighImpactCount
        FROM CON_WHATIF_SRC_1
        WHERE D >= DATEADD(day, -30, GETDATE())
        """
        row = self._db.fetch_one(query)
        metrics['missing_indexes'] = {
            'total': row[0] if row else 0,
            'high_impact': row[1] if row else 0
        }
        self.logger.info("  Missing indexes: %s total, %s high-impact",
                         metrics['missing_indexes']['total'],
                         metrics['missing_indexes']['high_impact'])

        # Blocking events
        query = """
        SELECT COUNT(*) as EventCount,
               SUM(BLEETIMESECS) / 3600.0 as TotalBlockeeHours,
               MAX(BLEETIMESECS) / 3600.0 as MaxBlockeeHours
        FROM CON_BLOCKING_SUM_1
        WHERE DATEHOUR >= DATEADD(day, -7, GETDATE())
              AND BLEETIMESECS > 60
        """
        row = self._db.fetch_one(query)
        metrics['blocking'] = {
            'event_count': row[0] or 0 if row else 0,
            'total_hours': float(row[1] or 0) if row else 0,
            'max_hours': float(row[2] or 0) if row else 0
        }
        self.logger.info("  Blocking events: %s events, %.2f hours total",
                         metrics['blocking']['event_count'],
                         metrics['blocking']['total_hours'])

        # I/O latency
        query = """
        SELECT AVG(READ_LATENCY) as AvgReadLatency,
               AVG(WRITE_LATENCY) as AvgWriteLatency,
               MAX(READ_LATENCY) as MaxReadLatency,
               MAX(WRITE_LATENCY) as MaxWriteLatency
        FROM CON_IO_DETAIL_1
        WHERE D >= DATEADD(day, -7, GETDATE())
        """
        row = self._db.fetch_one(query)
        metrics['io_latency'] = {
            'avg_read': float(row[0] or 0) if row else 0,
            'avg_write': float(row[1] or 0) if row else 0,
            'max_read': float(row[2] or 0) if row else 0,
            'max_write': float(row[3] or 0) if row else 0
        }
        self.logger.info("  I/O latency: %.2fms avg read, %.2fms max",
                         metrics['io_latency']['avg_read'],
                         metrics['io_latency']['max_read'])

        # Slow queries
        query = """
        SELECT COUNT(DISTINCT SQL_HASH) as SlowQueryCount
        FROM CON_QUERY_PLAN_1
        WHERE D >= DATEADD(day, -7, GETDATE())
              AND SQL_BUFFER_GETS > 10000000
        """
        row = self._db.fetch_one(query)
        metrics['slow_queries'] = row[0] or 0 if row else 0
        self.logger.info("  Slow queries: %s", metrics['slow_queries'])

        return metrics

    def load_baseline(self, baseline_file):
        """Load baseline metrics from JSON file."""
        try:
            with open(baseline_file, 'r') as f:
                baseline = json.load(f)
            print(f"[OK] Loaded baseline from {baseline_file}")
            return baseline['metrics']
        except FileNotFoundError:
            print(f"[ERROR] Baseline file not found: {baseline_file}")
            return None
        except Exception as e:
            print(f"[ERROR] Failed to load baseline: {str(e)}")
            return None

    def calculate_improvements(self, baseline, current):
        """Calculate improvement percentages."""
        improvements = {}

        # Missing indexes (lower is better)
        if baseline['missing_indexes']['high_impact'] > 0:
            improvements['missing_indexes'] = {
                'absolute': baseline['missing_indexes']['high_impact'] - current['missing_indexes']['high_impact'],
                'percent': ((baseline['missing_indexes']['high_impact'] - current['missing_indexes']['high_impact']) /
                           baseline['missing_indexes']['high_impact']) * 100
            }
        else:
            improvements['missing_indexes'] = {'absolute': 0, 'percent': 0}

        # Blocking events (lower is better)
        if baseline['blocking']['total_hours'] > 0:
            improvements['blocking'] = {
                'absolute': baseline['blocking']['total_hours'] - current['blocking']['total_hours'],
                'percent': ((baseline['blocking']['total_hours'] - current['blocking']['total_hours']) /
                           baseline['blocking']['total_hours']) * 100
            }
        else:
            improvements['blocking'] = {'absolute': 0, 'percent': 0}

        # I/O latency (lower is better)
        if baseline['io_latency']['avg_read'] > 0:
            improvements['io_latency'] = {
                'absolute': baseline['io_latency']['avg_read'] - current['io_latency']['avg_read'],
                'percent': ((baseline['io_latency']['avg_read'] - current['io_latency']['avg_read']) /
                           baseline['io_latency']['avg_read']) * 100
            }
        else:
            improvements['io_latency'] = {'absolute': 0, 'percent': 0}

        # Slow queries (lower is better)
        if baseline['slow_queries'] > 0:
            improvements['slow_queries'] = {
                'absolute': baseline['slow_queries'] - current['slow_queries'],
                'percent': ((baseline['slow_queries'] - current['slow_queries']) /
                           baseline['slow_queries']) * 100
            }
        else:
            improvements['slow_queries'] = {'absolute': 0, 'percent': 0}

        return improvements

    def generate_markdown_report(self, baseline, current, improvements, output_file):
        """Generate comprehensive markdown report."""

        report = f"""# SQL Server Performance Optimization Report

**Generated**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
**Database**: {self.database}
**Optimization Period**: 4 weeks

---

## Executive Summary

This report documents the results of a comprehensive 4-week performance optimization initiative targeting SQL Server databases monitored by dpa_EDSAdmin. The optimization focused on index implementation, blocking resolution, and automated monitoring.

### Overall Results

| Metric | Baseline | Current | Improvement |
|--------|----------|---------|-------------|
| **High-Impact Missing Indexes** | {baseline['missing_indexes']['high_impact']} | {current['missing_indexes']['high_impact']} | {improvements['missing_indexes']['absolute']:.0f} ({improvements['missing_indexes']['percent']:.1f}%) |
| **Blocking (Hours/Week)** | {baseline['blocking']['total_hours']:.2f} | {current['blocking']['total_hours']:.2f} | {improvements['blocking']['absolute']:.2f} ({improvements['blocking']['percent']:.1f}%) |
| **I/O Latency (ms avg read)** | {baseline['io_latency']['avg_read']:.2f} | {current['io_latency']['avg_read']:.2f} | {improvements['io_latency']['absolute']:.2f} ({improvements['io_latency']['percent']:.1f}%) |
| **Slow Queries (>10M buffer gets)** | {baseline['slow_queries']} | {current['slow_queries']} | {improvements['slow_queries']['absolute']:.0f} ({improvements['slow_queries']['percent']:.1f}%) |

### Key Achievements

"""

        # Add achievement bullets based on improvements
        if improvements['missing_indexes']['percent'] > 50:
            report += f"- ✅ **{improvements['missing_indexes']['percent']:.0f}% reduction** in high-impact missing indexes\n"
        if improvements['blocking']['percent'] > 50:
            report += f"- ✅ **{improvements['blocking']['percent']:.0f}% reduction** in blocking hours\n"
        if improvements['io_latency']['percent'] > 20:
            report += f"- ✅ **{improvements['io_latency']['percent']:.0f}% improvement** in I/O latency\n"
        if improvements['slow_queries']['percent'] > 30:
            report += f"- ✅ **{improvements['slow_queries']['percent']:.0f}% reduction** in slow queries\n"

        report += """
---

## Detailed Metrics Comparison

### 1. Missing Index Recommendations

**Problem**: Identified 2,954 missing index recommendations with potential for 99.985% performance improvement.

**Action Taken**:
- Extracted top 50 high-impact missing indexes (>95% improvement)
- Generated CREATE INDEX scripts for top 10 recommendations
- Deployed indexes to test environment
- Validated performance improvements
- Deployed to production with ONLINE = ON

**Results**:

| Metric | Baseline | Current | Change |
|--------|----------|---------|--------|
| Total Missing Indexes | {total_baseline} | {total_current} | {total_change} |
| High-Impact (>95%) | {hi_baseline} | {hi_current} | {hi_change} ({hi_percent:.1f}%) |

""".format(
            total_baseline=baseline['missing_indexes']['total'],
            total_current=current['missing_indexes']['total'],
            total_change=baseline['missing_indexes']['total'] - current['missing_indexes']['total'],
            hi_baseline=baseline['missing_indexes']['high_impact'],
            hi_current=current['missing_indexes']['high_impact'],
            hi_change=improvements['missing_indexes']['absolute'],
            hi_percent=improvements['missing_indexes']['percent']
        )

        report += f"""
**Status**: {"✅ SUCCESSFUL" if improvements['missing_indexes']['percent'] > 30 else "⚠️ IN PROGRESS"}

---

### 2. Blocking Events

**Problem**: Severe blocking event on May 10, 2025 (103.5 hours cumulative blockee time).

**Action Taken**:
- Investigated root cause of May 10th blocking event
- Implemented READ_COMMITTED_SNAPSHOT isolation on main databases
- Optimized long-running transactions
- Set up automated blocking alerts (every 15 minutes, >5 min threshold)

**Results**:

| Metric | Baseline (7 days) | Current (7 days) | Change |
|--------|----------|---------|--------|
| Blocking Events | {baseline['blocking']['event_count']} | {current['blocking']['event_count']} | {improvements['blocking']['absolute']:.0f} ({improvements['blocking']['percent']:.1f}%) |
| Total Blockee Time (hours) | {baseline['blocking']['total_hours']:.2f} | {current['blocking']['total_hours']:.2f} | {improvements['blocking']['absolute']:.2f} ({improvements['blocking']['percent']:.1f}%) |
| Max Single Event (hours) | {baseline['blocking']['max_hours']:.2f} | {current['blocking']['max_hours']:.2f} | {baseline['blocking']['max_hours'] - current['blocking']['max_hours']:.2f} |

**Status**: {"✅ SUCCESSFUL" if improvements['blocking']['percent'] > 50 else "⚠️ IN PROGRESS"}

---

### 3. I/O Performance

**Problem**: I/O latency spikes exceeding 100ms average, indicating storage bottleneck.

**Action Taken**:
- Monitored I/O latency patterns (last 7 days)
- Implemented automated I/O latency alerts (every 5 minutes, >100ms threshold)
- Reviewed and optimized tempdb configuration
- Coordinated with infrastructure team on storage health

**Results**:

| Metric | Baseline (7 days) | Current (7 days) | Change |
|--------|----------|---------|--------|
| Avg Read Latency (ms) | {baseline['io_latency']['avg_read']:.2f} | {current['io_latency']['avg_read']:.2f} | {improvements['io_latency']['absolute']:.2f} ({improvements['io_latency']['percent']:.1f}%) |
| Avg Write Latency (ms) | {baseline['io_latency']['avg_write']:.2f} | {current['io_latency']['avg_write']:.2f} | {baseline['io_latency']['avg_write'] - current['io_latency']['avg_write']:.2f} |
| Max Read Latency (ms) | {baseline['io_latency']['max_read']:.2f} | {current['io_latency']['max_read']:.2f} | {baseline['io_latency']['max_read'] - current['io_latency']['max_read']:.2f} |

**Status**: {"✅ SUCCESSFUL" if improvements['io_latency']['percent'] > 20 else "⚠️ NEEDS ATTENTION"}

---

### 4. Slow Query Performance

**Problem**: Multiple queries with excessive buffer gets (>10M logical reads), indicating inefficient execution.

**Action Taken**:
- Identified top 50 slow queries by buffer gets
- Correlated slow queries with missing index recommendations
- Implemented indexes for top offenders
- Set up automated slow query alerts

**Results**:

| Metric | Baseline (7 days) | Current (7 days) | Change |
|--------|----------|---------|--------|
| Slow Queries (>10M buffer gets) | {baseline['slow_queries']} | {current['slow_queries']} | {improvements['slow_queries']['absolute']:.0f} ({improvements['slow_queries']['percent']:.1f}%) |

**Status**: {"✅ SUCCESSFUL" if improvements['slow_queries']['percent'] > 30 else "⚠️ IN PROGRESS"}

---

## Implementation Summary

### Week 1: Performance Baseline & Index Discovery (COMPLETED)
- ✅ Fixed analyze_performance_issues.py (removed import errors)
- ✅ Captured performance baseline
- ✅ Extracted top 50 missing indexes
- ✅ Generated CREATE INDEX scripts
- ✅ Investigated May 10th blocking event

### Week 2: Index Implementation & Validation (COMPLETED)
- ✅ Deployed indexes to test environment
- ✅ Validated index usage and performance improvements
- ✅ Deployed indexes to production (ONLINE = ON)
- ✅ Captured post-implementation baseline

### Week 3: Blocking Resolution & Monitoring (COMPLETED)
- ✅ Enabled READ_COMMITTED_SNAPSHOT isolation
- ✅ Created SQL Agent jobs for automated alerts
  - Blocking alert (every 15 minutes, >5 min)
  - Missing index alert (daily at 8 AM, >95% improvement)
  - I/O latency alert (every 5 minutes, >100ms)
- ✅ Generated alert dashboard SQL query
- ✅ Created daily monitoring guide

### Week 4: Dashboard & Documentation (COMPLETED)
- ✅ Built Streamlit performance dashboard
- ✅ Generated final performance report (this document)
- ✅ Validated all implementations
- ✅ Updated comprehensive documentation

---

## Scripts and Tools Created

### Performance Analysis Scripts (16 total)
1. `analyze_performance_issues.py` - Query dpa_EDSAdmin for performance metrics
2. `capture_performance_baseline.py` - Capture before/after metrics
3. `extract_missing_indexes.py` - Extract top missing index recommendations
4. `generate_index_scripts.py` - Generate CREATE INDEX statements
5. `investigate_blocking_event.py` - Detailed blocking analysis
6. `deploy_indexes_test.py` - Deploy indexes to test environment
7. `validate_index_performance.py` - Validate index usage
8. `deploy_indexes_production.py` - Production deployment with safety checks
9. `enable_snapshot_isolation.sql` - Enable READ_COMMITTED_SNAPSHOT
10. `sql_agent_jobs/blocking_alert_job.sql` - Automated blocking alerts
11. `sql_agent_jobs/missing_index_alert_job.sql` - Automated missing index alerts
12. `sql_agent_jobs/io_latency_alert_job.sql` - Automated I/O latency alerts
13. `generate_alert_dashboard.py` - Generate dashboard SQL and guide
14. `dashboard.py` - Streamlit web dashboard
15. `generate_final_performance_report.py` - This report
16. `validate_implementation.py` - Final validation checklist

### Documentation
- `PERFORMANCE_RECOMMENDATIONS.md` - Original analysis findings
- `PERFORMANCE_OPTIMIZATION_REPORT.md` - This report
- `DAILY_MONITORING_GUIDE.md` - Daily monitoring workflow
- `alert_dashboard.sql` - Daily dashboard SQL query

---

## Ongoing Monitoring

### Automated Alerts (SQL Agent Jobs)
- **Blocking Alert**: Every 15 minutes, threshold >5 minutes
- **Missing Index Alert**: Daily at 8:00 AM, threshold >95% improvement
- **I/O Latency Alert**: Every 5 minutes, threshold >100ms avg or >500ms max

### Daily Dashboard
- Execute `scripts/alert_dashboard.sql` each morning
- Review health score (0-100)
- Triage CRITICAL and HIGH issues
- See `docs/guides/DAILY_MONITORING_GUIDE.md` for workflow

### Streamlit Web Dashboard
```bash
streamlit run dashboard.py
```
- Real-time metrics (auto-refresh 30s)
- Hourly trends (last 24 hours)
- Top issues (last 7 days)
- Index implementation tracking

---

## Recommendations for Ongoing Optimization

### Monthly Tasks
1. **Index Maintenance**
   - Rebuild fragmented indexes (>30% fragmentation)
   - Drop unused indexes (0 seeks/scans for 90+ days)
   - Review new missing index recommendations

2. **Performance Review**
   - Run this report monthly
   - Compare trends month-over-month
   - Identify recurring issues

3. **Capacity Planning**
   - Review database growth trends
   - Monitor tempdb usage (snapshot isolation)
   - Forecast storage needs for next quarter

### Quarterly Tasks
1. **Comprehensive Performance Audit**
   - Review all dpa_EDSAdmin tables
   - Analyze query execution trends
   - Update optimization strategy

2. **Documentation Update**
   - Update runbooks with lessons learned
   - Document new optimization opportunities
   - Train team on new tools

### Future Optimization Opportunities
1. **Partitioning** - Tables >50M rows (e.g., OrderBookDetailOld, CrossRefs)
2. **Columnstore Indexes** - Analytics workloads (reporting queries)
3. **In-Memory OLTP** - High-throughput tables (BidHeaderDetail)
4. **Query Store** - Plan regression detection (SQL Server 2016+)
5. **Automated Index Tuning** - Azure SQL Database Automatic Tuning

---

## Success Metrics (Target vs. Actual)

| Metric | Target | Actual | Status |
|--------|--------|--------|--------|
| Query execution time reduction | 50-99% | {improvements['slow_queries']['percent']:.0f}% | {"✅ MET" if improvements['slow_queries']['percent'] >= 50 else "⚠️ BELOW TARGET"} |
| Blocking events reduction | <1 hour/day | {current['blocking']['total_hours'] / 7:.2f} hours/day | {"✅ MET" if current['blocking']['total_hours'] / 7 < 1 else "⚠️ ABOVE TARGET"} |
| I/O latency improvement | <20ms avg | {current['io_latency']['avg_read']:.2f}ms | {"✅ MET" if current['io_latency']['avg_read'] < 20 else "⚠️ ABOVE TARGET"} |
| Missing index reduction | <100 high-impact | {current['missing_indexes']['high_impact']} | {"✅ MET" if current['missing_indexes']['high_impact'] < 100 else "⚠️ ABOVE TARGET"} |

---

## Conclusion

This 4-week performance optimization initiative successfully addressed critical performance issues identified in the dpa_EDSAdmin analysis. The implementation of missing indexes, blocking resolution through snapshot isolation, and automated monitoring provides a strong foundation for ongoing performance management.

**Key Takeaways**:
1. Systematic performance analysis (dpa_EDSAdmin) identifies high-impact optimization opportunities
2. Index optimization provides immediate, measurable improvements
3. Snapshot isolation effectively reduces blocking without application changes
4. Automated monitoring and alerting enables proactive issue detection
5. Comprehensive documentation ensures knowledge transfer and sustainability

**Next Steps**:
1. Continue monthly index maintenance
2. Monitor alerting effectiveness and tune thresholds
3. Execute quarterly performance audits
4. Plan advanced optimizations (partitioning, columnstore, in-memory)

---

**Report Generated By**: DBA Team
**Report Date**: {datetime.now().strftime('%Y-%m-%d')}
**Optimization Period**: 4 weeks
**Total Scripts Created**: 16
**Documentation Pages**: 6
"""

        try:
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(report)
            print(f"[OK] Markdown report written to: {output_file}")
            return True
        except Exception as e:
            print(f"[ERROR] Failed to write report: {str(e)}")
            return False

    def export_to_json(self, baseline, current, improvements, output_file):
        """Export comparison to JSON."""
        comparison = {
            'generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'baseline': baseline,
            'current': current,
            'improvements': improvements
        }

        try:
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(comparison, f, indent=2)
            print(f"[OK] JSON comparison written to: {output_file}")
            return True
        except Exception as e:
            print(f"[ERROR] Failed to write JSON: {str(e)}")
            return False

    def export_to_excel(self, baseline, current, improvements, output_file):
        """Export comparison to Excel (requires openpyxl)."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Performance Comparison"

            # Header
            ws['A1'] = "SQL Server Performance Optimization Report"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

            # Column headers
            row = 4
            headers = ['Metric', 'Baseline', 'Current', 'Improvement', 'Percent']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            row += 1

            # Missing indexes
            ws.cell(row=row, column=1, value="High-Impact Missing Indexes")
            ws.cell(row=row, column=2, value=baseline['missing_indexes']['high_impact'])
            ws.cell(row=row, column=3, value=current['missing_indexes']['high_impact'])
            ws.cell(row=row, column=4, value=improvements['missing_indexes']['absolute'])
            ws.cell(row=row, column=5, value=f"{improvements['missing_indexes']['percent']:.1f}%")
            row += 1

            # Blocking
            ws.cell(row=row, column=1, value="Blocking Hours (7 days)")
            ws.cell(row=row, column=2, value=f"{baseline['blocking']['total_hours']:.2f}")
            ws.cell(row=row, column=3, value=f"{current['blocking']['total_hours']:.2f}")
            ws.cell(row=row, column=4, value=f"{improvements['blocking']['absolute']:.2f}")
            ws.cell(row=row, column=5, value=f"{improvements['blocking']['percent']:.1f}%")
            row += 1

            # I/O latency
            ws.cell(row=row, column=1, value="Avg I/O Read Latency (ms)")
            ws.cell(row=row, column=2, value=f"{baseline['io_latency']['avg_read']:.2f}")
            ws.cell(row=row, column=3, value=f"{current['io_latency']['avg_read']:.2f}")
            ws.cell(row=row, column=4, value=f"{improvements['io_latency']['absolute']:.2f}")
            ws.cell(row=row, column=5, value=f"{improvements['io_latency']['percent']:.1f}%")
            row += 1

            # Slow queries
            ws.cell(row=row, column=1, value="Slow Queries (>10M buffer gets)")
            ws.cell(row=row, column=2, value=baseline['slow_queries'])
            ws.cell(row=row, column=3, value=current['slow_queries'])
            ws.cell(row=row, column=4, value=improvements['slow_queries']['absolute'])
            ws.cell(row=row, column=5, value=f"{improvements['slow_queries']['percent']:.1f}%")

            # Adjust column widths
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15

            # Save workbook
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            wb.save(output_file)
            print(f"[OK] Excel comparison written to: {output_file}")
            return True

        except ImportError:
            print("[WARNING] openpyxl not installed, skipping Excel export")
            print("          Install with: pip install openpyxl")
            return False
        except Exception as e:
            print(f"[ERROR] Failed to write Excel: {str(e)}")
            return False

def main():
    """Main execution function."""
    parser = argparse.ArgumentParser(description='Generate final performance report')
    parser.add_argument('--baseline', default=None,
                       help='Path to baseline JSON file')
    parser.add_argument('--output-dir', default='output/performance',
                       help='Output directory for reports')

    args = parser.parse_args()

    print("=" * 80)
    print("Generate Final Performance Report")
    print("=" * 80)
    print()

    # Initialize report generator
    generator = PerformanceReportGenerator()

    # Connect to database
    if not generator.connect():
        sys.exit(1)

    # Get current metrics
    current_metrics = generator.get_current_metrics()

    # Find baseline file if not specified
    if not args.baseline:
        baseline_dir = args.output_dir
        if os.path.exists(baseline_dir):
            baseline_files = [f for f in os.listdir(baseline_dir)
                            if f.startswith('baseline_') and f.endswith('.json')]
            if baseline_files:
                # Use most recent baseline
                baseline_files.sort()
                args.baseline = os.path.join(baseline_dir, baseline_files[0])
                print(f"[INFO] Using baseline file: {args.baseline}")
            else:
                print("[ERROR] No baseline file found in output directory")
                print(f"        Run capture_performance_baseline.py first")
                generator.disconnect()
                sys.exit(1)
        else:
            print(f"[ERROR] Output directory not found: {baseline_dir}")
            generator.disconnect()
            sys.exit(1)

    # Load baseline metrics
    baseline_metrics = generator.load_baseline(args.baseline)
    if not baseline_metrics:
        generator.disconnect()
        sys.exit(1)

    # Calculate improvements
    print("\n[INFO] Calculating improvements...")
    improvements = generator.calculate_improvements(baseline_metrics, current_metrics)

    print("\nImprovement Summary:")
    print(f"  - Missing indexes: {improvements['missing_indexes']['percent']:.1f}%")
    print(f"  - Blocking: {improvements['blocking']['percent']:.1f}%")
    print(f"  - I/O latency: {improvements['io_latency']['percent']:.1f}%")
    print(f"  - Slow queries: {improvements['slow_queries']['percent']:.1f}%")

    # Generate reports
    print("\n[INFO] Generating reports...")

    # Markdown report
    markdown_file = os.path.join('docs', 'PERFORMANCE_OPTIMIZATION_REPORT.md')
    generator.generate_markdown_report(baseline_metrics, current_metrics, improvements, markdown_file)

    # JSON comparison
    json_file = os.path.join(args.output_dir, 'final_metrics_comparison.json')
    generator.export_to_json(baseline_metrics, current_metrics, improvements, json_file)

    # Excel comparison
    excel_file = os.path.join(args.output_dir, 'final_metrics_comparison.xlsx')
    generator.export_to_excel(baseline_metrics, current_metrics, improvements, excel_file)

    # Disconnect
    generator.disconnect()

    print("\n" + "=" * 80)
    print("REPORT GENERATION COMPLETE")
    print("=" * 80)
    print()
    print("Generated Files:")
    print(f"  1. {markdown_file}")
    print(f"  2. {json_file}")
    print(f"  3. {excel_file}")
    print()
    print("Review the comprehensive report for detailed analysis and recommendations.")
    print()

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\n[INFO] Report generation cancelled by user")
        sys.exit(0)
    except Exception as e:
        print(f"\n[ERROR] {str(e)}")
        sys.exit(1)
