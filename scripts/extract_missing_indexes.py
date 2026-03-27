"""
Extract Missing Index Recommendations
Queries dpa_EDSAdmin for top missing index recommendations.

Usage:
    python scripts/extract_missing_indexes.py [--days DAYS] [--min-saving PERCENT]

Arguments:
    --days: Look back period in days (default: 30)
    --min-saving: Minimum estimated saving percentage (default: 90)

Outputs:
    - output/performance/missing_indexes_top_50.csv
    - output/performance/missing_indexes_top_50.xlsx
    - output/performance/missing_indexes_full.json
"""

import sys
import os
import json
import csv
from datetime import datetime
from typing import List, Dict, Any, Optional

# Add scripts directory to path for imports
sys.path.insert(0, os.path.dirname(__file__))

from db_utils import DatabaseConnection, DatabaseConnectionError, DatabaseQueryError
from config import get_config, get_output_path, validate_params
from logging_config import setup_logging, LogContext


class MissingIndexExtractor:
    """Extracts missing index recommendations from dpa_EDSAdmin."""

    def __init__(self, database: str = None):
        self.config = get_config()
        self.database = database or self.config.database.name
        self.output_dir = get_output_path()
        self.logger = setup_logging('extract_missing_indexes')
        self._db: Optional[DatabaseConnection] = None

    def __enter__(self) -> 'MissingIndexExtractor':
        self.connect()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.disconnect()

    def connect(self) -> None:
        """Connect to database."""
        self.logger.info("Connecting to %s...", self.database)
        self._db = DatabaseConnection(database=self.database)
        self._db.connect()
        self.logger.info("Connected successfully")

    def disconnect(self) -> None:
        """Close database connection."""
        if self._db:
            self._db.disconnect()
            self._db = None
            self.logger.info("Disconnected")

    def extract_missing_indexes(
        self,
        days: int = None,
        min_saving: int = None
    ) -> List[Dict[str, Any]]:
        """
        Extract top 50 missing index recommendations.

        Args:
            days: Look back period in days
            min_saving: Minimum estimated saving percentage

        Returns:
            List of missing index recommendations
        """
        days = days or self.config.analysis.missing_indexes_days
        min_saving = min_saving or self.config.thresholds.missing_index.high

        validate_params(days=days, min_saving=min_saving)

        self.logger.info("=" * 80)
        self.logger.info("MISSING INDEX RECOMMENDATIONS EXTRACTION")
        self.logger.info("=" * 80)
        self.logger.info("Period: Last %d days", days)
        self.logger.info("Minimum estimated saving: %d%%", min_saving)

        # Parameterized query - prevents SQL injection
        query = """
        SELECT TOP 50
            w.ID,
            CONVERT(VARCHAR, w.D, 120) as DateIdentified,
            w.SQL_HASH as QueryHash,
            w.SQL_EXECS as Executions,
            w.SQL_WAIT as WaitTimeMS,
            w.EST_SAVING as EstimatedSavingPercent,
            w.IDX_ID as IndexID,
            fq.ODATABASE as DatabaseName,
            fq.OSCHEMA as SchemaName,
            fq.ONAME as TableName,
            fq.OINDEX_TABLE as IndexName,
            fq.OINDEX_COLUMNS as IndexColumns,
            fq.OINDEX_INCLUDED_COLUMNS as IncludedColumns,
            SUBSTRING(st.ST, 1, 1000) as SQLText
        FROM CON_WHATIF_SRC_1 w
        LEFT JOIN CON_FQ_OBJECT_1 fq ON w.IDX_ID = fq.ID
        LEFT JOIN CONST_1 st ON w.SQL_HASH = st.H
        WHERE w.D >= DATEADD(day, -?, GETDATE())
            AND w.EST_SAVING >= ?
        ORDER BY w.EST_SAVING DESC, w.SQL_EXECS DESC
        """

        with LogContext(self.logger, "Querying missing indexes"):
            rows = self._db.execute_query(query, (days, min_saving))

        results = []
        for row in rows:
            results.append({
                'ID': row[0],
                'DateIdentified': str(row[1]) if row[1] else None,
                'QueryHash': row[2],
                'Executions': row[3],
                'WaitTimeMS': row[4],
                'EstimatedSavingPercent': float(row[5]) if row[5] else 0,
                'IndexID': row[6],
                'DatabaseName': row[7],
                'SchemaName': row[8],
                'TableName': row[9],
                'IndexName': row[10],
                'IndexColumns': row[11],
                'IncludedColumns': row[12],
                'SQLText': row[13]
            })

        self.logger.info("[OK] Found %d missing index recommendations", len(results))

        if results:
            self._print_summary(results)

        return results

    def _print_summary(self, results: List[Dict]) -> None:
        """Print summary of top recommendations."""
        self.logger.info("TOP 10 MISSING INDEX RECOMMENDATIONS:")
        self.logger.info("=" * 80)
        self.logger.info(
            "%-4s %-10s %-10s %-15s %-30s",
            "#", "Saving%", "Execs", "Database", "Table"
        )
        self.logger.info("-" * 80)

        for i, r in enumerate(results[:10], 1):
            db = (r['DatabaseName'] or 'N/A')[:14]
            table = (r['TableName'] or 'N/A')[:29]
            saving = r['EstimatedSavingPercent']
            execs = r['Executions'] or 0
            self.logger.info(
                "%-4d %-10.3f %-10d %-15s %-30s",
                i, saving, execs, db, table
            )

        self.logger.info("=" * 80)

    def save_to_csv(self, results: List[Dict]) -> None:
        """Save results to CSV file."""
        output_file = self.output_dir / 'missing_indexes_top_50.csv'

        if not results:
            self.logger.info("[INFO] No results to save to CSV")
            return

        fieldnames = [
            'ID', 'DateIdentified', 'EstimatedSavingPercent', 'Executions',
            'WaitTimeMS', 'DatabaseName', 'SchemaName', 'TableName',
            'IndexColumns', 'IncludedColumns', 'QueryHash'
        ]

        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(
                f, fieldnames=fieldnames, extrasaction='ignore'
            )
            writer.writeheader()
            writer.writerows(results)

        self.logger.info("[OK] CSV saved to: %s", output_file)

    def save_to_excel(self, results: List[Dict]) -> None:
        """Save results to Excel file."""
        output_file = self.output_dir / 'missing_indexes_top_50.xlsx'

        if not results:
            self.logger.info("[INFO] No results to save to Excel")
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Missing Indexes"

            headers = [
                'ID', 'Date Identified', 'Est. Saving %', 'Executions',
                'Wait Time (ms)', 'Database', 'Schema', 'Table',
                'Index Columns', 'Included Columns', 'Query Hash'
            ]

            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row_idx, record in enumerate(results, 2):
                ws.cell(row=row_idx, column=1, value=record['ID'])
                ws.cell(row=row_idx, column=2, value=record['DateIdentified'])
                ws.cell(row=row_idx, column=3, value=record['EstimatedSavingPercent'])
                ws.cell(row=row_idx, column=4, value=record['Executions'])
                ws.cell(row=row_idx, column=5, value=record['WaitTimeMS'])
                ws.cell(row=row_idx, column=6, value=record['DatabaseName'])
                ws.cell(row=row_idx, column=7, value=record['SchemaName'])
                ws.cell(row=row_idx, column=8, value=record['TableName'])
                ws.cell(row=row_idx, column=9, value=record['IndexColumns'])
                ws.cell(row=row_idx, column=10, value=record['IncludedColumns'])
                ws.cell(row=row_idx, column=11, value=record['QueryHash'])

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(output_file)
            self.logger.info("[OK] Excel saved to: %s", output_file)

        except ImportError:
            self.logger.warning("openpyxl not installed, skipping Excel export")
            self.logger.info("Install with: pip install openpyxl")

    def save_to_json(self, results: List[Dict]) -> None:
        """Save full results to JSON file."""
        output_file = self.output_dir / 'missing_indexes_full.json'

        if not results:
            self.logger.info("[INFO] No results to save to JSON")
            return

        data = {
            'extraction_date': datetime.now().isoformat(),
            'total_recommendations': len(results),
            'recommendations': results
        }

        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)

        self.logger.info("[OK] JSON saved to: %s", output_file)


def main() -> int:
    """Main execution."""
    import argparse

    parser = argparse.ArgumentParser(
        description='Extract missing index recommendations'
    )
    parser.add_argument(
        '--days', type=int, default=30,
        help='Look back period in days (default: 30)'
    )
    parser.add_argument(
        '--min-saving', type=int, default=90,
        help='Minimum estimated saving percentage (default: 90)'
    )
    parser.add_argument(
        '--log-level', choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
        default='INFO', help='Logging level'
    )

    args = parser.parse_args()

    logger = setup_logging('extract_missing_indexes', log_level=args.log_level)

    logger.info("=" * 80)
    logger.info("MISSING INDEX RECOMMENDATIONS EXTRACTOR")
    logger.info("=" * 80)
    logger.info("Start Time: %s", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

    try:
        with MissingIndexExtractor() as extractor:
            results = extractor.extract_missing_indexes(
                days=args.days,
                min_saving=args.min_saving
            )

            if results:
                extractor.save_to_csv(results)
                extractor.save_to_excel(results)
                extractor.save_to_json(results)
            else:
                logger.info("No missing index recommendations found")

        logger.info("=" * 80)
        logger.info("EXTRACTION COMPLETE")
        logger.info("=" * 80)
        logger.info("End Time: %s", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        logger.info("Total Recommendations: %d", len(results))
        logger.info("=" * 80)

        return 0

    except DatabaseConnectionError as e:
        logger.error("Connection failed: %s", e)
        return 1
    except DatabaseQueryError as e:
        logger.error("Query failed: %s", e)
        return 1
    except ValueError as e:
        logger.error("Invalid parameter: %s", e)
        return 1
    except Exception as e:
        logger.exception("Extraction failed: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())
