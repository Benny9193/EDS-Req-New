import pyodbc
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# --- DB Connection ---
conn = pyodbc.connect(
    'DRIVER={ODBC Driver 17 for SQL Server};'
    'SERVER=eds-sqlserver.eastus2.cloudapp.azure.com;'
    'DATABASE=EDS;UID=EDSAdmin;PWD=Consultant~!'
)
cursor = conn.cursor()

print("Querying data for Berkeley Heights (DistrictId=252)...")

query = """
SELECT
    Category.Name                                      AS CategoryName,
    Detail.ItemCode,
    Detail.Description                                 AS ItemDescription,
    isnull(Vendors.DisplayAs, Vendors.Name)            AS VendorName,
    SUM(Detail.Quantity)                               AS Quantity,
    Detail.UnitCode,
    Detail.BidPrice,

    -- CatalogPrice logic mirrored from vw_SavingsDetail1
    CASE Category.CategoryId
        WHEN 12 THEN ROUND((Detail.BidPrice / .6), 2)
        WHEN 15 THEN (
            SELECT TOP 1 CatalogPrice
            FROM CrossRefs
            JOIN Catalog ON Catalog.CatalogId = CrossRefs.CatalogId
                        AND Catalog.Name = 'EDS'
                        AND Catalog.Active = 1
            WHERE CrossRefs.ItemId = Detail.ItemId
              AND CrossRefs.Active = 1
              AND CrossRefs.CatalogPrice < 9999
            ORDER BY CatalogPrice DESC
        )
        WHEN 44 THEN ROUND((Detail.BidPrice / .6), 2)
        ELSE ISNULL(Detail.CatalogPrice, Detail.BidPrice / .6)
    END                                                AS CatalogPrice,

    -- Discount %
    CASE ISNULL(
        CASE Category.CategoryId
            WHEN 12 THEN ROUND((Detail.BidPrice / .6), 2)
            WHEN 15 THEN (
                SELECT TOP 1 CatalogPrice FROM CrossRefs
                JOIN Catalog ON Catalog.CatalogId = CrossRefs.CatalogId AND Catalog.Name='EDS' AND Catalog.Active=1
                WHERE CrossRefs.ItemId=Detail.ItemId AND CrossRefs.Active=1 AND CrossRefs.CatalogPrice<9999
                ORDER BY CatalogPrice DESC
            )
            WHEN 44 THEN ROUND((Detail.BidPrice / .6), 2)
            ELSE ISNULL(Detail.CatalogPrice, Detail.BidPrice / .6)
        END, Detail.BidPrice / .6)
    WHEN 0 THEN 0
    ELSE 1 - (Detail.BidPrice / ISNULL(
        CASE Category.CategoryId
            WHEN 12 THEN ROUND((Detail.BidPrice / .6), 2)
            WHEN 15 THEN (
                SELECT TOP 1 CatalogPrice FROM CrossRefs
                JOIN Catalog ON Catalog.CatalogId = CrossRefs.CatalogId AND Catalog.Name='EDS' AND Catalog.Active=1
                WHERE CrossRefs.ItemId=Detail.ItemId AND CrossRefs.Active=1 AND CrossRefs.CatalogPrice<9999
                ORDER BY CatalogPrice DESC
            )
            WHEN 44 THEN ROUND((Detail.BidPrice / .6), 2)
            ELSE ISNULL(Detail.CatalogPrice, Detail.BidPrice / .6)
        END, Detail.BidPrice / .6))
    END                                                AS DiscountPct,

    SUM(Detail.Quantity * Detail.BidPrice)             AS BidExtended,

    -- CatalogExtended mirrored from view
    CASE Category.CategoryId
        WHEN 12 THEN ROUND((Detail.BidPrice / .6), 2)
        WHEN 15 THEN CASE
            WHEN (
                SELECT TOP 1 CatalogPrice FROM CrossRefs
                JOIN Catalog ON Catalog.CatalogId=CrossRefs.CatalogId AND Catalog.Name='EDS' AND Catalog.Active=1
                WHERE CrossRefs.ItemId=Detail.ItemId AND CrossRefs.Active=1 AND CrossRefs.CatalogPrice<9999
                ORDER BY CatalogPrice DESC
            ) > COALESCE(Detail.CatalogPrice,0)
            THEN COALESCE((
                SELECT TOP 1 CatalogPrice FROM CrossRefs
                JOIN Catalog ON Catalog.CatalogId=CrossRefs.CatalogId AND Catalog.Name='EDS' AND Catalog.Active=1
                WHERE CrossRefs.ItemId=Detail.ItemId AND CrossRefs.Active=1 AND CrossRefs.CatalogPrice<9999
                ORDER BY CatalogPrice DESC
            ), ROUND((Detail.BidPrice / .4), 2))
            ELSE COALESCE(Detail.CatalogPrice, ROUND((Detail.BidPrice / .4), 2))
        END
        WHEN 44 THEN ROUND((Detail.BidPrice / .6), 2)
        ELSE ISNULL(Detail.CatalogPrice, Detail.BidPrice / .6)
    END * SUM(Detail.Quantity)                         AS CatalogExtended,

    Awards.StateContractDiscount / 100.0               AS StateContractDiscount,

    Budgets.BudgetId,
    Requisitions.RequisitionId

FROM Detail
JOIN Awards       ON Awards.AwardId         = Detail.AwardId
JOIN Vendors      ON Vendors.VendorId        = Awards.VendorId
JOIN Requisitions ON Requisitions.RequisitionId = Detail.RequisitionId
                 AND ISNULL(Requisitions.StatusId, 0) != 4
JOIN Budgets      ON Budgets.BudgetId        = Requisitions.BudgetId
JOIN District     ON District.DistrictId     = Budgets.DistrictId
JOIN Category     ON Category.CategoryId     = Requisitions.CategoryId
                 AND (
                        (ISNULL(Category.OnSavingsReport,0) = 1)
                     OR (ISNULL(Category.OnSavingsReport,0) != 1)
                 )
WHERE District.DistrictId = 252
  AND Budgets.BudgetId = 1685524
  AND ISNULL(Detail.ItemMustBeBid, 0) = 0
  AND ISNULL(Detail.VendorId, 7691) != 7691
  AND CASE
        WHEN CASE Category.CategoryId
               WHEN 12 THEN ROUND((Detail.BidPrice / .6), 2)
               WHEN 15 THEN (
                   SELECT TOP 1 CatalogPrice FROM CrossRefs
                   JOIN Catalog ON Catalog.CatalogId=CrossRefs.CatalogId AND Catalog.Name='EDS' AND Catalog.Active=1
                   WHERE CrossRefs.ItemId=Detail.ItemId AND CrossRefs.Active=1 AND CrossRefs.CatalogPrice<9999
                   ORDER BY CatalogPrice DESC
               )
               WHEN 44 THEN ROUND((Detail.BidPrice / .6), 2)
               ELSE ISNULL((
                   SELECT TOP 1 CatalogPrice FROM CrossRefs
                   JOIN Catalog ON Catalog.CatalogId=CrossRefs.CatalogId AND Catalog.Active=1
                   WHERE CrossRefs.ItemId=Detail.ItemId AND CrossRefs.Active=1 AND CrossRefs.CatalogPrice<9999
                   ORDER BY CatalogPrice DESC
               ), ISNULL(Detail.CatalogPrice, Detail.BidPrice))
             END > ISNULL(Detail.CatalogPrice, 0)
        THEN CASE Category.CategoryId
               WHEN 12 THEN ROUND((Detail.BidPrice / .6), 2)
               WHEN 15 THEN (
                   SELECT TOP 1 CatalogPrice FROM CrossRefs
                   JOIN Catalog ON Catalog.CatalogId=CrossRefs.CatalogId AND Catalog.Name='EDS' AND Catalog.Active=1
                   WHERE CrossRefs.ItemId=Detail.ItemId AND CrossRefs.Active=1 AND CrossRefs.CatalogPrice<9999
                   ORDER BY CatalogPrice DESC
               )
               WHEN 44 THEN ROUND((Detail.BidPrice / .6), 2)
               ELSE ISNULL((
                   SELECT TOP 1 CatalogPrice FROM CrossRefs
                   JOIN Catalog ON Catalog.CatalogId=CrossRefs.CatalogId AND Catalog.Active=1
                   WHERE CrossRefs.ItemId=Detail.ItemId AND CrossRefs.Active=1 AND CrossRefs.CatalogPrice<9999
                   ORDER BY CatalogPrice DESC
               ), ISNULL(Detail.CatalogPrice, Detail.BidPrice))
             END
        ELSE Detail.CatalogPrice
      END >= ISNULL(Detail.BidPrice, 0)

GROUP BY
    Category.CategoryId, Category.Name,
    Detail.ItemCode, Detail.Description,
    Vendors.VendorId, Vendors.Name, Vendors.DisplayAs,
    Detail.UnitCode, Detail.BidPrice, Detail.ItemId,
    Detail.CatalogPrice, Awards.StateContractDiscount,
    Budgets.BudgetId, Requisitions.RequisitionId

ORDER BY Category.Name, VendorName, Detail.ItemCode
"""

cursor.execute(query)
rows = cursor.fetchall()
cols = [col[0] for col in cursor.description]
conn.close()

print(f"Retrieved {len(rows)} rows.")

# --- Build Excel ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Savings Detail"

# EDS brand colors
COLOR_PRIMARY   = "1c1a83"
COLOR_SECONDARY = "4a4890"
COLOR_ACCENT    = "b70c0d"
COLOR_WHITE     = "FFFFFF"
COLOR_LIGHT_BG  = "F0F0F8"
COLOR_SUBTOTAL  = "D8D8F0"

# Styles
hdr_font    = Font(name="Calibri", bold=True, color=COLOR_WHITE, size=11)
hdr_fill    = PatternFill("solid", fgColor=COLOR_PRIMARY)
hdr_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)

cat_font    = Font(name="Calibri", bold=True, color=COLOR_WHITE, size=10)
cat_fill    = PatternFill("solid", fgColor=COLOR_SECONDARY)

sub_font    = Font(name="Calibri", bold=True, size=10)
sub_fill    = PatternFill("solid", fgColor=COLOR_SUBTOTAL)

body_font   = Font(name="Calibri", size=10)
alt_fill    = PatternFill("solid", fgColor=COLOR_LIGHT_BG)

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

fmt_currency = '$#,##0.00'
fmt_pct      = '0.00%'
fmt_number   = '#,##0'

# --- Title block ---
ws.merge_cells("A1:K1")
title_cell = ws["A1"]
title_cell.value = "Berkeley Heights Public Schools — Line by Line Savings Report (2025–2026)"
title_cell.font = Font(name="Calibri", bold=True, size=14, color=COLOR_WHITE)
title_cell.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:K2")
date_cell = ws["A2"]
date_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
date_cell.font = Font(name="Calibri", italic=True, size=10, color="555555")
date_cell.alignment = Alignment(horizontal="center")
ws.row_dimensions[2].height = 18

ws.row_dimensions[3].height = 6  # spacer

# --- Column headers (row 4) ---
headers = [
    "Category", "Item Code", "Description", "Vendor",
    "Qty", "Unit", "Bid Price", "Catalog Price",
    "Discount %", "Bid Extended", "Catalog Extended"
]
col_widths = [22, 14, 42, 28, 8, 8, 12, 14, 12, 16, 18]

for c_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=4, column=c_idx, value=hdr)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = border
    ws.column_dimensions[get_column_letter(c_idx)].width = width

ws.row_dimensions[4].height = 30
ws.freeze_panes = "A5"

# --- Data rows ---
data_row = 5
current_category = None
cat_start_row = data_row

grand_bid_ext = 0.0
grand_cat_ext = 0.0

# Totals tracking
cat_bid_ext = 0.0
cat_cat_ext = 0.0

def write_category_subtotal(ws, row, cat_name, bid_total, cat_total):
    savings = cat_total - bid_total
    ws.merge_cells(f"A{row}:E{row}")
    lbl = ws.cell(row=row, column=1,
                  value=f"  {cat_name} Subtotal")
    lbl.font = sub_font
    lbl.fill = sub_fill
    lbl.alignment = Alignment(horizontal="left", vertical="center")
    lbl.border = border

    for c in range(2, 6):
        ws.cell(row=row, column=c).fill = sub_fill
        ws.cell(row=row, column=c).border = border

    # blank cols 6-9
    for c in range(6, 10):
        ws.cell(row=row, column=c).fill = sub_fill
        ws.cell(row=row, column=c).border = border

    bid_cell = ws.cell(row=row, column=10, value=bid_total)
    bid_cell.font = sub_font; bid_cell.fill = sub_fill
    bid_cell.number_format = fmt_currency; bid_cell.border = border

    cat_cell = ws.cell(row=row, column=11, value=cat_total)
    cat_cell.font = sub_font; cat_cell.fill = sub_fill
    cat_cell.number_format = fmt_currency; cat_cell.border = border

    # Savings label in a 12th column
    ws.column_dimensions["L"].width = 16
    sav_cell = ws.cell(row=row, column=12, value=savings)
    sav_cell.font = Font(name="Calibri", bold=True, size=10,
                         color="006100" if savings >= 0 else COLOR_ACCENT)
    sav_cell.fill = sub_fill
    sav_cell.number_format = fmt_currency
    sav_cell.border = border

    sav_hdr = ws.cell(row=4, column=12, value="Savings")
    sav_hdr.font = hdr_font; sav_hdr.fill = hdr_fill
    sav_hdr.alignment = hdr_align; sav_hdr.border = border

    ws.row_dimensions[row].height = 20

for i, row_data in enumerate(rows):
    d = dict(zip(cols, row_data))

    cat_name  = d["CategoryName"] or ""
    bid_ext   = float(d["BidExtended"] or 0)
    cat_ext   = float(d["CatalogExtended"] or 0)

    # Category header row
    if cat_name != current_category:
        # Write subtotal for previous category
        if current_category is not None:
            write_category_subtotal(ws, data_row, current_category, cat_bid_ext, cat_cat_ext)
            data_row += 1
            # blank spacer
            ws.row_dimensions[data_row].height = 6
            data_row += 1

        # Category header
        ws.merge_cells(f"A{data_row}:L{data_row}")
        ch = ws.cell(row=data_row, column=1, value=f"  {cat_name}")
        ch.font = cat_font
        ch.fill = cat_fill
        ch.alignment = Alignment(horizontal="left", vertical="center")
        ch.border = border
        for c in range(2, 13):
            cell = ws.cell(row=data_row, column=c)
            cell.fill = cat_fill
            cell.border = border
        ws.row_dimensions[data_row].height = 22
        data_row += 1

        current_category = cat_name
        cat_bid_ext = 0.0
        cat_cat_ext = 0.0

    # Data row
    use_alt = (i % 2 == 0)
    row_fill = alt_fill if use_alt else None

    values = [
        cat_name,
        d["ItemCode"],
        d["ItemDescription"],
        d["VendorName"],
        d["Quantity"],
        d["UnitCode"],
        float(d["BidPrice"] or 0),
        float(d["CatalogPrice"] or 0),
        float(d["DiscountPct"] or 0),
        bid_ext,
        cat_ext,
    ]
    fmts = [None, None, None, None, fmt_number, None,
            fmt_currency, fmt_currency, fmt_pct, fmt_currency, fmt_currency]
    aligns = ["left","left","left","left","center","center",
              "right","right","right","right","right"]

    for c_idx, (val, fmt, aln) in enumerate(zip(values, fmts, aligns), start=1):
        cell = ws.cell(row=data_row, column=c_idx, value=val)
        cell.font = body_font
        cell.alignment = Alignment(horizontal=aln, vertical="center")
        cell.border = border
        if fmt:
            cell.number_format = fmt
        if row_fill:
            cell.fill = row_fill

    ws.row_dimensions[data_row].height = 16
    cat_bid_ext += bid_ext
    cat_cat_ext += cat_ext
    grand_bid_ext += bid_ext
    grand_cat_ext += cat_ext
    data_row += 1

# Last category subtotal
if current_category is not None:
    write_category_subtotal(ws, data_row, current_category, cat_bid_ext, cat_cat_ext)
    data_row += 2

# --- Grand Total row ---
grand_savings = grand_cat_ext - grand_bid_ext
ws.merge_cells(f"A{data_row}:E{data_row}")
gt_lbl = ws.cell(row=data_row, column=1, value="GRAND TOTAL")
gt_lbl.font = Font(name="Calibri", bold=True, size=12, color=COLOR_WHITE)
gt_fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
gt_lbl.fill = gt_fill
gt_lbl.alignment = Alignment(horizontal="left", vertical="center")
gt_lbl.border = border
for c in range(2, 13):
    ws.cell(row=data_row, column=c).fill = gt_fill
    ws.cell(row=data_row, column=c).border = border

gt_bid = ws.cell(row=data_row, column=10, value=grand_bid_ext)
gt_bid.font = Font(name="Calibri", bold=True, size=12, color=COLOR_WHITE)
gt_bid.fill = gt_fill; gt_bid.number_format = fmt_currency; gt_bid.border = border

gt_cat = ws.cell(row=data_row, column=11, value=grand_cat_ext)
gt_cat.font = Font(name="Calibri", bold=True, size=12, color=COLOR_WHITE)
gt_cat.fill = gt_fill; gt_cat.number_format = fmt_currency; gt_cat.border = border

gt_sav = ws.cell(row=data_row, column=12, value=grand_savings)
gt_sav.font = Font(name="Calibri", bold=True, size=12, color=COLOR_WHITE)
gt_sav.fill = gt_fill; gt_sav.number_format = fmt_currency; gt_sav.border = border

ws.row_dimensions[data_row].height = 26

# --- Auto-filter on header row ---
ws.auto_filter.ref = f"A4:L4"

# --- Summary Tab ---
ws2 = wb.create_sheet("Summary", 0)  # insert before detail sheet

# Title
ws2.merge_cells("A1:G1")
t = ws2["A1"]
t.value = "Berkeley Heights Public Schools — Savings Summary by Category (2025–2026)"
t.font = Font(name="Calibri", bold=True, size=14, color=COLOR_WHITE)
t.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

ws2.merge_cells("A2:G2")
d2 = ws2["A2"]
d2.value = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
d2.font = Font(name="Calibri", italic=True, size=10, color="555555")
d2.alignment = Alignment(horizontal="center")
ws2.row_dimensions[2].height = 18
ws2.row_dimensions[3].height = 6

# Headers
sum_headers = ["Category", "Quantity", "Bid Total", "Catalog Total", "Savings $", "Savings %", "Avg Discount %"]
sum_widths  = [28, 12, 16, 16, 16, 14, 16]
for c_idx, (hdr, width) in enumerate(zip(sum_headers, sum_widths), start=1):
    cell = ws2.cell(row=4, column=c_idx, value=hdr)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = border
    ws2.column_dimensions[get_column_letter(c_idx)].width = width
ws2.row_dimensions[4].height = 30
ws2.freeze_panes = "A5"

# Aggregate from rows data
from collections import defaultdict
cat_summary = defaultdict(lambda: {"qty": 0, "bid_ext": 0.0, "cat_ext": 0.0})
for row_data in rows:
    d = dict(zip(cols, row_data))
    cat = d["CategoryName"] or ""
    cat_summary[cat]["qty"]     += int(d["Quantity"] or 0)
    cat_summary[cat]["bid_ext"] += float(d["BidExtended"] or 0)
    cat_summary[cat]["cat_ext"] += float(d["CatalogExtended"] or 0)

sum_row = 5
sum_grand_qty = 0
sum_grand_bid = 0.0
sum_grand_cat = 0.0

for i, (cat_name, vals) in enumerate(sorted(cat_summary.items())):
    qty     = vals["qty"]
    bid_ext = vals["bid_ext"]
    cat_ext = vals["cat_ext"]
    savings = cat_ext - bid_ext
    sav_pct = (savings / cat_ext) if cat_ext else 0
    avg_disc = (1 - (bid_ext / cat_ext)) if cat_ext else 0

    use_alt = (i % 2 == 0)
    rfill = alt_fill if use_alt else None

    row_vals = [cat_name, qty, bid_ext, cat_ext, savings, sav_pct, avg_disc]
    row_fmts = [None, fmt_number, fmt_currency, fmt_currency, fmt_currency, fmt_pct, fmt_pct]
    row_alns = ["left","center","right","right","right","right","right"]

    for c_idx, (val, fmt, aln) in enumerate(zip(row_vals, row_fmts, row_alns), start=1):
        cell = ws2.cell(row=sum_row, column=c_idx, value=val)
        cell.font = body_font
        cell.alignment = Alignment(horizontal=aln, vertical="center")
        cell.border = border
        if fmt:
            cell.number_format = fmt
        if rfill:
            cell.fill = rfill
        # Color savings cells
        if c_idx == 5:
            cell.font = Font(name="Calibri", size=10,
                             color="006100" if savings >= 0 else COLOR_ACCENT)

    ws2.row_dimensions[sum_row].height = 18
    sum_grand_qty += qty
    sum_grand_bid += bid_ext
    sum_grand_cat += cat_ext
    sum_row += 1

# Grand total row
sum_grand_sav     = sum_grand_cat - sum_grand_bid
sum_grand_sav_pct = (sum_grand_sav / sum_grand_cat) if sum_grand_cat else 0

gt_vals = ["GRAND TOTAL", sum_grand_qty, sum_grand_bid, sum_grand_cat,
           sum_grand_sav, sum_grand_sav_pct, sum_grand_sav_pct]
gt_fmts = [None, fmt_number, fmt_currency, fmt_currency, fmt_currency, fmt_pct, fmt_pct]
gt_alns = ["left","center","right","right","right","right","right"]

for c_idx, (val, fmt, aln) in enumerate(zip(gt_vals, gt_fmts, gt_alns), start=1):
    cell = ws2.cell(row=sum_row, column=c_idx, value=val)
    cell.font = Font(name="Calibri", bold=True, size=12, color=COLOR_WHITE)
    cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal=aln, vertical="center")
    cell.border = border
    if fmt:
        cell.number_format = fmt
ws2.row_dimensions[sum_row].height = 26

ws2.auto_filter.ref = f"A4:G4"

# --- Save ---
out_path = r"C:\EDS\output\Berkeley_Heights_Savings_Report_2025-26.xlsx"
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print(f"\nReport saved to: {out_path}")
print(f"Grand Bid Extended:     ${grand_bid_ext:,.2f}")
print(f"Grand Catalog Extended: ${grand_cat_ext:,.2f}")
print(f"Grand Savings:          ${grand_savings:,.2f}")
