"""
Generate combined Nasco report: all items summed across all bids,
sortable by Vendor Item Code. Plus the existing bid-level sheets.
"""
import pickle, os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Load data ---
with open(r"C:\EDS\nasco_items_raw3.pkl", "rb") as f:
    data = pickle.load(f)

pdi_df = pd.DataFrame(data["pdi"]["rows"], columns=data["pdi"]["cols"])
items_df = pd.DataFrame(data["items"]["rows"], columns=data["items"]["cols"])
awards_df = pd.DataFrame(data["awards"]["rows"], columns=data["awards"]["cols"])

df = pdi_df.merge(items_df, on="ItemId", how="left")
df = df.merge(awards_df, on="AwardId", how="left")
df["LineAmount"] = df["Quantity"] * df["BidPrice"]

# --- Aggregate 1: Combined by VendorItemCode (across all bids) ---
combined = df.groupby([
    "VendorItemCode", "ItemCode", "Description"
]).agg(
    TotalQtyOrdered=("Quantity", "sum"),
    MinBidPrice=("BidPrice", "min"),
    MaxBidPrice=("BidPrice", "max"),
    AvgBidPrice=("BidPrice", "mean"),
    TotalAmount=("LineAmount", "sum"),
    POCount=("POId", "nunique"),
    BidCount=("BidId", "nunique"),
    BidHeaderIds=("BidHeaderId", lambda x: ", ".join(sorted(str(v) for v in x.dropna().unique()))),
    Categories=("CategoryName", lambda x: ", ".join(sorted(x.dropna().unique()))),
).reset_index()

combined = combined.sort_values("VendorItemCode", ascending=True)

# --- Aggregate 2: Items by bid ---
detail_agg = df.groupby([
    "CategoryName", "CategoryCode",
    "BidId", "BidName", "VendorBidNumber",
    "BidHeaderId", "BidHeaderDescription",
    "EffectiveFrom", "EffectiveUntil",
    "ItemCode", "VendorItemCode", "Description"
]).agg(
    TotalQtyOrdered=("Quantity", "sum"),
    MinBidPrice=("BidPrice", "min"),
    MaxBidPrice=("BidPrice", "max"),
    TotalAmount=("LineAmount", "sum"),
    POCount=("POId", "nunique"),
).reset_index()
detail_agg = detail_agg.sort_values(["CategoryName", "BidName", "TotalAmount"], ascending=[True, True, False])

# --- Aggregate 3: Bid summary ---
bid_agg = df.groupby([
    "CategoryName", "CategoryCode",
    "BidId", "BidName", "VendorBidNumber",
    "BidHeaderId", "BidHeaderDescription",
    "EffectiveFrom", "EffectiveUntil",
]).agg(
    UniqueItems=("ItemCode", "nunique"),
    TotalQtyOrdered=("Quantity", "sum"),
    TotalAmount=("LineAmount", "sum"),
    POCount=("POId", "nunique"),
    FirstPODate=("PODate", "min"),
    LastPODate=("PODate", "max"),
).reset_index()
bid_agg = bid_agg.sort_values(["CategoryName", "TotalAmount"], ascending=[True, False])

print(f"Combined items: {len(combined)}")
print(f"Detail rows: {len(detail_agg)}")
print(f"Bid summary rows: {len(bid_agg)}")

# ===================== EXCEL GENERATION =====================
EDS_PRIMARY   = "1C1A83"
EDS_SECONDARY = "4A4890"
EDS_ACCENT    = "B70C0D"

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color=EDS_PRIMARY, end_color=EDS_PRIMARY, fill_type="solid")
data_font = Font(name="Calibri", size=10)
money_fmt = '#,##0.00'
qty_fmt = '#,##0'
date_fmt = 'MM/DD/YYYY'
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
alt_fill = PatternFill(start_color="F2F2F7", end_color="F2F2F7", fill_type="solid")

def write_header_row(ws, row, headers):
    for col_idx, (hdr, width) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def write_data_rows(ws, start_row, df_iter, col_formats):
    """col_formats: list of (format_string or None) per column"""
    for row_idx, row in enumerate(df_iter, start_row):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            fmt = col_formats[col_idx - 1] if col_idx <= len(col_formats) else None
            if fmt:
                cell.number_format = fmt
            if row_idx % 2 == 0:
                cell.fill = alt_fill
    return start_row + len(list(df_iter))  # won't work since generator consumed

def add_total_row(ws, row, col_specs, start_data_row):
    """col_specs: list of (col_idx, fmt)"""
    ws.cell(row=row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=11, color=EDS_ACCENT)
    for col_idx, fmt in col_specs:
        cell = ws.cell(row=row, column=col_idx)
        cl = get_column_letter(col_idx)
        cell.value = f"=SUM({cl}{start_data_row}:{cl}{row-1})"
        cell.font = Font(name="Calibri", bold=True, size=11)
        cell.border = Border(top=Side(style="double"))
        cell.number_format = fmt

wb = Workbook()

# ========== Sheet 1: Combined Total by Vendor Item Code ==========
ws1 = wb.active
ws1.title = "Combined by Item"

ws1.merge_cells("A1:L1")
ws1["A1"].value = "Nasco Education (0518) — All Items Combined — Dec 2024 – Nov 2025"
ws1["A1"].font = Font(name="Calibri", bold=True, color=EDS_PRIMARY, size=14)

ws1.merge_cells("A2:L2")
ws1["A2"].value = f"Total summation across all bids  |  {len(combined)} unique items  |  ${df['LineAmount'].sum():,.2f} total spend"
ws1["A2"].font = Font(name="Calibri", italic=True, color="666666", size=10)

combined_headers = [
    ("Vendor Item Code", 18), ("Item Code", 14), ("Item Description", 45),
    ("Total Qty", 12), ("Min Price", 12), ("Max Price", 12), ("Avg Price", 12),
    ("Total Amount", 16), ("PO Count", 10), ("Bid Count", 10),
    ("Bid Header IDs", 22), ("Categories", 40),
]
write_header_row(ws1, 4, combined_headers)

col_fmts = [None, None, None, qty_fmt, money_fmt, money_fmt, money_fmt, money_fmt, qty_fmt, qty_fmt, None, None]
for row_idx, row in enumerate(combined.itertuples(index=False), 5):
    vals = [
        row.VendorItemCode, row.ItemCode, row.Description,
        row.TotalQtyOrdered, row.MinBidPrice, row.MaxBidPrice, round(row.AvgBidPrice, 4),
        row.TotalAmount, row.POCount, row.BidCount,
        row.BidHeaderIds, row.Categories,
    ]
    for col_idx, val in enumerate(vals, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        if col_fmts[col_idx - 1]:
            cell.number_format = col_fmts[col_idx - 1]
        if row_idx % 2 == 0:
            cell.fill = alt_fill

total_row = 5 + len(combined)
add_total_row(ws1, total_row, [(4, qty_fmt), (8, money_fmt), (9, qty_fmt)], 5)
ws1.freeze_panes = "A5"
ws1.auto_filter.ref = f"A4:L{total_row - 1}"

# ========== Sheet 2: Bid Summary ==========
ws2 = wb.create_sheet("Bid Summary")

ws2.merge_cells("A1:O1")
ws2["A1"].value = "Nasco Education (0518) — Bid-Level Summary — Dec 2024 – Nov 2025"
ws2["A1"].font = Font(name="Calibri", bold=True, color=EDS_PRIMARY, size=14)
ws2.merge_cells("A2:O2")
ws2["A2"].value = f"Bid-Level Summary  |  {len(bid_agg)} bids  |  ${df['LineAmount'].sum():,.2f} total spend"
ws2["A2"].font = Font(name="Calibri", italic=True, color="666666", size=10)

bid_headers = [
    ("Category", 22), ("Cat Code", 10), ("Bid ID", 8), ("Bid Name", 35),
    ("Vendor Bid #", 14), ("Bid Header ID", 14), ("Bid Header Desc", 30),
    ("Effective From", 14), ("Effective Until", 14),
    ("Unique Items", 12), ("Total Qty", 12), ("Total Amount", 16),
    ("PO Count", 10), ("First PO Date", 14), ("Last PO Date", 14),
]
write_header_row(ws2, 4, bid_headers)

bid_fmts = [None, None, None, None, None, None, None, date_fmt, date_fmt, qty_fmt, qty_fmt, money_fmt, qty_fmt, date_fmt, date_fmt]
for row_idx, row in enumerate(bid_agg.itertuples(index=False), 5):
    vals = [
        row.CategoryName, row.CategoryCode, row.BidId, row.BidName,
        row.VendorBidNumber, row.BidHeaderId, row.BidHeaderDescription,
        row.EffectiveFrom, row.EffectiveUntil,
        row.UniqueItems, row.TotalQtyOrdered, row.TotalAmount,
        row.POCount, row.FirstPODate, row.LastPODate,
    ]
    for col_idx, val in enumerate(vals, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        if bid_fmts[col_idx - 1]:
            cell.number_format = bid_fmts[col_idx - 1]
        if row_idx % 2 == 0:
            cell.fill = alt_fill

tr2 = 5 + len(bid_agg)
add_total_row(ws2, tr2, [(10, qty_fmt), (11, qty_fmt), (12, money_fmt), (13, qty_fmt)], 5)
ws2.freeze_panes = "A5"
ws2.auto_filter.ref = f"A4:O{tr2 - 1}"

# ========== Sheet 3: Item Detail by Bid ==========
ws3 = wb.create_sheet("Item Detail by Bid")

ws3.merge_cells("A1:Q1")
ws3["A1"].value = "Nasco Education (0518) — Item Detail by Bid — Dec 2024 – Nov 2025"
ws3["A1"].font = Font(name="Calibri", bold=True, color=EDS_PRIMARY, size=14)
ws3.merge_cells("A2:Q2")
ws3["A2"].value = f"Item-Level Detail  |  {len(detail_agg)} item-bid combinations"
ws3["A2"].font = Font(name="Calibri", italic=True, color="666666", size=10)

detail_headers = [
    ("Category", 22), ("Cat Code", 10), ("Bid ID", 8), ("Bid Name", 35),
    ("Vendor Bid #", 14), ("Bid Header ID", 14), ("Bid Header Desc", 30),
    ("Effective From", 14), ("Effective Until", 14),
    ("Item Code", 14), ("Vendor Item Code", 16), ("Item Description", 40),
    ("Total Qty", 12), ("Min Price", 12), ("Max Price", 12),
    ("Total Amount", 16), ("PO Count", 10),
]
write_header_row(ws3, 4, detail_headers)

det_fmts = [None, None, None, None, None, None, None, date_fmt, date_fmt, None, None, None, qty_fmt, money_fmt, money_fmt, money_fmt, qty_fmt]
for row_idx, row in enumerate(detail_agg.itertuples(index=False), 5):
    vals = [
        row.CategoryName, row.CategoryCode, row.BidId, row.BidName,
        row.VendorBidNumber, row.BidHeaderId, row.BidHeaderDescription,
        row.EffectiveFrom, row.EffectiveUntil,
        row.ItemCode, row.VendorItemCode, row.Description,
        row.TotalQtyOrdered, row.MinBidPrice, row.MaxBidPrice,
        row.TotalAmount, row.POCount,
    ]
    for col_idx, val in enumerate(vals, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        if det_fmts[col_idx - 1]:
            cell.number_format = det_fmts[col_idx - 1]
        if row_idx % 2 == 0:
            cell.fill = alt_fill

tr3 = 5 + len(detail_agg)
add_total_row(ws3, tr3, [(13, qty_fmt), (16, money_fmt), (17, qty_fmt)], 5)
ws3.freeze_panes = "A5"
ws3.auto_filter.ref = f"A4:Q{tr3 - 1}"

# ---- Save ----
out_dir = r"C:\EDS\output"
os.makedirs(out_dir, exist_ok=True)
out_path = os.path.join(out_dir, "Nasco_0518_Complete_Items_Report_Dec2024_Nov2025.xlsx")
wb.save(out_path)
print(f"\nSaved: {out_path}")
print(f"  Sheet 1 - Combined by Item: {len(combined)} unique items (sorted by Vendor Item Code)")
print(f"  Sheet 2 - Bid Summary: {len(bid_agg)} bids")
print(f"  Sheet 3 - Item Detail by Bid: {len(detail_agg)} item-bid combos")
print(f"  Total Spend: ${df['LineAmount'].sum():,.2f}")
