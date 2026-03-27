"""Generate Scarsdale 2025-2026 Requisitions Export Excel file."""
import pyodbc
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
import os

OUTPUT_PATH = r"C:\Users\conno\Desktop\Scarsdale Export 2025-2026.xlsx"

BUDGET_ID = 1685383  # Scarsdale 2025-2026

SQL = """
    SELECT
        r.RequisitionNumber AS Requisition,
        ISNULL(st.Name, 'On Hold') AS Status,
        cat.Name AS Category,
        sch.Name AS School,
        r.AccountCode AS [Account Code],
        r.Attention AS [Attention To],
        u.CometId AS [User #],
        CASE r.OrderType WHEN 1 THEN 'Annual' WHEN 2 THEN 'Incidental' ELSE 'Annual' END AS [Order Type],
        r.NotesCount AS Notes,
        r.TotalRequisitionCost AS Total
    FROM Requisitions r
    LEFT JOIN StatusTable st ON st.StatusId = r.StatusId
    LEFT JOIN Category cat ON cat.CategoryId = r.CategoryId
    LEFT JOIN School sch ON sch.SchoolId = r.SchoolId
    LEFT JOIN Users u ON u.UserId = r.UserId
    WHERE r.BudgetId = ?
    ORDER BY cat.Name, r.RequisitionNumber
"""

HEADERS = [
    "Requisition", "Status", "Category", "School",
    "Account Code", "Attention To", "User #", "Order Type",
    "Notes", "Total"
]

# Accounting format matching the sample file
ACCOUNTING_FMT = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'
TEXT_FMT = "@"

def main():
    print("Connecting to database...")
    conn = pyodbc.connect(
        'DRIVER={ODBC Driver 17 for SQL Server};'
        'SERVER=eds-sqlserver.eastus2.cloudapp.azure.com;'
        'DATABASE=EDS;UID=EDSAdmin;PWD=Consultant~!;timeout=30'
    )
    conn.autocommit = True
    cursor = conn.cursor()

    print("Querying Scarsdale 2025-2026 requisitions...")
    cursor.execute(SQL, BUDGET_ID)
    rows = cursor.fetchall()
    conn.close()
    print(f"  Fetched {len(rows)} rows.")

    print("Building Excel workbook...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # --- Write headers ---
    header_font = Font(name="Aptos Narrow", size=11, bold=True)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")

    # --- Write data rows ---
    data_font = Font(name="Aptos Narrow", size=11)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font

            if col_idx == 10:  # Total column (J) - numeric, accounting format
                # Convert Decimal to float for Excel
                if value is not None:
                    cell.value = float(value)
                else:
                    cell.value = None
                cell.number_format = ACCOUNTING_FMT
                cell.alignment = Alignment(horizontal="right")
            else:
                # Text columns A-I
                if value is None:
                    cell.value = None
                else:
                    # Store as string for text-formatted columns
                    cell.value = str(value) if not isinstance(value, str) else value
                cell.number_format = TEXT_FMT
                cell.alignment = Alignment(horizontal="left")

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    # --- Set column widths (approximate match to sample) ---
    col_widths = {
        1: 12,   # Requisition
        2: 22,   # Status
        3: 25,   # Category
        4: 30,   # School
        5: 22,   # Account Code
        6: 25,   # Attention To
        7: 10,   # User #
        8: 14,   # Order Type
        9: 8,    # Notes
        10: 14,  # Total
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    print(f"Saving to {OUTPUT_PATH} ...")
    wb.save(OUTPUT_PATH)
    print(f"Done. {len(rows)} requisitions exported.")

if __name__ == "__main__":
    main()
